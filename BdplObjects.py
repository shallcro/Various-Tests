#!/usr/bin/env python3

import bagit
import chardet
from collections import OrderedDict
from collections import Counter
import csv
import datetime
import errno
import fnmatch
import glob
import hashlib
from lxml import etree
import math
import openpyxl
import os
import pickle
import psutil
import re
import shelve
import shutil
import sqlite3
import subprocess
import sys
import tarfile
import time
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from urllib.parse import unquote
import urllib.request
import uuid
import webbrowser
import zipfile

# from dfxml project
import Objects

class Unit:
    def __init__(self, controller):
        self.controller = controller
        self.unit_name = self.controller.unit_name.get()
        self.unit_home = os.path.join(self.controller.bdpl_work_dir, self.unit_name)
        self.ingest_dir = os.path.join(self.unit_home, 'ingest')
        self.media_image_dir = os.path.join(self.controller.bdpl_work_dir, 'media-images', self.unit_name)
        self.completed_shpt_dir = os.path.join(self.unit_home, 'completed_shipments')
        
    def move_media_images(self):
    
        #make sure unit value is not empty and that 
        if self.unit_name == '':
            print('\n\nError; please make sure you have entered a unit ID abbreviation.')
            return 
                
        if len(os.listdir(self.media_image_dir)) == 0:
            print('\n\nNo images of media at {}'.format(self.media_image_dir))
            return
        
        # get a list of barcodes in each shipment
        all_barcode_folders = list(filter(lambda f: os.path.isdir(f), glob.glob('{}\\*\\*'.format(self.unit_home))))

        #list of files with no parent
        bad_file_list = []
        
        #loop through a list of all images in this folder; try to find match in list of barcodes; if not, add to 'bad file list'
        for f in os.listdir(self.media_image_dir):
            pic = f.split('-')[0]
            barcode_folder = [s for s in all_barcode_folders if pic in s]
            if len(barcode_folder) == 1:
                media_pics = os.path.join(barcode_folder[0], 'metadata', 'media-image')
                if not os.path.exists(media_pics):
                    os.makedirs(media_pics)
                try:
                    shutil.move(os.path.join(self.media_image_dir, f), media_pics)
                except shutil.Error as e:
                    print('NOTE: ', e)
                    print('\n\nCheck the media image folder to determine if a file already exists or a filename is being duplicated.')
            else:
                bad_file_list.append(f)
            
        if len(bad_file_list) > 0:
            print('\n\nFilenames for the following images do not match current barcodes:\n{}'.format('\n'.join(bad_file_list)))
            print('\nPlease correct filenames and try again.')
        else:
            print('\n\nMedia images successfully copied!')

class Shipment(Unit):
    def __init__(self, controller):
        Unit.__init__(self, controller)
        self.controller = controller
        self.shipment_date = self.controller.shipment_date.get()
        self.ship_dir = os.path.join(self.ingest_dir, self.shipment_date)
        self.spreadsheet = os.path.join(self.ship_dir, '{}_{}.xlsx'.format(self.unit_name, self.shipment_date)) 
            
    def verify_spreadsheet(self):
        #check what is in the shipment dir
        found = glob.glob(os.path.join(self.ship_dir, '*.xlsx'))

        if len(found) == 0:
            return (False, '\nWARNING: No .XLSX spreadsheet found in {}. Check {} dropbox or consult with digital preservation librarian.'.format(self.ship_dir, self.unit_name))

        elif len(found) > 1:
            if self.spreadsheet in found:
                found.remove(self.spreadsheet)
                return (True, '\nNOTE: In addition to the shipment manifest, {} contains the following spreadsheet(s):\n\n\t{}'.format(self.ship_dir, '\n\t'.join(found)))
            else:
                return (False, '\nWARNING: the following spreadsheets do not meet the BDPL naming convention of {}_{}.xlsx:\n\n\t{}'.format(self.unit_name, self.shipment_date, '\n\t'.join(found)))

        elif found[0] == self.spreadsheet:
            return (True, '\nSpreadsheet identified.')
            
        else:
            return (False, '\n\tWARNING: {} only contains the following spreadsheet: {}'.format(self.ship_dir, found[0]))

class DigitalObject(Shipment):
    def __init__(self, controller):
        Shipment.__init__(self, controller)
        self.controller = controller
        self.identifier = self.controller.identifier.get()

        '''SET UP FOLDERS'''
        #main folders
        self.barcode_dir = os.path.join(self.ship_dir, self.identifier)
        self.image_dir = os.path.join(self.barcode_dir, "disk-image")
        self.files_dir = os.path.join(self.barcode_dir, "files")
        self.metadata_dir = os.path.join(self.barcode_dir, "metadata")
        self.temp_dir = os.path.join(self.barcode_dir, "temp")
        self.reports_dir = os.path.join(self.metadata_dir, "reports")
        self.log_dir = os.path.join(self.metadata_dir, "logs")
        self.bulkext_dir = os.path.join(self.barcode_dir, "bulk_extractor")
        self.folders = [self.barcode_dir, self.image_dir, self.files_dir, self.metadata_dir, self.temp_dir, self.reports_dir, self.log_dir, self.bulkext_dir, self.media_image_dir]

        '''SET UP FILES'''
        #assets
        self.imagefile = os.path.join(self.image_dir, '{}.dd'.format(self.identifier))
        self.paranoia_out = os.path.join(self.files_dir, '{}.wav'.format(self.identifier))
        self.tar_file = os.path.join(self.ship_dir, '{}.tar'.format(self.identifier))

        #files related to disk imaging with ddrescue and FC5025
        self.mapfile = os.path.join(self.log_dir, '{}.map'.format(self.identifier))
        self.fc5025_log = os.path.join(self.log_dir, 'fcimage.log')

        #log files
        self.virus_log = os.path.join(self.log_dir, 'viruscheck-log.txt')
        self.bulkext_log = os.path.join(self.log_dir, 'bulkext-log.txt')
        self.lsdvd_out = os.path.join(self.reports_dir, "{}_lsdvd.xml".format(self.identifier))
        self.paranoia_log = os.path.join(self.log_dir, '{}-cdparanoia.log'.format(self.identifier))

        #reports
        self.disk_info_report = os.path.join(self.reports_dir, '{}-cdrdao-diskinfo.txt'.format(self.identifier))
        self.sf_file = os.path.join(self.reports_dir, 'siegfried.csv')
        self.dup_report = os.path.join(self.reports_dir, 'duplicates.csv')
        self.disktype_output = os.path.join(self.reports_dir, 'disktype.txt')
        self.fsstat_output = os.path.join(self.reports_dir, 'fsstat.txt')
        self.ils_output = os.path.join(self.reports_dir, 'ils.txt')
        self.mmls_output = os.path.join(self.reports_dir, 'mmls.txt')
        self.tree_dest = os.path.join(self.reports_dir, 'tree.txt')
        self.new_html = os.path.join(self.reports_dir, 'report.html')
        self.formatcsv = os.path.join(self.reports_dir, 'formats.csv')
        self.assets_target = os.path.join(self.reports_dir, 'assets')

        #temp files
        self.ffmpeg_temp_dir = os.path.join(self.temp_dir, 'ffmpeg')
        self.siegfried_db = os.path.join(self.temp_dir, 'siegfried.sqlite')
        self.cumulative_be_report = os.path.join(self.bulkext_dir, 'cumulative.txt')
        self.lsdvd_temp = os.path.join(self.temp_dir, 'lsdvd.txt')
        self.temp_dfxml = os.path.join(self.temp_dir, 'temp_dfxml.txt')
        self.dummy_audio = os.path.join(self.temp_dir, 'added_silence.mpg')
        self.cdr_scan = os.path.join(self.temp_dir, 'cdr_scan.txt')
        self.droid_profile = os.path.join(self.temp_dir, 'droid.droid')
        self.droid_out = os.path.join(self.temp_dir, 'droid.csv')
        self.temp_html = os.path.join(self.temp_dir, 'temp.html')
        self.assets_dir = 'C:\\BDPL\\resources\\assets'
        self.duplicates = os.path.join(self.temp_dir, 'duplicates.txt')
        self.folders_created = os.path.join(self.temp_dir, 'folders-created.txt')
        self.sqlite_done = os.path.join(self.temp_dir, 'sqlite_done.txt')
        self.stats_done = os.path.join(self.temp_dir, 'stats_done.txt')
        self.done_file = os.path.join(self.temp_dir, 'done.txt')
        self.final_stats = os.path.join(self.temp_dir, 'final_stats.txt')
        self.checksums_dvd = os.path.join(self.temp_dir, 'checksums_dvd.txt')
        self.checksums = os.path.join(self.temp_dir, 'checksums.txt')

        #metadata files
        self.dfxml_output = os.path.join(self.metadata_dir, '{}-dfxml.xml'.format(self.identifier))
        self.premis_xml_file = os.path.join(self.metadata_dir, '{}-premis.xml'.format(self.identifier))
        
        self.current_dict = self.pickle_load('dict', 'current_dict')
        
        #special vars for RipstationBatch
        if self.controller.get_current_tab() == 'RipStation Ingest':
            self.rs_wav_file = os.path.join(self.files_dir, "{}.wav".format(self.identifier))
            self.rs_wav_cue = os.path.join(self.files_dir, "{}.cue".format(self.identifier))
            self.rs_cdr_bin = os.path.join(self.image_dir, "{}-01.bin".format(self.identifier))
            self.rs_cdr_toc = os.path.join(self.image_dir, "{}-01.toc".format(self.identifier))
            self.rs_cdr_cue = os.path.join(self.image_dir, "{}-01.cue".format(self.identifier))
            self.ripstation_item_log = os.path.join(self.log_dir, 'ripstation.txt')
            self.ripstation_orig_imagefile = os.path.join(self.image_dir, '{}.iso'.format(self.identifier))
        
    def prep_barcode(self):
        
        shipment_spreadsheet = Spreadsheet(self.controller)

        #verify spreadsheet--make sure we only have 1 & that it follows naming conventions
        if self.controller.get_current_tab() == 'BDPL Ingest':
            status, msg = shipment_spreadsheet.verify_spreadsheet()
            if not status:
                return (status, msg)

        #make sure spreadsheet is not open
        if shipment_spreadsheet.already_open():
            return (False, '\n\nWARNING: {} is currently open.  Close file before continuing and/or contact digital preservation librarian if other users are involved.'.format(shipment_spreadsheet.spreadsheet))
            
        #open spreadsheet and make sure current item exists in spreadsheet; if not, return
        shipment_spreadsheet.open_wb()
        status, row = shipment_spreadsheet.return_row(shipment_spreadsheet.inv_ws)
        if not status:
            return (False, '\n\nWARNING: barcode was not found in spreadsheet.  Make sure value is entered correctly and/or check spreadsheet for value.  Consult with digital preservation librarian as needed.')
        
        #load metadata into item object
        self.load_item_metadata(shipment_spreadsheet, row)
        
        #assign variables to GUI
        if self.controller.get_current_tab() == 'BDPL Ingest':
            self.controller.content_source_type.set(self.current_dict['content_source_type'])
            self.controller.collection_title.set(self.current_dict['collection_title'])
            self.controller.collection_creator.set(self.current_dict['collection_creator'])
            self.controller.item_title.set(self.current_dict.get('item_title', '-'))
            self.controller.label_transcription.set(self.current_dict['label_transcription'])
            self.controller.item_description.set(self.current_dict.get('item_description', '-'))
            self.controller.appraisal_notes.set(self.current_dict['appraisal_notes'])
            self.controller.bdpl_instructions.set(self.current_dict['bdpl_instructions'])
        
        #create folders
        if not self.check_ingest_folders(): 
            self.create_folders() 
        
        return (True, '\n\nRecord loaded successfully; ready for next operation.')
    
    def check_barcode_status(self):
        #If a 'done' file exists, we know the whole process was completed
        done_file = os.path.join(self.temp_dir, 'done.txt')
        if os.path.exists(done_file): 
            print('\n\nNOTE: this item barcode has completed the entire BDPL Ingest workflow.  Consult with the digital preservation librarian if you believe additional procedures are needed.')
            
        #if no 'done' file, see where we are with the item...
        else:
            premis_list = self.pickle_load('ls', 'premis_list')
            if len(premis_list) > 0:
                print('\n\nIngest of item has been initiated; the following procedures have been completed:\n\t{}'.format('\n\t'.join(list(set((i['eventType'] for i in premis_list))))))
                
    def load_item_metadata(self, shipment_spreadsheet):
        
        #if dict is empty, get info from Inventory spreadsheet
        if len(self.current_dict) == 0:
            #get info from inventory sheet
            ws_columns = shipment_spreadsheet.get_spreadsheet_columns(shipment_spreadsheet.inv_ws)
            
            item_row = shipment_spreadsheet.return_row(shipment_spreadsheet.inv_ws)[1]
            
            for key in ws_columns.keys():
                if key == 'identifier':
                    self.current_dict['identifier'] = self.identifier
                else:
                    self.current_dict[key] = shipment_spreadsheet.inv_ws.cell(row=item_row, column=ws_columns[key]).value

        #now check if we need to update with any info from appraisal worksheet
        status, row = shipment_spreadsheet.return_row(shipment_spreadsheet.app_ws)        
        if status:
            ws_columns = shipment_spreadsheet.get_spreadsheet_columns(shipment_spreadsheet.app_ws)
        
            for key in ws_columns.keys():
                if key == 'identifier':
                    self.current_dict['identifier'] = self.identifier
                else:
                    self.current_dict[key] = shipment_spreadsheet.app_ws.cell(row=row, column=ws_columns[key]).value
        
        #add additional elements required for SDA deposit
        if not self.current_dict.get['unit_name']:
            self.current_dict['unit_name'] = self.unit_name
        if not self.current_dict.get['shipment_date']:
            self.current_dict['unit_name'] = self.shipment_date
        
        #clean up any None values
        for val in self.current_dict:
            if self.current_dict[val] is None:
                self.current_dict[val] = '-'
        
        #save a copy so we can access later
        self.pickle_dump('current_dict', self.current_dict)
    
    def check_ingest_folders(self):
        
        for f in self.folders:
            if not os.path.exists(f):
                return False
        
        return True
    
    def create_folders(self):
        #folders-created file will help us check for completion

        #if file doesn't exist, create folders
        for target in self.folders:
            try:
                os.makedirs(target)
            except OSError as exception:
                if exception.errno != errno.EEXIST:
                    raise
        #create file so we can check for completion later, if need be
        open(self.folders_created, 'w').close()
        
    def verify_analysis_details(self): 
    
        #make sure main variables--unit_name, shipment_date, and barcode--are included.  Return if either is missing
        status, msg = self.controller.check_main_vars()
        if not status:
            return (status, msg)
        
        #make sure we have already initiated a session for this barcode
        if not self.check_ingest_folders():
            return (False, '\n\nWARNING: load record before proceeding')
        
        if not self.controller.job_type.get() in ['Copy_only', 'Disk_image', 'CDDA', 'DVD']:
            return (False, '\nWARNING: Indicate the appropriate job type for this item and then run transfer again.')
        else:
            self.job_type = self.controller.job_type.get()
        
        self.re_analyze = self.controller.re_analyze.get()
        
        self.current_dict = self.pickle_load('ls', 'current_dict')
        
        return (True, 'Ready to analyze!')
    
    def verify_transfer_details(self):

        if self.controller.job_type.get() is None:
            return (False, '\nWARNING: Indicate the appropriate job type for this item and then run transfer again.')
            
        else:
            self.job_type = self.controller.job_type.get()
        
        #set copy_only variables
        if self.job_type == 'Copy_only':
            
            if self.controller.path_to_content.get() == '':
                return (False, '\nERROR: no path to content provided.  Be sure to click the "Browse" button and navigate to appropriate source.')
                
            if not os.path.exists(self.controller.path_to_content.get()):
                return (False, '\nWARNING: {} does not exist.  Make sure path is entered correctly and try transfer again.')

            self.path_to_content = self.controller.path_to_content.get().replace('/', '\\')
            
            #if source is in 'Z:/bdpl_transfer_list', the path_to_content is a file
            if 'bdpl_transfer_list' in self.path_to_content:
                self.path_to_content = os.path.join(self.path_to_content, '{}.txt'.format(self.identifier))
                
            return (True, 'Ready to transfer')
        
        #set other variables
        else:
            self.media_attached = self.controller.media_attached.get()
            self.source_device = self.controller.source_device.get()
            self.other_device = self.controller.other_device.get()
            self.disk_525_type = self.controller.disk_525_type.get()
            self.disk_type_options = { 'Apple DOS 3.3 (16-sector)' : 'apple33', 'Apple DOS 3.2 (13-sector)' : 'apple32', 'Apple ProDOS' : 'applepro', 'Commodore 1541' : 'c1541', 'TI-99/4A 90k' : 'ti99', 'TI-99/4A 180k' : 'ti99ds180', 'TI-99/4A 360k' : 'ti99ds360', 'Atari 810' : 'atari810', 'MS-DOS 1200k' : 'msdos12', 'MS-DOS 360k' : 'msdos360', 'North Star MDS-A-D 175k' : 'mdsad', 'North Star MDS-A-D 350k' : 'mdsad350', 'Kaypro 2 CP/M 2.2' : 'kaypro2', 'Kaypro 4 CP/M 2.2' : 'kaypro4', 'CalComp Vistagraphics 4500' : 'vg4500', 'PMC MicroMate' : 'pmc', 'Tandy Color Computer Disk BASIC' : 'coco', 'Motorola VersaDOS' : 'versa' }
        
        #make sure media has been attached
        if not self.media_attached:
            return (False, '\nWARNING: Make sure media is in drive and/or attached.  Check the "Attached?" button and launch transfer again.')
        
        #make sure we are using the optical drive for DVD and CDDA jobs
        if self.job_type in ['DVD', 'CDDA'] and self.source_device != '/dev/sr0':
            return (False, '\nWARNING: DVD and CDDA jobs must select the "CD/DVD" media source. Check settings and try transfer again.')
        else:
            self.ddrescue_target = self.source_device
            return (True, 'Ready to transfer')
        
        #we'll assign 'ddrescue_target' variable here
        if self.job_type == 'Disk_image':
            #must have a source device selected.
            if self.source_device is None:
                return (False, '\nWARNING: Indicate the appropriate source media/device for this item and then run transfer again.')
            
            #make sure that a disk type is selected if this is a 5.25" floppy    
            if self.source_device == '5.25':
                if self.disk_525_type == 'N/A':
                    return (False, '\nWARNING: Select a 5.25" disk type from the drop-down menu and try again.')
                else:
                    return (True, 'Ready to transfer')
            
            elif self.source_device in ['/dev/sr0', '/dev/fd0']:
                self.ddrescue_target = self.source_device
                return (True, 'Ready to transfer')
                
            else:
                
                #get POSIX device names from /proc/partitions
                posix_names = subprocess.check_output('cat /proc/partitions', shell=True, text=True)
                
                #get all physical drives and associated drive letters using PowerShell
                ps_cmd = "Get-Partition | % {New-Object PSObject -Property @{'DiskModel'=(Get-Disk $_.DiskNumber).Model; 'DriveLetter'=$_.DriveLetter}}"
                cmd = 'powershell.exe "{}"'.format(ps_cmd)
                drive_letters = subprocess.check_output(cmd, shell=True, text=True)
                
                #verify Zip drive device name
                if self.source_device == 'Zip':
                    for letter in drive_letters.splitlines():
                        if 'ZIP 100' in letter:
                            drive_ltr = letter.split()[2]
                            
                    #verify that Zip drive was recognized and drive letter variable was assigned
                    try:
                        drive_ltr
                    except UnboundLocalError:
                        return (False, '\nWARNING: Zip drive not recognized.  Re-insert disk into drive, allow device to complete initial loading, and attempt transfer again.')
                    
                    #match drive letter with POSIX device name
                    for line in drive_letters.splitlines():
                        if len(line.split()) == 5 and drive_ltr in line.split()[4]:
                            self.ddrescue_target = '/dev/{}'.format(line.split()[3])
                            return (True, 'Ready to transfer')
                    
                    #if unable to match drive letter and posix name, return false
                    return (False, '\nWARNING: Zip drive not recognized.  Re-insert disk into drive, allow device to complete initial loading, and attempt transfer again.')
                
                elif self.source_device == 'Other':
                    if self.other_device in posix_names:
                        self.ddrescue_target = '/dev/{}'.format(self.other_device)
                        return (True, 'Ready to transfer')
                    else:
                        return (False, '\nWARNING: device "{}" was not found in /proc/partitions; verify name, re-enter information, and attempt transfer again.'.format(self.other_device))
    
    def get_size(self, start_path):
        total_size = 0
        if os.path.isfile(start_path):
            total_size = os.path.getsize(start_path)
        else:
            for dirpath, dirnames, filenames in os.walk(start_path):
                for f in filenames:
                    fp = os.path.join(dirpath, f)
                    # skip if it is symbolic link
                    if not os.path.islink(fp):
                        total_size += os.path.getsize(fp)
        return total_size

    def secure_copy(self, content_source):

        #function takes the file source and destination as well as  a specific premis event to be used in documenting action
        print('\n\nFILE REPLICATION: TERACOPY\n\n\tSOURCE: {} \n\tDESTINATION: {}'.format(content_source, self.files_dir))
        
        #set variables for premis
        timestamp = str(datetime.datetime.now())             
        teracopy_ver = "TeraCopy v3.26"
        
        destination = self.files_dir.replace('/', '\\')
        
        #set variables for copy operation; note that if we are using a file list, TERACOPY requires a '*' before the source. 
        if os.path.isfile(content_source):
            copycmd = 'TERACOPY COPY *"{}" "{}" /SkipAll /CLOSE'.format(content_source, destination)
        else:
            copycmd = 'TERACOPY COPY "{}" "{}" /SkipAll /CLOSE'.format(content_source, destination)
        
        try:
            exitcode = subprocess.call(copycmd, shell=True, text=True)
        except subprocess.CalledProcessError as e:
            print('\n\tFile replication failed:\n\n\t{}'.format(e))
            return
                
        #need to find Teracopy SQLITE db and export list of copied files to csv log file
        list_of_files = glob.glob(os.path.join(os.path.expandvars('C:\\Users\%USERNAME%\AppData\Roaming\TeraCopy\History'), '*'))
        tera_db = max(list_of_files, key=os.path.getctime)
        
        conn = sqlite3.connect(tera_db)
        conn.text_factory = str
        cur = conn.cursor()
        results = cur.execute("SELECT * from Files")
        
        #now write the results to a csv file
        tera_log = os.path.join(self.log_dir, 'teracopy_log.csv')
        with open(tera_log, 'w', encoding='utf8') as output:
            writer = csv.writer(output, lineterminator='\n')
            header = ['Source', 'Offset', 'State', 'Size', 'Attributes', 'IsFolder', 'Creation', 'Access', 'Write', 'SourceCRC', 'TargetCRC', 'TargetName', 'Message', 'Marked', 'Hidden']
            writer.writerow(header)
            writer.writerows(results)

        cur.close()
        conn.close()    
        
        #get count of files that were actually moved
        with open(tera_log, 'r', encoding='utf8') as input:
            csvreader = csv.reader(input)
            count = sum(1 for row in csvreader) - 1
        print('\n\t{} files successfully transferred to {}.'.format(count, self.files_dir))
        
        #record premis
        self.record_premis(timestamp, 'replication', exitcode, copycmd, 'Created a copy of an object that is, bit-wise, identical to the original.', teracopy_ver)       
            
        print('\n\tFile replication completed; proceed to content analysis.')
        
    def fc5025_image(self):
    
        print('\n\n\DISK IMAGE CREATION: DeviceSideData FC5025\n\n\tSOURCE: 5.25" floppy disk \n\tDESTINATION: {}\n\n'.format(self.imagefile))       

        timestamp = str(datetime.datetime.now())
        copycmd = 'fcimage -f {} {} | tee -a {}'.format(self.disk_type_options[self.disk_525_type], self.imagefile, self.fc5025_log)
        exitcode = subprocess.call(copycmd, shell=True, text=True)
        
        #NOTE: FC5025 will return non-zero exitcode if any errors detected.  As disk image creation may still be 'successful', we will fudge the results a little bit.  Failure == no disk image.
        if exitcode != 0:
            if os.stat(imagefile).st_size > 0:
                exitcode = 0
            else:
                print('\n\nWARNING: Disk image not successfully created. Verify you have selected the correct disk type and try again (if possible).  Otherwise, indicate issues in note to collecting unit.')
                return
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'disk image creation', exitcode, copycmd, 'Extracted a disk image from the physical information carrier.', 'FCIMAGE v1309')
        
        print('\n\n\tDisk image created; proceeding to next step...')  
    
    def ddrescue_image(self):
                        
        print('\n\nDISK IMAGE CREATION: DDRESCUE\n\n\tSOURCE: {} \n\tDESTINATION: {}'.format(self.ddrescue_target, self.imagefile))
        
        dd_ver = subprocess.check_output('ddrescue -V', shell=True, text=True).split('\n', 1)[0]  
        timestamp1 = str(datetime.datetime.now())
        image_cmd1 = 'ddrescue -n {} {} {}'.format(self.ddrescue_target, self.imagefile, self.mapfile)
 
        print('\n--------------------------------------First pass with ddrescue------------------------------------\n')
        exitcode1 = subprocess.call(image_cmd1, shell=True, text=True)
        
        #record event in PREMIS metadata
        self.record_premis(timestamp1, 'disk image creation', exitcode1, image_cmd1, 'First pass; extracted a disk image from the physical information carrier.', dd_ver)
        
        #new timestamp for second pass (recommended by ddrescue developers)
        timestamp2 = str(datetime.datetime.now())
        image_cmd2 = 'ddrescue -d -r2 {} {} {}'.format(self.ddrescue_target, self.imagefile, self.mapfile)
        
        print('\n\n--------------------------------------Second pass with ddrescue------------------------------------\n')
        exitcode2 = subprocess.call(image_cmd2, shell=True, text=True)
        
        #record event in PREMIS metadata if successful
        if os.path.exists(self.imagefile) and os.stat(self.imagefile).st_size > 0:
            print('\n\n\tDisk image created; proceeding to next step...')
            exitcode2 = 0
            self.record_premis(timestamp2, 'disk image creation', exitcode2, image_cmd2, 'Second pass; extracted a disk image from the physical information carrier.', dd_ver)
        else:
            print('\n\nDISK IMAGE CREATION FAILED: Indicate any issues in note to collecting unit.')
    
    def disk_image_info(self):
        
        print('\n\nDISK IMAGE METADATA EXTRACTION: FSSTAT, ILS, MMLS')
    
        #run disktype to get information on file systems on disk
        disktype_command = 'disktype {} > {}'.format(self.imagefile, self.disktype_output)    
        timestamp = str(datetime.datetime.now())
        exitcode = subprocess.call(disktype_command, shell=True, text=True)
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'forensic feature analysis', exitcode, disktype_command, 'Determined disk image file system information.', 'disktype v9')
        
        #get disktype output; check character encoding just in case there's something funky...
        with open(self.disktype_output, 'rb') as f:
            charenc = chardet.detect(f.read())
        
        with open(self.disktype_output, 'r', encoding=charenc['encoding']) as f:
            dt_out = f.read()
        
        #print disktype output to screen
        print(dt_out, end="")
        
        #get a list of output
        dt_info = dt_out.split('Partition ')
        
        #now loop through the list to get all file systems ID'd by disktype.  Split results so we just get the name of the file system (and make lower case to avoid issues)
        fs_list = []
        for dt in dt_info:
            if 'file system' in dt:
                fs_list.append([d for d in dt.split('\n') if ' file system' in d][0].split(' file system')[0].lstrip().lower())
        
        #save file system list for later...
        self.pickle_dump('fs_list', fs_list)
        
        #run fsstat: get range of meta-data values (inode numbers) and content units (blocks or clusters)
        fsstat_ver = 'fsstat: {}'.format(subprocess.check_output('fsstat -V', shell=True, text=True).strip())
        fsstat_command = 'fsstat {} > {} 2>&1'.format(self.imagefile, self.fsstat_output)
        timestamp = str(datetime.datetime.now())
        
        try:
            exitcode = subprocess.call(fsstat_command, shell=True, text=True, timeout=60)   
        #if process times out, kill it and mark as failed
        except subprocess.TimeoutExpired:
            for proc in psutil.process_iter():
                if proc.name() == 'fsstat.exe':
                    psutil.Process(proc.pid).terminate()
            exitcode = 1
            
        #record event in PREMIS metadata    
        self.record_premis(timestamp, 'forensic feature analysis', exitcode, fsstat_command, 'Determined range of meta-data values (inode numbers) and content units (blocks or clusters)', fsstat_ver)

        #run ils to document inode information
        ils_ver = 'ils: {}'.format(subprocess.check_output('ils -V', shell=True, text=True).strip())
        ils_command = 'ils -e {} > {} 2>&1'.format(self.imagefile, self.ils_output)
        timestamp = str(datetime.datetime.now())
        try:
            exitcode = subprocess.call(ils_command, shell=True, text=True, timeout=60)
        #if the command times out, kill the process and report as a failure
        except subprocess.TimeoutExpired:
            for proc in psutil.process_iter():
                if proc.name() == 'ils.exe':
                    psutil.Process(proc.pid).terminate()
            exitcode = 1
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'forensic feature analysis', exitcode, ils_command, 'Documented all inodes found on disk image.', ils_ver)
        
        #run mmls to document the layout of partitions in a volume system
        mmls_ver = 'mmls: {}'.format(subprocess.check_output('mmls -V', shell=True, text=True).strip())
        mmls_command = 'mmls {} > {} 2>NUL'.format(self.imagefile, self.mmls_output)
        timestamp = str(datetime.datetime.now())
        exitcode = subprocess.call(mmls_command, shell=True, text=True) 
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'forensic feature analysis', exitcode, mmls_command, 'Determined the layout of partitions in a volume system.', mmls_ver)
        
        #check mmls output for partition information; first make sure there's actually data in the mmls output file
        partition_info_list = []
        
        if os.stat(self.mmls_output).st_size > 0:
            
            with open(self.mmls_output, 'r', encoding='utf8') as f:
                mmls_info = [m.split('\n') for m in f.read().splitlines()[5:]] 
            
            #loop through mmls output; match file system info (block start/end and partition#) with what came from disktype
            for mm in mmls_info:
                temp = {}
                for dt in dt_info:
                    if 'file system' in dt and ', {} sectors from {})'.format(mm[0].split()[4].lstrip('0'), mm[0].split()[2].lstrip('0')) in dt:
                        fsname = [d for d in dt.split('\n') if ' file system' in d][0].split(' file system')[0].lstrip().lower()
                        temp['start'] = mm[0].split()[2]
                        temp['desc'] = fsname
                        temp['slot'] = mm[0].split()[1]
                        #now save this dictionary to our list of partition info
                        if not temp in partition_info_list:
                            partition_info_list.append(temp)
            
            #save partition info for later
            self.pickle_dump('partition_info_list', partition_info_list)
    
    def disk_image_replication(self):    

        print('\n\nDISK IMAGE FILE REPLICATION: ')
        
        #get our software versions for unhfs and tsk_recover
        cmd = 'unhfs 2>&1'
        unhfs_tool_ver = subprocess.check_output(cmd, shell=True, text=True).splitlines()[0]
        tsk_tool_ver = 'tsk_recover: {}'.format(subprocess.check_output('tsk_recover -V', text=True).strip())
        
        #now get information on filesystems and (if present) partitions.  We will need to choose which tool to use based on file system; if UDF or ISO9660 present, use TeraCopy; otherwise use unhfs or tsk_recover
        secure_copy_list = ['udf', 'iso9660']
        unhfs_list = ['osx', 'hfs', 'apple', 'apple_hfs', 'mfs', 'hfs plus']
        tsk_list = ['ntfs', 'fat', 'fat12', 'fat16', 'fat32', 'exfat', 'ext2', 'ext3', 'ext4', 'ufs', 'ufs1', 'ufs2', 'ext', 'yaffs2', 'hfs+']
        
        #retrieve saved lists
        fs_list = self.pickle_load('ls', 'fs_list')
        partition_info_list = self.pickle_load('ls','partition_info_list')
        
        #Proceed if any file systems were found; return if none identified
        if len(fs_list) > 0:
            print('\n\tDisktype has identified the following file system(s): ', ', '.join(fs_list))
            
            #now check for any partitions; if none, go ahead and use teracopy, tsk_recover, or unhfs depending on the file system
            if len(partition_info_list) <= 1:

                print('\n\tNo partition information...')
                
                if any(fs in ' '.join(fs_list) for fs in secure_copy_list):
                    if self.controller.get_current_tab() == 'RipStation Ingest':
                        os.rename(self.imagefile, self.ripstation_orig_imagefile)
                        self.mount_iso()
                        drive_letter = self.get_iso_drive_letter()
                        self.secure_copy(drive_letter)
                        self.dismount_iso()
                        os.rename(self.ripstation_orig_imagefile, self.imagefile)
                    else:
                        self.secure_copy(self.optical_drive_letter())

                elif any(fs in ' '.join(fs_list) for fs in unhfs_list):
                    self.carve_files('unhfs', unhfs_tool_ver, '', self.files_dir)
                
                elif any(fs in ' '.join(fs_list) for fs in tsk_list): 
                    self.carve_files('tsk_recover', tsk_tool_ver, '', self.files_dir)
                
                else:
                    print('\n\tCurrent tools unable to address file system.')
                    return
                    
            #if there are one or more partitions, use tsk_recover or unhfs        
            elif len(partition_info_list) > 1:
            
                for partition in partition_info_list:

                    outfolder = os.path.join(self.files_dir, 'partition_{}'.format(partition['slot']))
                    
                    if partition['desc'] in unhfs_list:
                        self.carve_files('unhfs', unhfs_tool_ver, partition['slot'], outfolder)
                                      
                    elif partition['desc'] in tsk_list:
                        carve_files('tsk_recover', tsk_tool_ver, partition['start'], outfolder)
        else:
            print('\n\tNo files to be replicated.')
    
    def optical_drive_letter(self):
        #NOTE: this assumes only 1 optical disk drive is connected to workstation
        drive_cmd = 'wmic logicaldisk get caption, drivetype | FINDSTR /C:"5"'
        drive_ltr = subprocess.check_output(drive_cmd, shell=True, text=True).split()[0]
        return drive_ltr
    
    def carve_files(self, tool, tool_ver, partition, outfolder): 
        
        if not os.path.exists(outfolder):
            os.makedirs(outfolder)
        
        if tool == 'unhfs':
            if partition == '':
                carve_cmd = 'unhfs -sfm-substitutions -resforks APPLEDOUBLE -o "{}" "{}" 2>nul'.format(outfolder, imagefile)
            else:
                carve_cmd = 'unhfs -sfm-substitutions -partition {} -resforks APPLEDOUBLE -o "{}" "{}" 2>nul'.format(partition, outfolder, imagefile)
        
        else:
            if partition == '':
                carve_cmd = 'tsk_recover -a {} {}'.format(imagefile, outfolder)
            else:
                carve_cmd = 'tsk_recover -a -o {} {} {}'.format(partition, imagefile, outfolder)
            
        print('\n\tTOOL: {}\n\n\tSOURCE: {} \n\n\tDESTINATION: {}\n'.format(tool, imagefile, outfolder))
        
        timestamp = str(datetime.datetime.now())  
        exitcode = subprocess.call(carve_cmd, shell=True)
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'replication', exitcode, carve_cmd, "Created a copy of an object that is, bit-wise, identical to the original.", tool_ver)
        
        #if no files were extracted, remove partition folder.
        if not self.check_files(outfolder) and outfolder != self.files_dir:
            os.rmdir(outfolder)
        
        #if tsk_recover has been run, go through and fix the file MAC times
        if tool == 'tsk_recover' and exitcode == 0:
            
            #generate DFXML with fiwalk
            if not os.path.exists(self.dfxml_output):
                self.produce_dfxml(self.imagefile)
            
            #use DFXML output to get correct MAC times and update files
            self.fix_dates(outfolder)
        
        elif tool == 'unhfs' and os.path.exists(outfolder):
            file_count = sum([len(files) for r, d, files in os.walk(outfolder)])
            print('\t{} files successfully transferred to {}.'.format(file_count, outfolder))
            
        print('\n\tFile replication completed; proceed to content analysis.')
    
    def produce_dfxml(self, target):
    
        timestamp = str(datetime.datetime.now())
        file_stats = []
        
        #use fiwalk if we have an image file
        if os.path.isfile(target):
            print('\n\nDIGITAL FORENSICS XML CREATION: FIWALK')
            dfxml_ver_cmd = 'fiwalk-0.6.3 -V'
            dfxml_ver = subprocess.check_output(dfxml_ver_cmd, shell=True, text=True).splitlines()[0]
            dfxml_cmd = 'fiwalk-0.6.3 -x {} > {}'.format(target, self.dfxml_output)
            exitcode = subprocess.call(dfxml_cmd, shell=True, text=True)
                    
            #parse dfxml to get info for later; because large DFXML files pose a challenge; use iterparse to avoid crashing (Note: for DVD jobs we will also get stats on the files themselves later on) 
            print('\n\tCollecting file statistics...\n')
            counter = 0
            for event, element in etree.iterparse(self.dfxml_output, events = ("end",), tag="fileobject"):
                
                #refresh dict for each fileobject
                file_dict = {}
                
                #default values; will make sure that we don't record info about non-allocated files and that we have a default timestamp value
                good = True
                mt = False
                mtime = 'undated'
                target = ''
                size = ''
                checksum = ''
                
                for child in element:
                    
                    if child.tag == "filename":
                        target = child.text
                    if child.tag == "name_type":
                        if child.text != "r":
                            element.clear()
                            good = False
                            break
                    if child.tag == "alloc":
                        if child.text != "1":
                            good = False
                            element.clear()
                            break
                    if child.tag == "unalloc":
                        if child.text == "1":
                            good = False
                            element.clear()
                            break
                    if child.tag == "filesize":
                        size = child.text
                    if child.tag == "hashdigest":
                        if child.attrib['type'] == 'md5':
                            checksum = child.text
                    if child.tag == "mtime":
                        mtime = datetime.datetime.utcfromtimestamp(int(child.text)).isoformat()
                        mt = True
                    if child.tag == "crtime" and mt == False:
                        mtime = datetime.datetime.utcfromtimestamp(int(child.text)).isoformat()
                
                if good and not '' in file_dict.values():
                    file_dict = { 'name' : target, 'size' : size, 'mtime' : mtime, 'checksum' : checksum}
                    file_stats.append(file_dict)
                    
                    counter+=1            
                    print('\r\tWorking on file #: {}'.format(counter), end='')

                element.clear()
                
            if self.job_type == 'DVD':
            
                #save info from DVD checksums to separate file
                with open (self.checksums_dvd, 'wb') as f:
                    pickle.dump(file_stats, f)
                
                #now compile stats for the normalized file versions
                file_stats = []
                for f in os.listdir(self.files_dir):
                    file = os.path.join(self.files_dir, f)
                    file_dict = {}
                    size = os.path.getsize(file)
                    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file)).isoformat()
                    ctime = datetime.datetime.fromtimestamp(os.path.getctime(file)).isoformat()
                    atime = datetime.datetime.fromtimestamp(os.path.getatime(file)).isoformat()[:-7]
                    checksum = self.md5(file)
                    
                    file_dict = { 'name' : file, 'size' : size, 'mtime' : mtime, 'ctime' : ctime, 'atime' : atime, 'checksum' : checksum}
                    file_stats.append(file_dict)  
     
        #use custom operation for other cases    
        elif os.path.isdir(target):
            print('\n\nDIGITAL FORENSICS XML CREATION: bdpl_ingest')
            
            dfxml_ver = 'https://github.com/IUBLibTech/bdpl_ingest'
            dfxml_cmd = 'bdpl_ingest.py'
            
            timestamp = str(datetime.datetime.now().isoformat())
            
            done_list = []

            if os.path.exists(self.temp_dfxml):
                with open(self.temp_dfxml, 'r', encoding='utf-8') as f:
                    done_so_far = f.read().splitlines()
                    for d in done_so_far:
                        line = d.split(' | ')
                        done_list.append(line[0])
                        file_dict = { 'name' : line[0], 'size' : line[1], 'mtime' : line[2], 'ctime' : line[3], 'atime' : line[4], 'checksum' : line[5], 'counter' : line[6] }
                        file_stats.append(file_dict)
            
            if len(file_stats) > 0:
                counter = int(file_stats[-1]['counter'])
            else:
                counter = 0
            
            print('\n')
            
            #get total number of files
            total = sum([len(files) for r, d, files in os.walk(target)])
            
            #now loop through, keeping count
            for root, dirnames, filenames in os.walk(target):
                for file in filenames:
                    
                    #check to make sure that we haven't already added info for this file
                    file_target = os.path.join(root, file)
                    
                    if file_target in done_list:
                        continue
                    
                    counter += 1
                    print('\r\tCalculating checksum for file {} out of {}'.format(counter, total), end='')
                    
                    size = os.path.getsize(file_target)
                    mtime = datetime.datetime.fromtimestamp(os.path.getmtime(file_target)).isoformat()
                    ctime = datetime.datetime.fromtimestamp(os.path.getctime(file_target)).isoformat()
                    atime = datetime.datetime.fromtimestamp(os.path.getatime(file_target)).isoformat()[:-7]
                    checksum = self.md5(file_target)
                    file_dict = { 'name' : file_target, 'size' : size, 'mtime' : mtime, 'ctime' : ctime, 'atime' : atime, 'checksum' : checksum, 'counter' : counter }                 
                    file_stats.append(file_dict)
                    done_list.append(file_target)
                    
                    #save this list to file just in case we crash...
                    raw_stats = "{} | {} | {} | {} | {} | {} | {}\n".format(file_target, size, mtime, ctime, atime, checksum, counter)
                    with open(self.temp_dfxml, 'a', encoding='utf8') as f:
                        f.write(raw_stats)
            
            print('\n')
            
            dc_namespace = 'http://purl.org/dc/elements/1.1/'
            dc = "{%s}" % dc_namespace
            NSMAP = {None : 'http://www.forensicswiki.org/wiki/Category:Digital_Forensics_XML',
                    'xsi': "http://www.w3.org/2001/XMLSchema-instance",
                    'dc' : dc_namespace}
            dfxml = etree.Element("dfxml", nsmap=NSMAP, version="1.0")
            metadata = etree.SubElement(dfxml, "metadata")
            dctype = etree.SubElement(metadata, dc + "type")
            dctype.text = "Hash List"
            creator = etree.SubElement(dfxml, 'creator')
            program = etree.SubElement(creator, 'program')
            program.text = 'bdpl_ingest'
            execution_environment = etree.SubElement(creator, 'execution_environment')
            start_time = etree.SubElement(execution_environment, 'start_time')
            start_time.text = timestamp
            
            for f in file_stats:
                fileobject = etree.SubElement(dfxml, 'fileobject')
                filename = etree.SubElement(fileobject, 'filename')
                filename.text = f['name']
                filesize = etree.SubElement(fileobject, 'filesize')
                filesize.text = str(f['size'])
                modifiedtime = etree.SubElement(fileobject, 'mtime')
                modifiedtime.text = f['mtime']
                createdtime = etree.SubElement(fileobject, 'ctime')
                createdtime.text = f['ctime']
                accesstime = etree.SubElement(fileobject, 'atime')
                accesstime.text = f['atime']
                hashdigest = etree.SubElement(fileobject, 'hashdigest', type='md5')
                hashdigest.text = f['checksum']

            tree = etree.ElementTree(dfxml)
            tree.write(self.dfxml_output, pretty_print=True, xml_declaration=True, encoding="utf-8")      
        
        else:
            print('\n\tERROR: {} does not appear to exist...'.format(target))
            return
        
        #save stats for reporting...            
        with open (self.checksums, 'wb') as f:
            pickle.dump(file_stats, f)
        
        #save PREMIS
        self.record_premis(timestamp, 'message digest calculation', 0, dfxml_cmd, 'Extracted information about the structure and characteristics of content, including file checksums.', dfxml_ver)
        
        print('\n\n\tDFXML creation completed; moving on to next step...')
    
    def fix_dates(self, outfolder):
        #adapted from Timothy Walsh's Disk Image Processor: https://github.com/CCA-Public/diskimageprocessor
               
        print('\n\nFILE MAC TIME CORRECTION (USING DFXML)')
        
        timestamp = str(datetime.datetime.now())
         
        try:
            for (event, obj) in Objects.iterparse(self.dfxml_output):
                # only work on FileObjects
                if not isinstance(obj, Objects.FileObject):
                    continue

                # skip directories and links
                if obj.name_type:
                    if obj.name_type not in ["r", "d"]:
                        continue

                # record filename
                dfxml_filename = obj.filename
                dfxml_filedate = int(time.time()) # default to current time

                # record last modified or last created date
                try:
                    mtime = obj.mtime
                    mtime = str(mtime)
                except:
                    pass

                try:
                    crtime = obj.crtime
                    crtime = str(crtime)
                except:
                    pass

                # fallback to created date if last modified doesn't exist
                if mtime and (mtime != 'None'):
                    mtime = time_to_int(mtime[:19])
                    dfxml_filedate = mtime
                elif crtime and (crtime != 'None'):
                    crtime = time_to_int(crtime[:19])
                    dfxml_filedate = crtime
                else:
                    continue

                # rewrite last modified date of corresponding file in objects/files
                exported_filepath = os.path.join(outfolder, dfxml_filename)
                if os.path.isdir(exported_filepath):
                    os.utime(exported_filepath, (dfxml_filedate, dfxml_filedate))
                elif os.path.isfile(exported_filepath):
                    os.utime(exported_filepath, (dfxml_filedate, dfxml_filedate)) 
                else:
                    continue

        except (ValueError, xml.etree.ElementTree.ParseError):
            print('\nUnable to read DFXML!')
            pass
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'metadata modification', 0, 'https://github.com/CCA-Public/diskimageprocessor/blob/master/diskimageprocessor.py#L446-L489', 'Corrected file timestamps to match information extracted from disk image.', 'Adapted from Disk Image Processor Version: 1.0.0 (Tim Walsh)')
    
    def time_to_int(self, str_time):
        """ Convert datetime to unix integer value """
        dt = time.mktime(datetime.datetime.strptime(str_time, 
            "%Y-%m-%dT%H:%M:%S").timetuple())
        return dt
    
    def lsdvd_check(self, drive_letter):
        
        #get lsdvd version
        lsdvd_ver = subprocess.run('lsdvd -V', shell=True, text=True, capture_output=True).stderr.split(' - ')[0]
        
        #now run lsdvd to get info about DVD, including # of titles
        timestamp = str(datetime.datetime.now())
        lsdvdcmd = 'lsdvd -Ox -x {} > {} 2> NUL'.format(drive_letter, self.lsdvd_out)
        exitcode = subprocess.call(lsdvdcmd, shell=True, text=True)
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'metadata extraction', exitcode, lsdvdcmd, 'Extracted content information from DVD, including titles, chapters, audio streams and video.', lsdvd_ver)
        
        #now verify how many titles are on the disk.  Set a default value of 0
        titlecount = 0
        
        #check file to see how many titles are on DVD using lsdvd XML output. 
        parser = etree.XMLParser(recover=True)

        try:
            doc = etree.parse(self.lsdvd_out, parser=parser)
            titlecount = int(doc.xpath("count(//lsdvd//track)"))
            
            #check for PAL content, just in case...
            formats = doc.xpath("//format")
            if [f for f in formats if f.text == 'PAL']:
                title_format = 'PAL'
            else:
                title_format = 'NTSC'
                
        #if lsdvd fails or information not in report, get the title count by parsing directory...
        except (OSError, lxml.etree.XMLSyntaxError):
            titlelist = glob.glob(os.path.join(drive_letter, '**/VIDEO_TS', '*_*_*.VOB'), recursive=True)
            count = []
            for t in titlelist:
                #parse VOB filenames to get # of titles
                count.append(int(t.rsplit('_', 2)[1]))
            if len(count) > 0:
                titlecount = max(set(count))
        
        #if we haven't identified titles (i.e., we do not have a DVD), delete lsdvd output
        if titlecount == 0:
            os.remove(self.lsdvd_out)
            
        return titlecount, title_format
    
    def normalize_dvd_content(self, titlecount, drive_letter):

        #check current directory; change to a temp directory to store files
        bdpl_cwd = 'C:\\BDPL\\scripts'
        
        if not os.path.exists(self.ffmpeg_temp_dir):
            os.makedirs(self.ffmpeg_temp_dir)
        
        os.chdir(self.ffmpeg_temp_dir)
        
        #get ffmpeg version
        ffmpeg_ver =  '; '.join(subprocess.check_output('"C:\\Program Files\\ffmpeg\\bin\\ffmpeg" -version', shell=True, text=True).splitlines()[0:2])
        
        print('\n\nMOVING IMAGE FILE NORMALIZATION: FFMPEG')
        
        #loop through titles and rip each one to mpeg using native streams
        for title in range(1, (titlecount+1)):
            titlelist = glob.glob(os.path.join(drive_letter, "**/VIDEO_TS", "VTS_{}_*.VOB".format(str(title).zfill(2))), recursive=True)
            #be sure list is sorted
            sorted(titlelist)
            
            if len(titlelist) > 0:
                
                #check if title track is missing audio--could make trouble for other tracks...
                audio_test = {}
                print('\n\tChecking audio streams...')
                for t in titlelist:
                    cmd = "ffprobe -i {} -hide_banner -show_streams -select_streams a -loglevel error".format(t)
                    try:
                        audio_check = subprocess.check_output(cmd, shell=True, text=True)
                        audio_test[t] = audio_check
                    except subprocess.CalledProcessError:
                        pass
                
                if len(audio_test) == 0:
                    print('\nWARNING: unable to access information on DVD. Moving image normalization has failed...')
                    return
                
                #if there's no audio in any track, it's OK
                if all(value == '' for value in audio_test.values()):
                    pass
                    
                #if our first track lacks audio, add a dummy track
                elif audio_test[titlelist[0]] == '':
                    
                    cmd = "ffmpeg -y -nostdin -loglevel warning -i {} -f lavfi -i anullsrc -c:v copy -c:a aac -shortest -target ntsc-dvd {{}".format(titlelist[0], self.dummy_audio)
                    
                    print('\n\tCorrecting missing audio on first track...')
                    
                    subprocess.call(cmd, text=True)
                    
                    #replace original item from list
                    del titlelist[0]
                    titlelist.insert(0, dummy_audio)
                
                timestamp = str(datetime.datetime.now())
                
                ffmpegout = os.path.join(self.files_dir, '{}-{}.mpg'.format(self.identifier, str(title).zfill(2)))
                ffmpeg_cmd = 'ffmpeg -y -nostdin -loglevel warning -report -stats -i "concat:{}" -c copy -target ntsc-dvd {}'.format('|'.join(titlelist), ffmpegout)
                
                print('\n\tGenerating title {} of {}: {}\n'.format(str(title), str(titlecount), ffmpegout))
                
                exitcode = subprocess.call(ffmpeg_cmd, shell=True, text=True)
                
                #record event in PREMIS metadata                
                self.record_premis(timestamp, 'normalization', exitcode, ffmpeg_cmd, 'Transformed object to an institutionally supported preservation format (.MPG) with a direct copy of all streams.', ffmpeg_ver)
                
                #move and rename ffmpeg log file
                ffmpeglog = glob.glob(os.path.join(self.ffmpeg_temp_dir, 'ffmpeg-*.log'))[0]
                shutil.move(ffmpeglog, os.path.join(self.log_dir, '{}-{}-ffmpeg.log'.format(identifier, str(title).zfill(2))))
                
        #move back to original directory
        os.chdir(bdpl_cwd)
        
        print('\n\tMoving image normalization completed; proceed to content analysis.')

    def cdda_image_creation(self):
        
        print('\n\nDISK IMAGE CREATION: CDRDAO\n\n\tSOURCE: {} \n\tDESTINATION: {}'.format(self.source_device, self.image_dir))
        
        #determine appropriate drive ID for cdrdao; save output of command to temp file
        scan_cmd = 'cdrdao scanbus > {} 2>&1'.format(self.cdr_scan)
        
        subprocess.check_output(scan_cmd, shell=True, text=True)

        #pull drive ID and cdrdao version from file
        with open(self.cdr_scan, 'r') as f:
            info = f.read().splitlines()
        cdrdao_ver = info[0].split(' - ')[0]
        drive_id = info[8].split(':')[0]
            
        #get info about CD using cdrdao; record this as a premis event, too.
        cdrdao_cmd = 'cdrdao disk-info --device {} --driver generic-mmc-raw > {} 2>&1'.format(drive_id, self.disk_info_report)
        timestamp = str(datetime.datetime.now())
        exitcode = subprocess.call(cdrdao_cmd, shell=True, text=True)
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'metadata extraction', exitcode, cdrdao_cmd, 'Extracted information about the CD-R, including medium, TOC type, number of sessions, etc.', cdrdao_ver)

        #read log file to determine # of sessions on disk.
        with open(self.disk_info_report, 'r') as f:
            for line in f:
                if 'Sessions             :' in line:
                    sessions = int(line.split(':')[1].strip())
        
        t2c_ver = subprocess.check_output('toc2cue -V', shell=True, text=True).strip()
        
        #for each session, create a bin/toc file
        for x in range(1, (sessions+1)):
            cdr_bin = os.path.join(self.image_dir, "{}-{}.bin").format(self.identifier, str(x).zfill(2))
            cdr_toc = os.path.join(self.image_dir, "{}-{}.toc").format(self.identifier, str(x).zfill(2))
            cdr_log = os.path.join(self.image_dir, "{}-{}.log").format(self.identifier, str(x).zfill(2))
            
            print('\n\tGenerating session {} of {}: {}\n\n'.format(str(x), str(sessions), cdr_bin))
            
            #create separate bin/cue for each session
            cdr_cmd = 'cdrdao read-cd --read-raw --session {} --datafile {} --device {} --driver generic-mmc-raw -v 3 {} | tee -a {}'.format(str(x), cdr_bin, drive_id, cdr_toc, cdr_log)
            
            timestamp = str(datetime.datetime.now())
            
            #record event in PREMIS metadata
            exitcode = subprocess.call(cdr_cmd, shell=True, text=True)
            
            self.record_premis(timestamp, 'disk image creation', exitcode, cdr_cmd, 'Extracted a disk image from the physical information carrier.', cdrdao_ver)
                        
            #convert TOC to CUE
            cue = os.path.join(self.image_dir, "{}-{}.cue").format(self.identifier, str(sessions).zfill(2))
            cue_log = os.path.join(self.log_dir, "{}-{}_toc2cue.log").format(self.identifier, str(sessions).zfill(2))
            t2c_cmd = 'toc2cue {} {} > {} 2>&1'.format(cdr_toc, cue, cue_log)
            timestamp = str(datetime.datetime.now())
            exitcode2 = subprocess.call(t2c_cmd, shell=True, text=True)
            
            #toc2cue may try to encode path information as binary data--let's fix that
            with open(cue, 'rb') as infile:
                cue_info = infile.readlines()[1:]
            
            with open(cue, 'w') as outfile:
                outfile.write('FILE "{}" BINARY\n'.format(os.path.basename(cdr_bin)))
            
            with open(cue, 'ab') as outfile:
                for line in cue_info:
                    outfile.write(line)           
            
            #record event in PREMIS metadata
            self.record_premis(timestamp, 'metadata modification', exitcode2, t2c_cmd, "Converted the CD's table of contents (TOC) file to the CUE format.", t2c_ver)
            
            #place a copy of the .cue file for the first session in files_dir for the forthcoming WAV; this session will have audio data
            if x == 1:
                new_cue = os.path.join(self.files_dir, '{}.cue'.format(self.identifier))
                
                #now write the new cue file
                with open(new_cue, 'w') as outfile:
                    outfile.write('FILE "{}.wav" WAVE\n'.format(self.identifier))
                    
                with open(new_cue, 'ab') as outfile:
                    for line in cue_info:
                        outfile.write(line)
        
        print('\n\tCDDA disk image created; moving on to next step...')

    def cdda_wav_creation(self):

        #get cdparanoia version
        ver_cmd = 'cd-paranoia -V'    
        paranoia_ver = subprocess.run(ver_cmd, shell=True, text=True, capture_output=True).stderr.splitlines()[0]
        
        print('\n\nAUDIO CONTENT NORMALIZATION: CDPARANOIA\n\n\tSOURCE: {} \n\tDESTINATION: {}\n'.format(self.source_device, self.paranoia_out))
        
        paranoia_cmd = 'cd-paranoia -l {} -w [00:00:00.00]- {}'.format(self.paranoia_log, self.paranoia_out)
        
        timestamp = str(datetime.datetime.now())
        exitcode = subprocess.call(paranoia_cmd, shell=True, text=True)
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'normalization', exitcode, paranoia_cmd, 'Transformed object to an institutionally supported preservation format (.WAV).', paranoia_ver)
        
        print('\n\tAudio normalization complete; proceed to content analysis.')
    
    def run_antivirus(self):
       
        #get version
        cmd = 'clamscan -V'
        av_ver = subprocess.check_output(cmd, text=True).rstrip()

        av_command = 'clamscan -i -l {} --recursive {}'.format(self.virus_log, self.files_dir)  
        
        timestamp = str(datetime.datetime.now())
        exitcode = subprocess.call(av_command, shell=True, text=True)
        
        #store virus scan results in current_dict
        with open(self.virus_log, 'r') as f:
            if "Infected files: 0" not in f.read():
                self.current_dict['virus_scan_results'] = 'WARNING! Virus or malware found; see {}.'.format(self.virus_log)
            else:
                self.current_dict['virus_scan_results'] = '-'

        #save current_dict to file, just in case
        self.pickle_dump('current_dict', self.current_dict)
        
        #save preservation metadata to PREMIS
        self.record_premis(timestamp, 'virus check', exitcode, av_command, 'Scanned files for malicious programs.', av_ver)
        
        print('\n\tVirus scan completed; moving on to next step...')

    def document_dir_tree(self):
        
        #make a directory tree to document original structure
        tree_dest = os.path.join(self.reports_dir, 'tree.txt')
        
        tree_ver = subprocess.check_output('tree --version', shell=True, text=True).split(' (')[0]
        tree_command = 'tree.exe -tDhR "{}" > "{}"'.format(self.files_dir, tree_dest)
        
        timestamp = str(datetime.datetime.now())
        exitcode = subprocess.call(tree_command, shell=True, text=True)
        
        self.record_premis(timestamp, 'metadata extraction', exitcode, tree_command, 'Documented the organization and structure of content within a directory tree.', tree_ver)
        
        print('\n\tDirectory structure documented; moving on to next step...')
    
    def run_bulkext(self):

        #get bulk extractor version for premis
        try:
            be_ver = subprocess.check_output(['bulk_extractor', '-V'], shell=True, text=True).rstrip()
        except subprocess.CalledProcessError as e:
            be_ver = e.output.rstrip()
        
        print('\n\tScan underway...be patient!\n')
        
        #use default command with buklk_extractor
        bulkext_command = 'bulk_extractor -x aes -x base64 -x elf -x exif -x gps -x hiberfile -x httplogs -x json -x kml -x net -x pdf -x sqlite -x winlnk -x winpe -x winprefetch -S ssn_mode=2 -q -1 -o "{}" -R "{}" > "{}"'.format(self.bulkext_dir, self.files_dir, self.bulkext_log)
        
        if os.path.exists(self.bulkext_dir):
            shutil.rmtree(self.bulkext_dir)
        
        try:
            os.makedirs(self.bulkext_dir)
        except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise

        #create timestamp
        timestamp = str(datetime.datetime.now())        

        exitcode = subprocess.call(bulkext_command, shell=True, text=True)
       
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'sensitive data scan', exitcode, bulkext_command, 'Scanned files for potentially sensitive information, including Social Security and credit card numbers.', be_ver)
        
        #create a cumulative BE report
        if os.path.exists(self.cumulative_be_report):
            os.remove(self.cumulative_be_report)
            
        for myfile in ('pii.txt', 'ccn.txt', 'email.txt', 'telephone.txt', 'find.txt'):
            myfile = os.path.join(self.bulkext_dir, myfile)
            if os.path.exists(myfile) and os.stat(myfile).st_size > 0:
                with open(myfile, 'rb') as filein:
                    data = filein.read().splitlines()    
                with open(self.cumulative_be_report, 'a', encoding='utf8') as outfile:
                    outfile.write('{}: {}\n'.format(os.path.basename(myfile), len(data[5:])))
        
        #if no results from the above, create file so we don't throw an error later
        if not os.path.exists(self.cumulative_be_report):         
            open(self.cumulative_be_report, 'a').close()
        #otherwise, move any b_e histogram files, if needed
        else:
            for myfile in ('email_domain_histogram.txt', 'find_histogram.txt', 'telephone_histogram.txt'):
                current_file = os.path.join(self.bulkext_dir, myfile)
                try:    
                    if os.stat(current_file).st_size > 0:
                        shutil.copy(current_file, self.reports_dir)
                except OSError:
                    continue
        
        print('\n\tSensitive data scan completed; moving on to next step...')
    
    def format_analysis(self):
    
        print('\n\tFile format identification with Siegfried...') 

        format_version = subprocess.check_output('sf -version', shell=True, text=True).replace('\n', ' ')
        
        #remove Siegrfried report if it already exists
        if os.path.exists(self.sf_file):
            os.remove(self.sf_file)                                                                 
                
        format_command = 'sf -z -csv "{}" > "{}"'.format(self.files_dir, self.sf_file)
        
        #create timestamp
        timestamp = str(datetime.datetime.now())
        
        exitcode = subprocess.call(format_command, shell=True, text=True)
        
        #if siegfried fails, then we'll run DROID
        if exitcode != 0 and os.path.getsize(sf_file) == 0:
            print('\n\tFile format identification with siegfried failed; now attempting with DROID...\n') 
            
            format_version = "DROID v{}".format(subprocess.check_output('droid -v', shell=True, text=True).strip())
            
            droid_cmd1 = 'droid -RAq -a "{}" -p "{}"'.format(self.files_dir, self.droid_profile)
            
            exitcode = subprocess.call(droid_cmd1, shell=True)
            
            droid_cmd2 = 'droid -p "{}" -e "{}"'.format(self.droid_profile, self.droid_out)
            
            subprocess.call(droid_cmd2, shell=True)
            
            #consolidate commands for premis
            format_command = "{} && {}".format(droid_cmd1, droid_cmd2)
            
            #now reformat droid output to be like sf output
            self.droid_to_siegfried()
        
        #record event in PREMIS metadata
        self.record_premis(timestamp, 'format identification', exitcode, format_command, 'Determined file format and version numbers for content using the PRONOM format registry.', format_version)
    
    def droid_to_siegfried(self):

        counter = 0

        with open(self.sf_file, 'w', newline='') as f1:
            csvWriter = csv.writer(f1)
            header = ['filename', 'filesize', 'modified', 'errors', 'namespace', 'id', 'format', 'version', 'mime', 'basis', 'warning']
            csvWriter.writerow(header)
            with open(self.droid_out, 'r', encoding='utf8') as f2:
                csvReader = csv.reader(f2)
                next(csvReader)
                for row in csvReader:
                    counter+=1
                    print('\rWorking on row {}'.format(counter), end='')
                    
                    if 'zip:file:' in row[2]:
                        filename = row[2].split('zip:file:/', 1)[1].replace('.zip!', '.zip#').replace('/', '\\')
                    else:
                        filename = row[2].split('file:/', 1)[1]
                    filename = unquote(filename)
                    
                    filesize = row[7]
                    modified = row[10]
                    errors = ''
                    namespace = 'pronom'
                    if row[14] == "":
                        id = 'UNKNOWN'
                    else:
                        id = row[14]
                    format = row[16]
                    version = row[17]
                    mime = row[15]
                    basis = ''
                    if row[11].lower() == 'true':
                        warning = 'extension mismatch'
                    else:
                        warning = ''
                    
                    data = [filename, filesize, modified, errors, namespace, id, format, version, mime, basis, warning]
                    
                    csvWriter.writerow(data)
    
    def import_csv(self):

        conn = sqlite3.connect(self.siegfried_db)
        conn.text_factory = str  # allows utf-8 data to be stored
        cursor = conn.cursor()

        print('\n\tImporting siegried file to sqlite3 database...')
        
        """Import csv file into sqlite db"""
        f = open(self.sf_file, 'r', encoding='utf8')
        
        try:
            reader = csv.reader(x.replace('\0', '') for x in f) # replace null bytes with empty strings on read
        except UnicodeDecodeError:
            f = (x.strip() for x in f) # skip non-utf8 encodable characters
            reader = csv.reader(x.replace('\0', '') for x in f) # replace null bytes with empty strings on read
        header = True
        for row in reader:
            if header:
                header = False # gather column names from first row of csv
                sql = "DROP TABLE IF EXISTS siegfried"
                cursor.execute(sql)
                sql = "CREATE TABLE siegfried (filename text, filesize text, modified text, errors text, namespace text, id text, format text, version text, mime text, basis text, warning text)"
                cursor.execute(sql)
                insertsql = "INSERT INTO siegfried VALUES ({})".format(", ".join([ "?" for column in row ]))
                rowlen = len(row)
            else:
                # skip lines that don't have right number of columns
                if len(row) == rowlen:
                    cursor.execute(insertsql, row)
        conn.commit()
        f.close()
        
        #create file to indicate that this operation has completed
        open(self.sqlite_done, 'a').close()
        
        cursor.close()
        conn.close()
    
    def sqlite_to_csv(self, sql, path, header, cursor):
        """Write sql query result to csv"""
        report = open(path, 'w', newline='', encoding='utf8')

        w = csv.writer(report, lineterminator='\n')
        w.writerow(header)
        for row in cursor.execute(sql):
            w.writerow(row)
        report.close()
    
    def get_stats(self):

        print('\n\tGetting statistics and generating reports about content...')
        
        #prepare sqlite database and variables
        conn = sqlite3.connect(self.siegfried_db)
        conn.text_factory = str  # allows utf-8 data to be stored
        cursor = conn.cursor()
        
        full_header = ['Filename', 'Filesize', 'Date modified', 'Errors', 
                    'Namespace', 'ID', 'Format', 'Format version', 'MIME type', 
                    'Basis for ID', 'Warning']
        
        #retrieve our 'file stats'
        file_stats = []
        try:
            with open(self.checksums, 'rb') as f:
                file_stats = pickle.load(f)
        except FileNotFoundError:
            pass
        
        # get total # of files
        cursor.execute("SELECT COUNT(*) from siegfried;") # total files
        self.num_files = cursor.fetchone()[0]

        # get # of empty files
        cursor.execute("SELECT COUNT(*) from siegfried where filesize='0';") # empty files
        self.empty_files = cursor.fetchone()[0]
            
        #Get stats on duplicates. Just in case the bdpl ingest tool crashes after compiling a duplicates list, we'll check to see if it already exists
        dup_list = []
        if os.path.exists(self.duplicates) and not self.re_analyze:
            dup_list = self.pickle_load('ls', 'duplicates')
        else:
            #next, create a new dictionary that IDs checksums that correspond to 1 or more files. NOTE: the 'file_stats' list will be empty for DVDs, so we'll skip this step in that case
            if len(file_stats) > 1:
                stat_dict = {}
                for dctnry in file_stats:
                    if int(dctnry['size']) > 0:
                        if dctnry['checksum'] in stat_dict:
                            stat_dict[dctnry['checksum']].append(dctnry['name'])
                        else:
                            stat_dict[dctnry['checksum']] = [dctnry['name']]
               
                #go through new dict and find checksums with duplicates
                for chksm in [key for key, values in stat_dict.items() if len(values) > 1]:
                    for fname in stat_dict[chksm]:
                        temp = [item for item in file_stats if item['checksum'] == chksm and item['name'] == fname][0]
                        dup_list.append([temp['name'], temp['size'], temp['mtime'], temp['checksum']])
                
            #save this duplicate file for later when we need to write to html
            self.pickle_dump('duplicates', dup_list)
        
        #total duplicates = total length of duplicate list
        self.all_dupes = len(dup_list)

        #distinct duplicates = # of unique checksums in the duplicates list
        self.distinct_dupes = len(set([c[3] for c in dup_list]))

        #duplicate copies = # of unique files that may have one or more copies
        duplicate_copies = int(self.all_dupes) - int(self.distinct_dupes) # number of duplicate copies of unique files
        self.duplicate_copies = str(duplicate_copies)
        
        distinct_files = int(self.num_files) - int(self.duplicate_copies)
        self.distinct_files = str(distinct_files)
        
        # generate sorted format list report;
        path = os.path.join(self.reports_dir, 'formats.csv')
        sql = "SELECT format, id, COUNT(*) as 'num' FROM siegfried GROUP BY format ORDER BY num DESC"
        format_header = ['Format', 'ID', 'Count']
        self.sqlite_to_csv(sql, path, format_header, cursor)
        
        #add top formats to current_dict
        fileformats = []
        formatcount = 0
        try:
            with open(path, 'r') as csvfile:
                formatreader = csv.reader(csvfile)
                next(formatreader)
                for row in formatreader:
                    formatcount += 1
                    fileformats.append(row[0])
                fileformats = [element or 'Unidentified' for element in fileformats] # replace empty elements with 'Unidentified'
                if formatcount > 0:
                    self.current_dict['format_overview'] = "Top file formats (out of {} total) are: {}".format(formatcount, ' | '.join(fileformats[:10]))
                else:
                    self.current_dict['format_overview'] = "-"
                
        except IOError:
            self.current_dict['format_overview'] = "ERROR! No formats.csv file to pull formats from."
        
        # generate sorted format and version list report
        path = os.path.join(self.reports_dir, 'formatVersions.csv')
        sql = "SELECT format, id, version, COUNT(*) as 'num' FROM siegfried GROUP BY format, version ORDER BY num DESC"
        version_header = ['Format', 'ID', 'Version', 'Count']
        self.sqlite_to_csv(sql, path, version_header, cursor)
        
        # get # of unidentified files and write list to CSV
        cursor.execute("SELECT COUNT(*) FROM siegfried WHERE id='UNKNOWN';") # unidentified files
        self.unidentified_files = cursor.fetchone()[0]
        
        sql = "SELECT * FROM siegfried WHERE id='UNKNOWN';"
        path = os.path.join(self.reports_dir, 'unidentified.csv')
        self.sqlite_to_csv(sql, path, full_header, cursor)
        
        # get sorted mimetype list report
        path = os.path.join(self.reports_dir, 'mimetypes.csv')
        sql = "SELECT mime, COUNT(*) as 'num' FROM siegfried GROUP BY mime ORDER BY num DESC"
        mime_header = ['MIME type', 'Count']
        self.sqlite_to_csv(sql, path, mime_header, cursor)
        
        #for dvd jobs, we need to use disk image metadata for dates; for CDDA jobs, we can only list date as unknown
        if self.job_type == 'DVD':
            try:
                with open(self.checksums_dvd, 'rb') as f:
                    file_stats = pickle.load(f)
            except FileNotFoundError:
                pass
                
        #For reporting purposes, we want to catch any files whose current 'mtime' was set during replication in the BDPL process.
        premis_list = self.pickle_load('ls', 'premis_list')
        
        #first, establish when we ran the replication operation.  If no replication operation, check timestamp of folders we created
        bdpl_time = datetime.datetime.fromtimestamp(os.path.getmtime(self.folders_created)).isoformat().replace('T', ' ').split('.')[0]
        
        bdpl_time = datetime.datetime.strptime(bdpl_time, "%Y-%m-%d %H:%M:%S")
        
        #next, go through or file list.  If the 'mtime' is more recent than the 'BDPL' replication action, that means we don't have the original file timestamp.  Only record older/original dates in a date_info list
        date_info = []
        undated_count = []
        if len(file_stats) > 0:
            for dctnry in file_stats:
                dt_time = dctnry['mtime'].replace('T', ' ').split('.')[0]
                dt_time = datetime.datetime.strptime(dt_time, "%Y-%m-%d %H:%M:%S")
                if dt_time < bdpl_time:
                    date_info.append(dctnry['mtime'])
                else:
                    undated_count.append('undated')
            
        #If we've collected any dates in our date_info list, set date ranges and then record years in separate list
        if len(date_info) > 0:
            self.begin_date = min(date_info)[:4]
            self.end_date = max(date_info)[:4]
            self.earliest_date = min(date_info)
            self.latest_date = max(date_info)   
            
            year_info = [x[:4] for x in date_info]
            
        #if date_info is empty, record 'undated' for date ranges
        else:
            self.begin_date = "undated"
            self.end_date = "undated"
            self.earliest_date = "undated"
            self.latest_date = "undated"
            
            year_info = undated_count
            
        #get frequency of each year for report       
        self.year_count = dict(Counter(year_info))
        
        #write year info to file
        path = os.path.join(self.reports_dir, 'years.csv')    
        with open(path, 'w', newline='') as f:
            writer = csv.writer(f)
            year_header = ['Year Last Modified', 'Count']
            writer.writerow(year_header)
            if len(self.year_count) > 0:
                for key, value in self.year_count.items():
                    writer.writerow([key, value])

        # get number of identfied file formats
        cursor.execute("SELECT COUNT(DISTINCT format) as formats from siegfried WHERE format <> '';")
        self.num_formats = cursor.fetchone()[0]

        # get number of siegfried errors and write errors to csv
        cursor.execute("SELECT COUNT(*) FROM siegfried WHERE errors <> '';") # number of siegfried errors
        self.num_errors = cursor.fetchone()[0]
        
        sql = "SELECT * FROM siegfried WHERE errors <> '';"
        path = os.path.join(self.reports_dir, 'errors.csv')
        self.sqlite_to_csv(sql, path, full_header, cursor)

        # calculate size from recursive dirwalk and format
        self.total_size_bytes = 0
        for root, dirs, files in os.walk(self.files_dir):
            for f in files:
                file_path = os.path.join(root, f)
                file_info = os.stat(file_path)
                self.total_size_bytes += file_info.st_size

        self.total_size = self.convert_size(self.total_size_bytes)
        
        # close database connections
        cursor.close()
        conn.close()
        
        #save information to current_dict     
        self.current_dict.update({'Source': self.identifier, 'begin_date': self.begin_date, 'end_date' : self.end_date, 'extent_normal': self.total_size, 'extent_raw': self.total_size_bytes, 'item_file_count': self.num_files, 'item_duplicate_count': self.distinct_dupes, 'FormatCount': self.num_formats, 'item_unidentified_count': self.unidentified_files})  
        
        #get additional metadata from PREMIS about transfer
        premis_list = self.pickle_load('ls', 'premis_list')
        
        if self.job_type in ['Disk_image', 'DVD', 'CDDA']:
            try:
                temp_dict = [f for f in premis_list if f['eventType'] == 'disk image creation'][-1]
            except IndexError:
                try: 
                    temp_dict = [f for f in premis_list if f['eventType'] == 'normalization'][-1]
                except IndexError:
                    temp_dict = {'linkingAgentIDvalue' : '-', 'timestamp' : '-', 'eventOutcomeDetail' : 'Operation not completed.'}
        elif self.job_type == 'Copy_only':
            try:
                temp_dict = [f for f in premis_list if f['eventType'] == 'replication'][-1]
            except IndexError:
                temp_dict = {'linkingAgentIDvalue' : '-', 'timestamp' : '-', 'eventOutcomeDetail' : 'Operation not completed.'}
        
        self.current_dict['job_type'] = self.job_type
        self.current_dict['transfer_method'] = temp_dict['linkingAgentIDvalue']
        self.current_dict['migration_date'] = temp_dict['timestamp']
        
        if temp_dict['eventOutcomeDetail'] == '0' or temp_dict['eventOutcomeDetail'] == 0:
            self.current_dict['migration_outcome'] = 'Success'
        else:
            self.current_dict['migration_outcome'] = 'Failure'
        
        #if using the GUI ingest tool, update any notes provided by technician
        if self.controller.get_current_tab() == 'BDPL Ingest':
            self.current_dict['technician_note'] = self.controller.tabs['BdplIngest'].bdpl_technician_note.get(1.0, tk.END)
        
        #add linked information
        self.current_dict['full_report'] = '=HYPERLINK(".\\{}\\metadata\\reports\\report.html", "View report")'.format(self.identifier)
        self.current_dict['transfer_link'] = '=HYPERLINK("{}", "View transfer folder")'.format(self.identifier)
        
        try:
            if self.current_dict['initial_appraisal'] in ["No appraisal needed", "Move to SDA", "Transfer to SDA"]:
                self.current_dict['final_appraisal'] = "Transfer to SDA"
            elif self.current_dict['initial_appraisal'] == 'Move to SDA and MCO':
                self.current_dict['final_appraisal'] = 'Transfer to SDA and MCO'
            elif self.current_dict['initial_appraisal'] == '-':
                del self.current_dict['final_appraisal']
        except KeyError:
            pass
        
        #save current_dict to file just in case...
        self.pickle_dump('current_dict', self.current_dict)
        
        #create temp file so we can check that this step was already completed
        open(self.stats_done, 'w').close()
    
    def generate_html(self):
    
        print('\n\tCreating HTML...')
        
        #set up html for report
        html_doc = open(self.temp_html, 'w', encoding='utf8') 
        
        # write html
        html_doc.write('<!DOCTYPE html>')
        html_doc.write('\n<html lang="en">')
        html_doc.write('\n<head>')
        html_doc.write('\n<title>IUL Born Digital Preservation Lab report: {}</title>'.format(self.identifier))
        html_doc.write('\n<meta http-equiv="Content-Type" content="text/html; charset=utf-8">')
        html_doc.write('\n<meta name="description" content="HTML report based upon a template developed by Tim Walsh and distributed as part of Brunnhilde v. 1.7.2">')
        html_doc.write('\n<link rel="stylesheet" href="./assets//css/bootstrap.min.css">')
        html_doc.write('\n</head>')
        html_doc.write('\n<body style="padding-top: 80px">')
        # navbar
        html_doc.write('\n<nav class="navbar navbar-expand-lg navbar-dark bg-dark fixed-top">')
        html_doc.write('\n<a class="navbar-brand" href="#">Brunnhilde</a>')
        html_doc.write('\n<button class="navbar-toggler" type="button" data-toggle="collapse" data-target="#navbarNavAltMarkup" aria-controls="navbarNavAltMarkup" aria-expanded="false" aria-label="Toggle navigation">')
        html_doc.write('\n<span class="navbar-toggler-icon"></span>')
        html_doc.write('\n</button>')
        html_doc.write('\n<div class="collapse navbar-collapse" id="navbarNavAltMarkup">')
        html_doc.write('\n<div class="navbar-nav">')
        html_doc.write('\n<a class="nav-item nav-link" href="#Provenance">Provenance</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Stats">Statistics</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#File formats">File formats</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#File format versions">Versions</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#MIME types">MIME types</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Last modified dates by year">Dates</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Unidentified">Unidentified</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Errors">Errors</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Duplicates">Duplicates</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Personally Identifiable Information (PII)">PII</a>')
        html_doc.write('\n<a class="nav-item nav-link" href="#Named Entities">Named Entities</a>')
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        html_doc.write('\n</nav>')
        # content
        html_doc.write('\n<div class="container-fluid">')
        html_doc.write('\n<h1 style="text-align: center; margin-bottom: 40px;">IUL BDPL Brunnhilde HTML report</h1>')
        # provenance
        html_doc.write('\n<a name="Provenance" style="padding-top: 40px;"></a>')
        html_doc.write('\n<div class="container-fluid" style="margin-bottom: 40px;">')
        html_doc.write('\n<div class="card">')
        html_doc.write('\n<h2 class="card-header">Provenance</h2>')
        html_doc.write('\n<div class="card-body">')
        '''need to check if disk image or not'''
        if self.job_type == 'Copy_only':
            html_doc.write('\n<p><strong>Input source: File directory</strong></p>')
        elif self.job_type == 'DVD':
            html_doc.write('\n<p><strong>Input source: DVD-Video (optical disc)</strong></p>')
        elif self.job_type == 'CDDA':
            html_doc.write('\n<p><strong>Input source: Compact Disc Digital Audio (optical disc)</strong></p>')
        elif self.job_type == 'Disk_image':
            html_doc.write('\n<p><strong>Input source: Physical media: {}</strong></p>'.format(self.current_dict.get('content_source_type', 'Unidentified')))
            
        html_doc.write('\n<p><strong>Item identifier:</strong> {}</p>'.format(self.identifier))
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        # statistics
        html_doc.write('\n<a name="Stats" style="padding-top: 40px;"></a>')
        html_doc.write('\n<div class="container-fluid" style="margin-bottom: 40px;">')
        html_doc.write('\n<div class="card">')
        html_doc.write('\n<h2 class="card-header">Statistics</h2>')
        html_doc.write('\n<div class="card-body">')
        html_doc.write('\n<h4>Overview</h4>')
        html_doc.write('\n<p><strong>Total files:</strong> {} (includes contents of archive files)</p>'.format(self.num_files))
        html_doc.write('\n<p><strong>Total size:</strong> {}</p>'.format(self.total_size))
        html_doc.write('\n<p><strong>Years (last modified):</strong> {} - {}</p>'.format(self.begin_date, self.end_date))
        html_doc.write('\n<p><strong>Earliest date:</strong> {}</p>'.format(self.earliest_date))
        html_doc.write('\n<p><strong>Latest date:</strong> {}</p>'.format(self.latest_date))
        html_doc.write('\n<h4>File counts and contents</h4>')
        html_doc.write('\n<p><em>Calculated by hash value. Empty files are not counted in first three categories. Total files = distinct + duplicate + empty files.</em></p>')
        html_doc.write('\n<p><strong>Distinct files:</strong> {}</p>'.format(self.distinct_files))
        html_doc.write('\n<p><strong>Distinct files with duplicates:</strong> {}</p>'.format(self.distinct_dupes))
        html_doc.write('\n<p><strong>Duplicate files:</strong> {}</p>'.format(self.duplicate_copies))
        html_doc.write('\n<p><strong>Empty files:</strong> {}</p>'.format(self.empty_files))
        html_doc.write('\n<h4>Format identification</h4>')
        html_doc.write('\n<p><strong>Identified file formats:</strong> {}</p>'.format(self.num_formats))
        html_doc.write('\n<p><strong>Unidentified files:</strong> {}</p>'.format(self.unidentified_files))
        html_doc.write('\n<h4>Errors</h4>')
        html_doc.write('\n<p><strong>Siegfried errors:</strong> {}</p>'.format(self.num_errors))
        html_doc.write('\n<h2>Virus scan report</h2>')
        with open(self.virus_log, 'r', encoding='utf-8') as f:
            virus_report = f.read().splitlines()
        html_doc.write('\n<p>')
        for line in virus_report:
            html_doc.write('\n{}<br>'.format(line))
        html_doc.write('\n</p>')
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        # detailed reports
        html_doc.write('\n<div class="container-fluid" style="margin-bottom: 40px;">')
        html_doc.write('\n<div class="card">')
        html_doc.write('\n<h2 class="card-header">Detailed reports</h2>')
        html_doc.write('\n<div class="card-body">')
        
        #now write reports to HTML
        report_info = {'File formats' : {'path' : os.path.join(self.reports_dir, 'formats.csv'), 'delimiter' : ','}, 'File format versions' : {'path' : os.path.join(self.reports_dir, 'formatVersions.csv'), 'delimiter' : ','}, 'MIME types' : {'path' : os.path.join(self.reports_dir, 'mimetypes.csv'), 'delimiter' : ','}, 'Last modified dates by year' : {'path' : os.path.join(self.reports_dir, 'years.csv'), 'delimiter' : ','}, 'Unidentified' : {'path' : os.path.join(self.reports_dir, 'unidentified.csv'), 'delimiter' : ','}, 'Errors' : {'path' : os.path.join(self.reports_dir, 'errors.csv'), 'delimiter' : ','}, 'Duplicates' : {'path' : self.pickle_load('ls', 'duplicates'), 'delimiter' : ','}, 'Personally Identifiable Information (PII)' : {'path' : os.path.join(self.bulkext_dir, 'cumulative.txt'), 'delimiter' : '\n'}}
        
        for header, info in report_info.items():
            self.reports_to_html(header, info['path'], info['delimiter'], html_doc)
        
        #Add JavaScript and write html_doc closing tags
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        html_doc.write('\n</div>')
        html_doc.write('\n<script src="./assets//js/jquery-3.3.1.slim.min.js"></script>')
        html_doc.write('\n<script src="./assets//js/popper.min.js"></script>')
        html_doc.write('\n<script src="./assets//js/bootstrap.min.js"></script>')
        html_doc.write('\n<script>$(".navbar-nav .nav-link").on("click", function(){ $(".navbar-nav").find(".active").removeClass("active"); $(this).addClass("active"); });</script>')
        html_doc.write('\n<script>$(".navbar-brand").on("click", function(){ $(".navbar-nav").find(".active").removeClass("active"); });</script>')
        html_doc.write('\n</body>')
        html_doc.write('\n</html>')
        
        # close HTML file
        html_doc.close()

        # write new html file, with hrefs for PRONOM IDs           
        if os.path.exists(self.new_html):
            os.remove(self.new_html)

        # insert pronom links in HTML
        in_file = open(self.temp_html, 'r', encoding='utf8')
        out_file = open(self.new_html, 'w', encoding='utf8')

        for line in in_file:
            regex = r"fmt\/[0-9]+|x\-fmt\/[0-9]+" #regex to match fmt/# or x-fmt/#
            pronom_links_to_replace = re.findall(regex, line)
            new_line = line
            for match in pronom_links_to_replace:
                new_line = line.replace(match, "<a href=\"http://nationalarchives.gov.uk/PRONOM/" + 
                        match + "\" target=\"_blank\">" + match + "</a>")
                line = new_line # allow for more than one match per line
            out_file.write(new_line)

        in_file.close()
        out_file.close()
    
    def reports_to_html(self, header, path, file_delimiter, html_doc):
        """Write csv file to html table"""

        # write header
        html_doc.write('\n<a name="{}" style="padding-top: 40px;"></a>'.format(header))
        html_doc.write('\n<h4>{}</h4>'.format(header))
        
        if header == 'Duplicates':
            html_doc.write('\n<p><em>Duplicates are grouped by hash value.</em></p>')
            dup_list = path
            numline = len(dup_list)
            
            if numline > 1: #aka more rows than just header
                # read md5s from csv and write to list
                hash_list = []
                for row in dup_list:
                    hash_list.append(row[3])
                # deduplicate md5_list
                hash_list = list(OrderedDict.fromkeys(hash_list))
                # for each hash in md5_list, print header, file info, and list of matching files
                for hash_value in hash_list:
                    html_doc.write('\n<p>Files matching checksum <strong>{}</strong>:</p>'.format(hash_value))
                    html_doc.write('\n<table class="table table-sm table-responsive table-bordered table-hover">')
                    html_doc.write('\n<thead>')
                    html_doc.write('\n<tr>')
                    html_doc.write('\n<th>Filename</th><th>Filesize</th>')
                    html_doc.write('<th>Date modified</th>')
                    html_doc.write('<th>Checksum</th>')
                    html_doc.write('\n</tr>')
                    html_doc.write('\n</thead>')
                    html_doc.write('\n<tbody>')
                    for row in dup_list:
                        if row[3] == hash_value:
                            # write data
                            html_doc.write('\n<tr>')
                            for column in row:
                                html_doc.write('\n<td>' + str(column) + '</td>')
                            html_doc.write('\n</tr>')
                    html_doc.write('\n</tbody>')
                    html_doc.write('\n</table>')
            
                #save a copy of the duplicates for the reports
                dup_report = os.path.join(self.reports_dir, 'duplicates.csv')
                with open(dup_report, "w", newline="", encoding='utf-8') as f:
                    writer = csv.writer(f)
                    dup_header = ['Filename', 'Filesize', 'Date modified', 'Checksum']
                    writer.writerow(dup_header)
                    for item in dup_list:
                        writer.writerow(item)
            else:
                html_doc.write('\nNone found.\n<br><br>')
            
        else:
            if not os.path.exists(path):
                open(path, 'w').close()
                
            in_file = open(path, 'r', encoding="utf-8")
            # count lines and then return to start of file
            numline = len(in_file.readlines())
            in_file.seek(0)

            #open csv reader
            r = csv.reader(in_file, delimiter="{}".format(file_delimiter))
            
            # if writing PII, handle separately
            if header == 'Personally Identifiable Information (PII)':
                html_doc.write('\n<p><em>Potential PII in source, as identified by bulk_extractor.</em></p>')  
                pii_list = []

                #check that there are any PII results.  Set value to begin; we will add any found values
                self.current_dict['pii_scan_results'] = '-'
                
                if os.stat(path).st_size > 0:
                    html_doc.write('\n<table class="table table-sm table-responsive table-hover">')
                    html_doc.write('\n<thead>')
                    html_doc.write('\n<tr>')
                    html_doc.write('\n<th>PII type</th>')
                    html_doc.write('\n<th># of matches (may be false)</th>')
                    html_doc.write('\n<th>More information (if available)</th>')
                    html_doc.write('\n</tr>')
                    html_doc.write('\n</thead>')
                    html_doc.write('\n<tbody>')
                    with open(path, 'r') as pii_info:
                        for line in pii_info:
                            html_doc.write('\n<tr>')
                            if 'pii.txt' in line:
                                # write data
                                html_doc.write('\n<td>SSNs, Account Nos., Birth Dates, etc.</td>')
                                html_doc.write('\n<td>' + line.split()[1] + '</td>')
                                html_doc.write('\n<td>Use BE_Viewer to verify results; report.xml file located at: {}.</td>'.format(self.bulkext_dir))
                                pii_list.append('ACCOUNT NOs')
                            if 'ccn.txt' in line:
                                html_doc.write('\n<td>Credit Card Nos.</td>')
                                html_doc.write('\n<td>' + line.split()[1] + '</td>')
                                html_doc.write('\n<td>Use BE_Viewer to verify results; report.xml file located at: {}.</td>'.format(self.bulkext_dir))
                                pii_list.append('CCNs')
                            if 'email.txt' in line:
                                html_doc.write('\n<td>Email address domains (may include 3rd party information)</td>')
                                html_doc.write('\n<td>' + line.split()[1] + '</td>')
                                html_doc.write('\n<td>See: <a href="./email_domain_histogram.txt">Email domain histogram</a></td>')
                                pii_list.append('EMAIL')
                            if 'telephone.txt' in line:
                                html_doc.write('\n<td>Telephone numbers (may include 3rd party information)</td>')
                                html_doc.write('\n<td>' + line.split()[1] + '</td>')
                                html_doc.write('\n<td>See: <a href="./telephone_histogram.txt">Telephone # histogram</a></td>')
                                pii_list.append('TELEPHONE NOs')
                            if 'find.txt' in line:
                                html_doc.write('\n<td>Sensitive terms and phrases</td>')
                                html_doc.write('\n<td>' + line.split()[1] + '</td>')
                                html_doc.write('\n<td>See: <a href="./find_histogram.txt">Keyword histogram</a></td>')
                                pii_list.append('TERMS')
                            html_doc.write('\n</tr>')   
                    html_doc.write('\n</tbody>')
                    html_doc.write('\n</table>')
                    
                    if len(pii_list) > 0:
                        self.current_dict['pii_scan_results'] = '{}.'.format(', '.join(pii_list))
            
                else:
                    html_doc.write('\nNone found.')
                
                self.pickle_dump('current_dict', self.current_dict)

            # otherwise write as normal
            else:
                if numline > 1: #aka more rows than just header
                    # add borders to table for full-width tables only
                    full_width_table_headers = ['Unidentified', 'Errors']
                    if header in full_width_table_headers:
                        html_doc.write('\n<table class="table table-sm table-responsive table-bordered table-hover">')
                    else:
                        html_doc.write('\n<table class="table table-sm table-responsive table-hover">')
                    # write header row
                    html_doc.write('\n<thead>')
                    html_doc.write('\n<tr>')
                    row1 = next(r)
                    for column in row1:
                        html_doc.write('\n<th>' + str(column) + '</th>')
                    html_doc.write('\n</tr>')
                    html_doc.write('\n</thead>')
                    # write data rows
                    html_doc.write('\n<tbody>')
                    for row in r:
                        # write data
                        html_doc.write('\n<tr>')
                        for column in row:
                            html_doc.write('\n<td>' + str(column) + '</td>')
                        html_doc.write('\n</tr>')
                    html_doc.write('\n</tbody>')
                    html_doc.write('\n</table>')
                else:
                    html_doc.write('\nNone found.\n<br><br>')
        
            in_file.close()
    
    def print_premis(self):   
    
        premis_list = self.pickle_load('ls', 'premis_list')
        
        attr_qname = etree.QName("http://www.w3.org/2001/XMLSchema-instance", "schemaLocation")

        PREMIS_NAMESPACE = "http://www.loc.gov/premis/v3"

        PREMIS = "{%s}" % PREMIS_NAMESPACE

        NSMAP = {'premis' : PREMIS_NAMESPACE,
                "xsi": "http://www.w3.org/2001/XMLSchema-instance"}

        events = []
        
        #if our premis file already exists, we'll just delete it and write a new one
        if os.path.exists(self.premis_xml_file):
            os.remove(self.premis_xml_file)
            
        root = etree.Element(PREMIS + 'premis', {attr_qname: "http://www.loc.gov/premis/v3 https://www.loc.gov/standards/premis/premis.xsd"}, version="3.0", nsmap=NSMAP)
        
        object = etree.SubElement(root, PREMIS + 'object', attrib={etree.QName(NSMAP['xsi'], 'type'): 'premis:file'})
        objectIdentifier = etree.SubElement(object, PREMIS + 'objectIdentifier')
        objectIdentifierType = etree.SubElement(objectIdentifier, PREMIS + 'objectIdentifierType')
        objectIdentifierType.text = 'local'
        objectIdentifierValue = etree.SubElement(objectIdentifier, PREMIS + 'objectIdentifierValue')
        objectIdentifierValue.text = self.identifier
        objectCharacteristics = etree.SubElement(object, PREMIS + 'objectCharacteristics')
        compositionLevel = etree.SubElement(objectCharacteristics, PREMIS + 'compositionLevel')
        compositionLevel.text = '0'
        format = etree.SubElement(objectCharacteristics, PREMIS + 'format')
        formatDesignation = etree.SubElement(format, PREMIS + 'formatDesignation')
        formatName = etree.SubElement(formatDesignation, PREMIS + 'formatName')
        formatName.text = 'Tape Archive Format'
        formatRegistry = etree.SubElement(format, PREMIS + 'formatRegistry')
        formatRegistryName = etree.SubElement(formatRegistry, PREMIS + 'formatRegistryName')
        formatRegistryName.text = 'PRONOM'
        formatRegistryKey = etree.SubElement(formatRegistry, PREMIS + 'formatRegistryKey')
        formatRegistryKey.text = 'x-fmt/265' 

        for entry in premis_list:
        
            event = etree.SubElement(root, PREMIS + 'event')
            eventID = etree.SubElement(event, PREMIS + 'eventIdentifier')
            eventIDtype = etree.SubElement(eventID, PREMIS + 'eventIdentifierType')
            eventIDtype.text = 'UUID'
            eventIDval = etree.SubElement(eventID, PREMIS + 'eventIdentifierValue')
            eventIDval.text = str(uuid.uuid4())

            eventType = etree.SubElement(event, PREMIS + 'eventType')
            eventType.text = entry['eventType']

            eventDateTime = etree.SubElement(event, PREMIS + 'eventDateTime')
            eventDateTime.text = entry['timestamp']

            eventDetailInfo = etree.SubElement(event, PREMIS + 'eventDetailInformation')
            eventDetail = etree.SubElement(eventDetailInfo, PREMIS + 'eventDetail')
            eventDetail.text = entry['eventDetailInfo']
            
            #include additional eventDetailInfo to clarify action; older transfers may not include this element, so skip if KeyError
            try:
                eventDetailInfo = etree.SubElement(event, PREMIS + 'eventDetailInformation')
                eventDetail = etree.SubElement(eventDetailInfo, PREMIS + 'eventDetail')
                eventDetail.text = entry['eventDetailInfo_additional']
            except KeyError:
                pass
                
            eventOutcomeInfo = etree.SubElement(event, PREMIS + 'eventOutcomeInformation')
            eventOutcome = etree.SubElement(eventOutcomeInfo, PREMIS + 'eventOutcome')
            eventOutcome.text = str(entry['eventOutcomeDetail'])
            eventOutDetail = etree.SubElement(eventOutcomeInfo, PREMIS + 'eventOutcomeDetail')
            eventOutDetailNote = etree.SubElement(eventOutDetail, PREMIS + 'eventOutcomeDetailNote')
            if entry['eventOutcomeDetail'] == '0':
                eventOutDetailNote.text = 'Successful completion'
            elif entry['eventOutcomeDetail'] == 0:
                eventOutDetailNote.text = 'Successful completion'
            else:
                eventOutDetailNote.text = 'Unsuccessful completion; refer to logs.'

            linkingAgentID = etree.SubElement(event, PREMIS + 'linkingAgentIdentifier')
            linkingAgentIDtype = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierType')
            linkingAgentIDtype.text = 'local'
            linkingAgentIDvalue = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierValue')
            linkingAgentIDvalue.text = 'IUL BDPL'
            linkingAgentRole = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentRole')
            linkingAgentRole.text = 'implementer'
            linkingAgentID = etree.SubElement(event, PREMIS + 'linkingAgentIdentifier')
            linkingAgentIDtype = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierType')
            linkingAgentIDtype.text = 'local'
            linkingAgentIDvalue = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentIdentifierValue')
            linkingAgentIDvalue.text = entry['linkingAgentIDvalue']
            linkingAgentRole = etree.SubElement(linkingAgentID, PREMIS + 'linkingAgentRole')
            linkingAgentRole.text = 'executing software'
            linkingObjectID = etree.SubElement(event, PREMIS + 'linkingObjectIdentifier')
            linkingObjectIDtype = etree.SubElement(linkingObjectID, PREMIS + 'linkingObjectIdentifierType')
            linkingObjectIDtype.text = 'local'
            linkingObjectIDvalue = etree.SubElement(linkingObjectID, PREMIS + 'linkingObjectIdentifierValue')
            linkingObjectIDvalue.text = self.identifier
        
        premis_tree = etree.ElementTree(root)
        
        premis_tree.write(self.premis_xml_file, pretty_print=True, xml_declaration=True, encoding="utf-8")
    
    def record_premis(self, timestamp, event_type, event_outcome, event_detail, event_detail_note, agent_id):
        
        #retrieve our premis_list
        premis_list = self.pickle_load('ls', 'premis_list')
        
        temp_dict = {}
        temp_dict['eventType'] = event_type
        temp_dict['eventOutcomeDetail'] = event_outcome
        temp_dict['timestamp'] = timestamp
        temp_dict['eventDetailInfo'] = event_detail
        temp_dict['eventDetailInfo_additional'] = event_detail_note
        temp_dict['linkingAgentIDvalue'] = agent_id
        
        premis_list.append(temp_dict)
        
        #JUST IN CASE: check to see if we've already written to a premis file (may happen if we have to rerun procedures)
        premis_xml_included = os.path.join(self.temp_dir, 'premis_xml_included.txt')
        if not os.path.exists(premis_xml_included) and os.path.exists(self.premis_xml_file):
        
            PREMIS_NAMESPACE = "http://www.loc.gov/premis/v3"
            NSMAP = {'premis' : PREMIS_NAMESPACE, "xsi": "http://www.w3.org/2001/XMLSchema-instance"}
            parser = etree.XMLParser(remove_blank_text=True)
            tree = etree.parse(self.premis_xml_file, parser=parser)
            root = tree.getroot()
            events = tree.xpath("//premis:event", namespaces=NSMAP)
            
            for e in events:
                temp_dict = {}
                temp_dict['eventType'] = e.findtext('./premis:eventType', namespaces=NSMAP)
                temp_dict['eventOutcomeDetail'] = e.findtext('./premis:eventOutcomeInformation/premis:eventOutcome', namespaces=NSMAP)
                temp_dict['timestamp'] = e.findtext('./premis:eventDateTime', namespaces=NSMAP)
                temp_dict['eventDetailInfo'] = e.findall('./premis:eventDetailInformation/premis:eventDetail', namespaces=NSMAP)[0].text
                temp_dict['eventDetailInfo_additional'] = e.findall('./premis:eventDetailInformation/premis:eventDetail', namespaces=NSMAP)[1].text
                temp_dict['linkingAgentIDvalue'] = e.findall('./premis:linkingAgentIdentifier/premis:linkingAgentIdentifierValue', namespaces=NSMAP)[1].text
                
                if not temp_dict in premis_list:
                    premis_list.append(temp_dict)
                
            #now sort based on ['timestamp'] to make sure we're in chronological order
            premis_list.sort(key=lambda x:x['timestamp'])
            
            #now create our premis_xml_included.txt file so we don't go through this again.
            open(premis_xml_included, 'a').close()

        #now save our premis list
        self.pickle_dump('premis_list', premis_list)
        
    def check_premis(self, term):
        #check to see if an event is already in our premis list--i.e., it's been successfully completed.  Currently only used for most resource-intensive operations: virus scheck, sensitive data scan, format id, and checksum calculation.
        
        #set up premis_list
        premis_list = self.pickle_load('ls', 'premis_list')
        
        #see if term has been recorded at all
        found = [dic for dic in premis_list if dic['eventType'] == term]
        
        #if not recorded, it hasn't been run
        if not found: 
            return False
        else:
            #for virus scans, we will assume that completion may have either a 0 or non-zero value.  No need to run again.
            if term == 'virus check':
                return True
            elif term == 'metadata extraction':
                if [dc for dc in found if 'tree v1.7.0' in dc['linkingAgentIDvalue']]:
                    return True
            #for other microservices, check if operation completed successfully; if so, return True, otherwise False
            else:
                if [dc for dc in found if dc['eventOutcomeDetail'] in ['0', 0]]:
                    return True
                else:
                    return False
                    
    def pickle_load(self, array_type, array_name):
        
        temp_file = os.path.join(self.temp_dir, '{}.txt'.format(array_name))
        
        if array_type == 'ls':
            temp_array = []
        elif array_type == 'dict':
            temp_array = {}
        
        #make sure there's something in the file
        if os.path.exists(temp_file) and os.path.getsize(temp_file) > 0:
            with open(temp_file, 'rb') as file:
                temp_array = pickle.load(file)
                        
        return temp_array

    def pickle_dump(self, array_name, array_instance):
        
        temp_file = os.path.join(self.temp_dir, '{}.txt'.format(array_name))
         
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
            
        with open(temp_file, 'wb') as file:
            pickle.dump(array_instance, file)

    def md5(self, fname):
        hash_md5 = hashlib.md5()
        with open(fname, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def convert_size(self, size):
        # convert size to human-readable form
        if (size == 0):
            return '0 bytes'
        size_name = ("bytes", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
        i = int(math.floor(math.log(size,1024)))
        p = math.pow(1024,i)
        s = round(size/p)
        s = str(s)
        s = s.replace('.0', '')
        return '{} {}'.format(s,size_name[i])

    def check_files(self, some_dir):
        #check to see if it exists
        if not os.path.exists(some_dir):
            print('\n\nError; folder "{}" does not exist.'.format(some_dir))
            return False
        
        #make sure there are files in the 'files' directory
        for dirpath, dirnames, contents in os.walk(some_dir):
            for file in contents:
                if os.path.isfile(os.path.join(dirpath, file)):
                    #return True as soon as a file is found
                    return True
                else: 
                    continue
        
        #if no files found, return false
        return False
        
    def mount_iso(self):
        print('\nMOUNTING .ISO DISK IMAGE FILE...')
        cmd = "Mount-DiskImage -ImagePath '%s'" % self.ripstation_orig_imagefile
        exitcode = subprocess.call('powershell "{}" > null 2>&1'.format(cmd))
        
        return exitcode
        
    def dismount_iso(self):
        print('\nDISMOUNTING DISK IMAGE FILE...')
        cmd = "Dismount-DiskImage -ImagePath '{}'".format(self.ripstation_orig_imagefile)
        exitcode = subprocess.call('powershell "{}" > null 2>&1'.format(cmd))
        
        return exitcode
        
    def get_iso_drive_letter(self):
        cmd = "(Get-DiskImage '{}' | Get-Volume).DriveLetter".format(self.ripstation_orig_imagefile)
        drive_letter = '{}:\\'.format(subprocess.check_output('powershell "%s"' % cmd, text=True).rstrip())
        
        return drive_letter
    
    def run_item_transfer(self):
    
        #Copy only job
        if self.job_type == 'Copy_only':
            self.secure_copy(self.path_to_content)
        
        #Disk image job type
        elif self.job_type == 'Disk_image':
            if self.source_device == '5.25':
                    self.fc5025_image()
            else:
                self.ddrescue_image()
                
            #next, get technical metadata from disk image and replicate files so we can run additional analyses (this step will also involve creating DFXML and correcting MAC times)
            self.disk_image_info()
            self.disk_image_replication()
        
        #DVD job
        elif self.job_type == 'DVD':

            self.ddrescue_image()
            
            #check DVD for title information
            drive_letter = "{}\\".format(self.optical_drive_letter())
            titlecount, title_format = self.lsdvd_check(drive_letter)
            
            #make surre this isn't PAL formatted: need to figure out solution. 
            if title_format == 'PAL':
                print('\n\nWARNING: DVD is PAL formatted! Notify digital preservation librarian so we can configure approprioate ffmpeg command; set disc aside for now...')
                return
            
            #if DVD has one or more titles, rip raw streams to .MPG
            if titlecount > 0:
                self.normalize_dvd_content(titlecount, drive_letter)
            else:
                print('\nWARNING: DVD does not appear to have any titles; job type should likely be Disk_image.  Manually review disc and re-transfer content if necessary.')
                return
        
        #CDDA job
        elif self.job_type == 'CDDA':
            #create a copy of raw pulse code modulated (PCM) audio data and then rip to WAV using cd-paranoia
            self.cdda_image_creation()
            self.cdda_wav_creation()

        print('\n\n--------------------------------------------------------------------------------------------------\n\n')
        
    def run_item_analysis(self):
        
        
        '''run antivirus'''
        print('\nVIRUS SCAN: clamscan.exe')
        if self.check_premis('virus check') and not self.re_analyze:
            print('\n\tVirus scan already completed; moving on to next step...')
        else:
            self.run_antivirus()
    
        '''create DFXML (if not already done so)'''
        if self.check_premis('message digest calculation') and not self.re_analyze:
            print('\n\nDIGITAL FORENSICS XML CREATION:')
            print('\n\tDFXML already created; moving on to next step...')
        else:
            if self.job_type == 'Disk_image':
                #DFXML creation for disk images will depend on the image's file system; check fs_list
                fs_list = self.pickle_load('ls', 'fs_list')
                
                #if it's an HFS+ file system, we can use fiwalk on the disk image; otherwise, use bdpl_ingest on the file directory
                if 'hfs+' in [fs.lower() for fs in fs_list]:
                    self.produce_dfxml(self.imagefile)
                else:
                    self.produce_dfxml(self.files_dir)
            
            elif self.job_type == 'Copy_only':
                self.produce_dfxml(self.files_dir)
            
            elif self.job_type == 'DVD':
                self.produce_dfxml(self.imagefile)
            
            elif self.job_type == 'CDDA':
                self.produce_dfxml(self.image_dir)
                
            '''document directory structure'''
            print('\n\nDOCUMENTING FOLDER/FILE STRUCTURE: TREE')
            if self.check_premis('metadata extraction') and not self.re_analyze:
                print('\n\tDirectory structure already documented with tree command; moving on to next step...')
            else:
                self.document_dir_tree() 
        
        '''run bulk_extractor to identify potential sensitive information (only if disk image or copy job type). Skip if b_e was run before'''
        print('\n\nSENSITIVE DATA SCAN: BULK_EXTRACTOR')
        if self.check_premis('sensitive data scan') and not self.re_analyze:
            print('\n\tSensitive data scan already completed; moving on to next step...')
        else:
            if self.job_type in ['Copy_only', 'Disk_image']:
                self.run_bulkext()
            else:
                print('\n\tSensitive data scan not required for DVD-Video or CDDA content; moving on to next step...')
                
        '''run siegfried to characterize file formats'''
        print('\n\nFILE FORMAT ANALYSIS')
        if self.check_premis('format identification') and not self.re_analyze:
            print('\n\tFile format analysis already completed; moving on to next operation...')
        else:
            self.format_analysis()
        
        #load siegfried.csv into sqlite database; skip if it's already completed
        if not os.path.exists(self.sqlite_done) or self.re_analyze:
            self.import_csv() # load csv into sqlite db
        
        '''generate statistics/reports'''
        if not os.path.exists(self.stats_done) or self.re_analyze:
            self.get_stats()
        
        '''write info to HTML'''
        if not os.path.exists(self.new_html) or self.re_analyze:
            self.generate_html()
        
        #generate PREMIS preservation metadata file
        self.print_premis()
        
        #write info to spreadsheet for collecting unit to review.  Create a spreadsheet object, make sure spreadsheet isn't already open, and if OK, proceed to open and write info.
        shipment_spreadsheet = Spreadsheet(self.controller)
        
        if shipment_spreadsheet.already_open():
            print('\n\nWARNING: {} is currently open.  Close file before continuing and/or contact digital preservation librarian if other users are involved.'.format(shipment_spreadsheet.spreadsheet))
            return
        
        shipment_spreadsheet.open_wb()
        shipment_spreadsheet.write_to_spreadsheet(self.current_dict)
           
        #create file to indicate that process was completed
        if not os.path.exists(self.done_file):
            open(self.done_file, 'a').close()
            
        #copy in .CSS and .JS files for HTML report
        if os.path.exists(self.assets_target):
            pass
        else:
            shutil.copytree(self.assets_dir, self.assets_target)
        
        '''clean up; delete disk image folder if empty and remove temp_html'''
        try:
            os.rmdir(self.image_dir)
        except (WindowsError, PermissionError):
            pass

        # remove temp html file
        try:
            os.remove(self.temp_html)
        except WindowsError:
            pass
        
        '''if using gui, print final details about item'''
        if self.controller.get_current_tab() == 'BDPL Ingest':
            print('\n\n--------------------------------------------------------------------------------------------------\n\nINGEST PROCESS COMPLETED FOR ITEM {}\n\nResults:\n'.format(self.identifier))
            
            du_cmd = 'du64.exe -nobanner "{}" > {}'.format(self.files_dir, self.final_stats)
            
            subprocess.call(du_cmd, shell=True, text=True)   
            
            if os.path.exists(self.image_dir):
                di_count = len(os.listdir(self.image_dir))
                if di_count > 0:
                    print('Disk Img(s):   {}'.format(di_count))
            du_list = ['Files:', 'Directories:', 'Size:', 'Size on disk:']
            with open(self.final_stats, 'r') as f:
                for line, term in zip(f.readlines(), du_list):
                    if "Directories:" in term:
                        print(term, ' ', str(int(line.split(':')[1]) - 1).rstrip())
                    else: 
                        print(term, line.split(':')[1].rstrip())
            print('\n\nReady for next item!') 
 
class Spreadsheet(Shipment):
    def __init__(self, controller):
        Shipment.__init__(self, controller)
        
        self.controller = controller        
        self.identifier = self.controller.identifier.get()
    
    def open_wb(self):
        self.wb = openpyxl.load_workbook(self.spreadsheet)
            
        if self.__class__.__name__ == 'MasterSpreadsheet':
            item_ws = self.wb['Item']
            cumulative_ws = self.wb['Cumulative']
            
        elif self.__class__.__name__ == 'Spreadsheet':
            self.inv_ws = self.wb['Inventory']
            self.app_ws = self.wb['Appraisal']
            self.info_ws = self.wb['Basic_Transfer_Information']
    
    def already_open(self):
        temp_file = os.path.join(os.path.dirname(self.spreadsheet), '~${}'.format(os.path.basename(self.spreadsheet)))
        if os.path.isfile(temp_file):
            return True
        else:
            return False
    
    def return_row(self, ws):
        #set initial Boolean value to false; change to True if barcode is found
        found = False
        row = self.app_ws.max_row+1
        
        if ws.title in ['Inventory', 'Appraisal', 'Item']:
            #if barcode exists in spreadsheet, set variable to that row
            for cell in ws['A']:
                if (cell.value is not None):
                    if self.identifier == str(cell.value).strip():
                        row = cell.row
                        found = True
                        break
        
        #for the cumulative worksheet in master_spreadsheet, we need to match on the unit_name and shipment_date
        elif ws.title == 'Cumulative':
            iterrows = ws.iter_rows()
            next(iterrows)
            
            for row in iterrows:    
                if not row[0].value is None:
                    if self.unit_name in row[0].value and self.shipment_date in row[1].value:
                        newrow = row[0].row
                        found = True
                        break
                        
        return found, row  
             
    def get_spreadsheet_columns(self, ws):

        spreadsheet_columns = {}

        for cell in ws[1]:
            if not cell.value is None:
                if 'identifier' in str(cell.value).lower():
                    spreadsheet_columns['identifier'] = cell.column
                elif cell.value.lower().strip() == 'unit':
                    spreadsheet_columns['unit_name'] = cell.column
                elif cell.value.lower().strip() == 'shipmentID':
                    spreadsheet_columns['shipment_date'] = cell.column
                elif 'accession' in cell.value.lower():
                    spreadsheet_columns['accession_number'] = cell.column
                elif 'collection title' in cell.value.lower():
                    spreadsheet_columns['collection_title'] = cell.column
                elif 'collection id' in cell.value.lower():
                    spreadsheet_columns['collection_id'] = cell.column
                elif 'creator' in cell.value.lower():
                    spreadsheet_columns['collection_creator'] = cell.column
                elif 'physical location' in cell.value.lower():
                    spreadsheet_columns['phys_loc'] = cell.column
                elif 'source type' in cell.value.lower():
                    spreadsheet_columns['content_source_type'] = cell.column
                elif cell.value.strip().lower() == 'title':
                    spreadsheet_columns['item_title'] = cell.column
                elif 'label transcription' in cell.value.lower():
                    spreadsheet_columns['label_transcription'] = cell.column
                elif cell.value.strip().lower() == 'description':
                    spreadsheet_columns['item_description'] = cell.column
                elif 'appraisal notes' in cell.value.lower():
                    spreadsheet_columns['appraisal_notes'] = cell.column
                elif 'content date range' in cell.value.lower():
                    spreadsheet_columns['assigned_dates'] = cell.column
                elif 'instructions' in cell.value.lower():
                    spreadsheet_columns['bdpl_instructions'] = cell.column
                elif 'restriction statement' in cell.value.lower():
                    spreadsheet_columns['restriction_statement'] = cell.column
                elif 'restriction end date' in cell.value.lower():
                    spreadsheet_columns['restriction_end_date'] = cell.column
                elif 'move directly to sda' in cell.value.lower():
                    spreadsheet_columns['initial_appraisal'] = cell.column
                elif 'transfer method' in cell.value.lower():
                    spreadsheet_columns['transfer_method'] = cell.column
                elif 'migration date' in cell.value.lower():
                    spreadsheet_columns['migration_date'] = cell.column
                elif 'migration notes' in cell.value.lower():
                    spreadsheet_columns['technician_note'] = cell.column
                elif 'migration outcome' in cell.value.lower():
                    spreadsheet_columns['migration_outcome'] = cell.column
                elif 'extent (normalized)' in cell.value.lower():
                    spreadsheet_columns['extent_normal'] = cell.column
                elif 'extent (raw)' or 'extracted files extent' in cell.value.lower():
                    spreadsheet_columns['extent_raw'] = cell.column
                elif 'no. of files' or 'extracted files number' in cell.value.lower():
                    spreadsheet_columns['item_file_count'] = cell.column
                elif 'no. of duplicate files' in cell.value.lower():
                    spreadsheet_columns['item_duplicate_count'] = cell.column
                elif 'no. of unidentified files' in cell.value.lower():
                    spreadsheet_columns['item_unidentified_count'] = cell.column
                elif 'file formats' in cell.value.lower():
                    spreadsheet_columns['format_overview'] = cell.column
                elif 'begin date' in cell.value.lower():
                    spreadsheet_columns['begin_date'] = cell.column
                elif 'end date' in cell.value.lower():
                    spreadsheet_columns['end_date'] = cell.column
                elif 'virus status' in cell.value.lower():
                    spreadsheet_columns['virus_scan_results'] = cell.column
                elif 'pii status' in cell.value.lower():
                    spreadsheet_columns['pii_scan_results'] = cell.column
                elif 'full report' in cell.value.lower():
                    spreadsheet_columns['full_report'] = cell.column
                elif 'link to transfer' in cell.value.lower():
                    spreadsheet_columns['transfer_link'] = cell.column
                elif 'appraisal results' in cell.value.lower():
                    spreadsheet_columns['final_appraisal'] = cell.column
                elif 'job type' in cell.value.lower():
                    spreadsheet_columns['job_type'] = cell.column
                    
                #additional elements for master_spreadsheet 'Item' sheet
                elif 'sip creation date' in cell.value.lower():
                    spreadsheet_columns['sip_creation_date'] = cell.column
                elif 'sip extent' in cell.value.lower():
                    spreadsheet_columns['sip_extent'] = cell.column
                elif 'sip md5' in cell.value.lower():
                    spreadsheet_columns['sip_md5'] = cell.column
                elif 'sip filename' in cell.value.lower():
                    spreadsheet_columns['sip_filename'] = cell.column
                    
                #additional elements for master_spreadsheet 'Cumulative'
                elif 'sip count' in cell.value.lower():
                    spreadsheet_columns['sip_count'] = cell.column
                elif 'sips extent' in cell.value.lower():
                    spreadsheet_columns['sips_extent'] = cell.column
                elif 'ingest start date' in cell.value.lower():
                    spreadsheet_columns['ingest_start_date'] = cell.column
                elif 'ingest end date' in cell.value.lower():
                    spreadsheet_columns['ingest_end_date'] = cell.column
                elif 'ingest duration' in cell.value.lower():
                    spreadsheet_columns['ingest_duration'] = cell.column

        return spreadsheet_columns
        
    def write_to_spreadsheet(self, current_dict, ws=None):
    
        if self.__class__.__name__ == 'Spreadsheet':
            ws = self.app_ws
        
        current_row = self.return_row(ws)[1]
        
        ws_cols = self.get_spreadsheet_columns(ws)
    
        for key in ws_cols.keys():
            if key in current_dict:
                self.ws.cell(row=current_row, column=ws_cols[key], value=current_dict[key])

        #save and close spreadsheet
        self.wb.save(self.spreadsheet) 
        
    def check_shipment_progress(self):
        
        #verify unit and shipment_date info has been entered
        if self.unit_name == '' or self.shipment_date == '':
            print('\n\nError; please make sure you have entered a unit ID abbreviation and shipment date.')
            return 
        
        #verify spreadsheet--make sure we only have 1 & that it follows naming conventions
        status, msg = self.verify_spreadsheet()
        if not status:
            print(msg)
            return
        
        self.open_wb()
        
        #get list of all barcodes on appraisal spreadsheet
        app_barcodes = []
        for col in self.app_ws['A'][1:]:
            if not col.value is None:
                app_barcodes.append(str(col.value))
        
        #get list of all barcodes on inventory spreadsheet
        inv_barcodes = {}
        for col in self.inv_ws['A'][1:]:
            if not col.value is None:
                inv_barcodes[str(col.value)] = col.row
        
        inv_list = list(inv_barcodes.keys())        
        
        #check to see if there are any duplicate barcodes in the inventory; print warning if so
        duplicate_barcodes = [item for item, count in Counter(inv_list).items() if count > 1]
        
        if duplicate_barcodes:
            print('\n\nWARNING! Inventory contains at least one duplicate barcode:')
            for dup in duplicate_barcodes:
                print('\t{}\tRow: {}'.format(dup, inv_barcodes[dup]))
        
        current_total = len(inv_list) - len(app_barcodes)
        
        items_not_done = list(set(inv_list) - set(app_barcodes))
        
        if len(items_not_done) > 0:
            print('\n\nThe following barcodes require ingest:\n{}'.format('\n'.join(items_not_done)))
        
        print('\n\nCurrent status: {} out of {} items have been ingested. \n\n{} remain.'.format(len(app_barcodes), len(inv_list), current_total))
        
class MasterSpreadsheet(Spreadsheet):
    def __init__(self, controller):
        Spreadsheet.__init__(self, controller)
        self.controller = controller        
        
        self.spreadsheet = self.controller.bdpl_master_spreadsheet
    
    def verify_master_spreadsheet(self):
        if not os.path.exists(self.spreadsheet):
            return (False, 'Unable to locate {}. Consult with digital preservation librarian.'.format(self.spreadsheet))
        else:
            return (True, "Let's roll!")
        
class ManualPremisEvent(tk.Toplevel):
    def __init__(self, controller):
        tk.Toplevel.__init__(self, controller)
        self.title('BDPL Ingest: Add PREMIS Event')
        self.iconbitmap(r'C:/BDPL/scripts/favicon.ico')
        self.protocol('WM_DELETE_WINDOW', self.close_top)
        self.attributes('-topmost', 'true')
        
        self.controller = controller
        
        status, msg = self.controller.check_main_vars()
        if not status:
            print(msg)
            self.close_top
            return
        
        #self.db = 
        
        if self.controller.get_current_tab() != 'BDPL Ingest' or self.controller.identifier.get()=='':
            self.get_info_frame = tk.Frame(self)
            self.get_info_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
            
            self.l = ttk.Label(self.get_info_frame, text='Enter barcode:', anchor='e', justify=tk.RIGHT, width=25)
            self.l.grid(row=0, column=0, padx=(10,0), pady=10)
            
            self.barcode_entry = tk.Entry(self.get_info_frame, justify=tk.LEFT, width=50)
            self.barcode_entry.grid(row=0, column=1, padx=(0,10), pady=10, sticky='w')
            
            tk.Button(self.get_info_frame, text = 'Use barcode', bg='light slate gray', command=self.add_barcode_value).grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
            tk.Button(self.get_info_frame, text = 'Cancel', bg='light slate gray', command=self.close_top).grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        
        self.barcode_item = DigitalObject(self.controller)

        self.event_frame = tk.LabelFrame(self, text='Item Barcode: {}'.format(self.controller.identifier.get()))
        self.event_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.timestamp_frame = tk.LabelFrame(self, text='Timestamp Information')
        self.timestamp_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.button_frame = tk.Frame(self)
        self.button_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.manual_event = tk.StringVar()
        self.manual_event.set('')
        self.manual_event.trace('w', self.update_fields)
        
        self.events = {
            'replication' : 'Created a copy of an object that is, bit-wise, identical to the original.', 
            'disk image creation' : 'Extracted a disk image from the physical information carrier.',
            'forensic feature analysis' : 'Forensically analyzed the disk image raw bitstream',
            'format identification' : 'Determined file format and version numbers for content recorded in the PRONOM format registry.', 
            'message digest calculation' : 'Extracted information about the structure and characteristics of content, including file checksums.',
            'metadata extraction' : 'Extracted metadata from the object.',
            'metadata modification' : 'Corrected file timestamps to match information extracted from disk image.',
            'normalization' : 'Transformed object to an institutionally supported preservation format.',
            'virus check' : 'Scanned files for malicious programs.'
        }
        
        self.current_event = {}
        widgets = {'event_combobox' : 'Select event:', 'event_software' : 'Software name:', 'event_software_version' : 'Version #:', 'event_command' : 'Command / Description:', 'event_description' : 'Describe preservation event:'}
        
        r = 0
        for name_, label_ in widgets.items():
            l = '{}_label'.format(name_)
            self.current_event[l] = ttk.Label(self.event_frame, text=label_, anchor='e', justify=tk.RIGHT, width=25)
            
            if name_ == 'event_combobox':
                self.current_event[name_] = ttk.Combobox(self.event_frame, textvariable=self.manual_event, values=list(self.events.keys()), justify=tk.LEFT, width=30)
                self.current_event[name_].bind("<<ComboboxSelected>>", self.update_fields)
            else:
                self.current_event[name_] = tk.Entry(self.event_frame, justify=tk.LEFT, width=50)
            
            if name_ != 'event_description':
                self.current_event[l].grid(row=r, column=0, padx=(10,0), pady=10)
                self.current_event[name_].grid(row=r, column=1, padx=(0,10), pady=10, sticky='w')               
            r+=1
        
        self.timestamp_source = tk.StringVar()
        self.timestamp_source.set(None)
        
        info = [['Use "now" for timestamp', 'now'], ['Get timestamp from file', 'file'], ['Get timestamp from folder', 'folder']]
        c = 0
        for i in info:
            ttk.Radiobutton(self.timestamp_frame, text = i[0], variable = self.timestamp_source, value = i[1], command=self.get_timestamp).grid(row=c, column=0, padx=10, pady=10, sticky='w')
            c += 1
        
        self.notice = ttk.Label(self.timestamp_frame, text='NOTE: folder contents will be copied to {}'.format(self.barcode_item.files_dir), wraplength=250)
        
        tk.Button(self.button_frame, text = 'Save Event', bg='light slate gray', command=self.create_manual_premis_event).grid(row=1, column=1, padx=20, pady=10, sticky="nsew")
        tk.Button(self.button_frame, text = 'Quit / Cancel', bg='light slate gray', command=self.close_top).grid(row=1, column=2, padx=20, pady=10, sticky="nsew")
        
        self.button_frame.grid_rowconfigure(0, weight=1)
        self.button_frame.grid_rowconfigure(2, weight=1)
        self.button_frame.grid_columnconfigure(0, weight=1)
        self.button_frame.grid_columnconfigure(3, weight=1)
    
    def add_barcode_value(self):
        if self.barcode_entry.get() == '':
            print('\n\nWARNING: Be sure to enter a barcode value')
            return
        else:
            self.controller.identifier.set(self.barcode_entry.get().trim())
        
        shipment_spreadsheet = Spreadsheet(self.controller)
        shipment_spreadsheet.open_wb()
        
        if not shipment_spreadsheet.return_row(shipment_spreadsheet.inv_ws)[0]:
            print('\n\nWARNING: Barcode value does not appear in spreadsheet')
            return
        else:
            self.get_info_frame.destroy()
    
    def update_fields(self, *args):
        if self.manual_event.get()=='replication':
            self.notice.grid(row=2, column=1, columnspan = 3, padx=10, pady=10, sticky='w')
        else:
            if self.notice.winfo_ismapped():
                self.notice.grid_forget()
                
        #if user adds a different event, we need to get a description of it.  Add fields.
        if not self.events.get(self.manual_event.get()):
            self.current_event['event_description_label'].grid(row=4, column=0, padx=(10,0), pady=10)
            self.current_event['event_description'].grid(row=4, column=1, columnspan=3, padx=(0,10), pady=10)
        #If the event is already recognized, we don't need to have extra fields.  Hide them if they exist. 
        else:
            if self.current_event['event_description_label'].winfo_ismapped():
                self.current_event['event_description_label'].grid_forget()
                self.current_event['event_description'].grid_forget()
            
    def get_timestamp(self):
        if self.timestamp_source.get() == 'now':
            ts = str(datetime.datetime.now())
            
        elif self.timestamp_source.get() == 'folder':
            self.selected_dir = filedialog.askdirectory(parent=self, initialdir=self.controller.bdpl_work_dir, title='Select a folder to extract timestamp from')
            ts = datetime.datetime.fromtimestamp(os.path.getmtime(self.selected_dir)).isoformat()
            
        elif self.timestamp_source.get() == 'file':
            selected_file = filedialog.askopenfilename(parent=self, initialdir=self.controller.bdpl_work_dir, title='Select a file to extract timestamp from')
            ts = datetime.datetime.fromtimestamp(os.path.getmtime(file_)).isoformat()
        
        self.timestamp = ts
        
    def create_manual_premis_event(self):
    
        if not self.events.get(self.manual_event.get()):
            event_desc = self.current_event['event_description'].get()
        else:
            event_desc = self.events[self.manual_event.get()]
        
        #concatenate software name and version #
        vers = '{} v{}'.format(self.current_event['event_software'].get(), self.current_event['event_software_version'].get())
        
        #save info in our 'premis list' for the item 
        self.barcode_item.record_premis(self.timestamp, self.manual_event.get(), 0, self.current_event['event_command'].get(), event_desc, vers)
        
        #if this is a replication event and we've identified a folder, move the folder.  We will also remove any existing DFXML file
        if self.manual_event.get() == 'replication' and self.timestamp_source.get() == 'folder':
            shutil.move(self.selected_dir, self.barcode_item.files_dir)
            
            if os.path.exists(self.barcode_item.dfxml_output):
                os.remove(self.barcode_item.dfxml_output)
                
        print('\nPreservation action ({}) has been succesfully added to PREMIS metadata.')
        
    def close_top(self):
        #close shelve
        
        #close window
        self.destroy()     
        
class RipstationBatch(Shipment):
    def __init__(self, controller):
        Shipment.__init__(self, controller)
        
        self.controller = controller 
        self.ripstation_userdata = self.contoller.ripstation_userdata.get()
        self.ripstation_log = self.controller.ripstation_log.get()
        self.ripstation_ingest_option = self.controller.ripstation_ingest_option.get()
        
        self.ripstation_reports = os.path.join(self.ship_dir, 'ripstation_reports')
        
        #reports
        self.failed_ingest_report = os.path.join(self.ripstation_reports, 'failed_ingest_ripstation.txt')
        self.replicated_report = os.path.join(self.ripstation_reports, 'replicated_ripstation.txt')
        self.analyzed_report = os.path.join(self.ripstation_reports, 'analyzed_ripstation.txt')
        
        #get a timestamp for ripstation batch
        self.rs_timestamp = datetime.datetime.fromtimestamp(os.path.getmtime(self.ripstation_log)).strftime('%Y-%m-%d')
            
        #get a list of barcodes; save to variable
        with open(self.ripstation_userdata, 'r') as ud:
            self.batch_barcodes = ud.read().splitlines()
        
    def set_up(self):
        #set up reports dir
        if not os.path.exists(self.ripstation_reports):
            os.makedirs(self.ripstation_reports)
          
    def ripstation_batch_ingest(self):
        
        #loop through our list of barcodes
        for item in self.batch_barcodes:
            
            print('\nWorking on item: {}'.format(item))
            
            #set our barcode variable and create barcode object
            self.controller.identifier.set(item)
            current_item = DigitalObject(self.controller)
                
            #if item has already failed, skip it.
            if self.controller.check_list(self.failed_ingest_report, current_item.identifier):
                print('\nThis item previously failed.  Moving on to next item...')
                continue
                
            #prep barcode; proceed to next item if any errors
            if not self.controller.check_list(self.replicated_report, current_item.identifier):
                
                print('\nLOADING METADATA AND CREATING FOLDERS...')
                
                status, msg = current_item.prep_barcode()
                if not status:
                    self.controller.write_list(self.failed_ingest_report, '{}\t{}'.format(current_item.identifier, msg))
                    continue
                
                if self.ripstation_ingest_option == 'CDs':
                    #set job_type
                    current_item.job_type = 'CDDA'
                    
                    #make sure .WAV and .CUE file were produced
                    try:
                        current_item.orig_rs_cue = glob.glob(os.path.join(current_item.files_dir, '*.cue'))[0]
                    except IndexError:
                        print("\nMissing '.cue' file; moving on to next item...")
                        self.controller.write_list(self.failed_ingest_report, '{}\tMissing .cue file'.format(current_item.identifier))
                        continue
                    
                    if not os.path.exists(current_item.rs_wav_file):
                        print("\nMissing '.wav' file; moving on to next item...")
                        self.controller.write_list(self.failed_ingest_report, '{}\tMissing .wav file'.format(current_item.identifier))
                        continue
                    
                    #write premis information for creating WAV; we assume that this operation was successful
                    timestamp = datetime.datetime.fromtimestamp(os.path.getmtime(current_item.rs_wav_file)).isoformat()
                    
                    current_item.record_premis(timestamp, 'normalization', 0, 'RipStation BR6-7604 batch .WAV file creation', 'Transformed object to an institutionally supported preservation format (.WAV).', 'RipStation V4.4.13.0')
                    
                    #save ripstation log information for disc to log_dir.  Have to get album # from txt file...
                    txt_file = glob.glob(os.path.join(current_item.files_dir, '*.txt'))[0]
                    
                    album_number = os.path.splitext(os.path.basename(txt_file))[0]

                    with open(current_item.ripstation_item_log, 'w') as outf:
                        outf.write('RipStation V4.4.13.0\n')
                        with open(self.ripstation_log, 'r') as inf:
                            for line in inf.read().splitlines():
                                if album_number in line:
                                    outf.write('{} {}\n'.format(self.rs_timestamp, line))
                    
                    print('\nSTEP 1: FORMAT NORMALIZATION TO .BIN\n\n')
                    
                    #get info about wav file
                    cmd = 'ffprobe -i {} -hide_banner -show_streams -select_streams a'.format(self.rs_wav_file)
                    
                    audio_info = subprocess.check_output(cmd, shell=True, text=True).split('\n')
                    
                    audio_dict = {}
                    
                    for a in audio_info:
                        if '=' in a:
                            audio_dict[a.split('=')[0]] = a.split('=')[1]
                    
                    sample_rate = audio_dict['sample_rate']
                    channels = audio_dict['channels']
                    
                    #now create bin file with raw 16 bit little-endian PCM 
                    cmd = 'ffmpeg -y -i {} -hide_banner -ar {} -ac {} -f s16le -acodec pcm_s16le {}'.format(current_item.rs_wav_file, sample_rate, channels, current_item.rs_cdr_bin)
                    
                    timestamp = str(datetime.datetime.now())
                    exitcode_bin = subprocess.call(cmd, shell=True)
                    
                    ffmpeg_ver = '; '.join(subprocess.check_output('"C:\\Program Files\\ffmpeg\\bin\\ffmpeg" -version', shell=True, text=True).splitlines()[0:2])
               
                    current_item.record_premis(timestamp, 'normalization', exitcode_bin, cmd, 'Transformed object to an institutionally supported preservation format (.BIN)', ffmpeg_ver)
                    
                    #correct cue file; save to file_dir.  
                    with open(current_item.rs_wav_cue, 'w') as outfile:
                        with open(current_item.orig_rs_cue, 'r') as infile:
                            for line in infile.readlines():
                                if line.startswith('FILE'):
                                    outfile.write(line.replace('WAV1', 'WAVE'))
                                elif line.startswith('  TRACK') or line.startswith('    INDEX'):
                                    outfile.write(line)
                    
                    #copy corrected cue file to image_dir; correct FILE reference
                    with open(current_item.rs_cdr_cue, 'w') as outfile:
                        with open(current_item.rs_wav_cue, 'r') as infile:
                            for line in infile.readlines():
                                if line.startswith('FILE'):
                                    outfile.write('FILE "{}" BINARY\n'.format(os.path.basename(current_item.rs_cdr_bin)))
                                elif line.startswith('  TRACK') or line.startswith('    INDEX'):
                                    outfile.write(line)
                    
                    #remove original cue and txt file        
                    os.remove(current_item.orig_rs_cue)
                    os.remove(txt_file)
                    
                    #create toc file
                    cue2toc_ver = subprocess.check_output('cue2toc -v', text=True).split('\n')[0]
                    timestamp = str(datetime.datetime.now())
                    cmd = 'cue2toc -o {} {}'.format(current_item.rs_cdr_toc, current_item.rs_cdr_cue)
                    exitcode = subprocess.call(cmd, shell=True, text=True)
                    
                    #record premis
                    current_item.record_premis(timestamp, 'metadata modification', exitcode, cmd, "Converted the CD's .CUE file to the table of contents (.TOC) format.", cue2toc_ver)
                    
                    #record successful completion
                    self.controller.write_list(self.replicated_report, current_item.identifier)
                    
                elif self.ripstation_ingest_option == 'DVD_Data':
                    
                    #make sure we can account for our original .ISO imagefile
                    if not os.path.exists(current_item.ripstation_orig_imagefile):
                    
                        if os.path.exists(current_item.imagefile):
                            print('\n.ISO file already changed to .DD; converting back to complete operations.')
                            os.rename(current_item.imagefile, current_item.ripstation_orig_imagefile)
                            
                        elif os.path.exists(os.path.join(current_item.image_dir, '{}.mdf'.format(current_item.identifier))):
                            print('\nWARNING: item is Compact Disc Digital Audio; unable to transfer using RipStation DataGrabber.')
                            self.controller.write_list(self.failed_ingest_report, '{}\tDisc is CDDA; transfer using original RipStation'.format(current_item.identifier))
                            continue
                            
                        else:
                            print('\nWARNING: disk image does not exist!  Moving on to next item...')
                            self.controller.write_list(self.failed_ingest_report, '{}\tDisk image does not exist'.format(current_item.identifier))
                            continue
                    
                    #write premis information for disk image creation.  Even if image is unreadable, we assume that this operation was successful
                    timestamp = datetime.datetime.fromtimestamp(os.path.getmtime(current_item.ripstation_orig_imagefile)).isoformat()
                    
                    current_item.record_premis(timestamp, 'disk image creation', 0, 'RipStation BR6-7604 ISO image batch operation', 'Extracted a disk image from the physical information carrier.', 'RipStation DataGrabber V1.0.35.0')
                    
                    #save ripstation log information for disc to log_dir.  Make sure it's only written once...
                    with open(current_item.ripstation_item_log, 'w') as outf:
                        outf.write('RipStation DataGrabber V1.0.35.0\n')
                        with open(self.ripstation_log, 'r') as inf:
                            for line in inf.read().splitlines():
                                if current_item.identifier in line:
                                    outf.write('{} {}\n' % (self.rs_timestamp, line))
                    
                    #mount .ISO so we can verify disk image type
                    exitcode = current_item.mount_iso()
                    if exitcode != 0:
                        print('\nWARNING: failed to mount disk image!  Moving on to next item...')
                        self.controller.write_list(self.failed_ingest_report, '{}\tFailed to mount disk image'.format(current_item.identifier))
                        continue
                    
                    #set media_attached variable to true: confirms that 'media' (mounted disk image) is present; required by bdpl_ingest functions
                    self.controller.media_attached.set(True)
                    current_item.media_attached = self.controller.media_attached.get()
                    
                    #get drive letter for newly mounted disk image
                    drive_letter = current_item.get_iso_drive_letter()
                    
                    #run lsdvd to determine if job_type is DVD or Disk_image
                    print('\nCHECKING IF DISC IS DATA OR DVD-VIDEO...')
                    titlecount, title_format = lsdvd_check(folders, identifier, drive_letter)
                    
                    #fail if disc is PAL-formatted
                    if title_format == 'PAL':
                        print('\nWARNING: PAL-formatted DVD; need to develop appropriate procedures...')
                        self.controller.write_list(self.failed_ingest_report, '{}\tFailed replication: PAL-formatted DVD'.format(current_item.identifier))
                        continue
                    
                    if titlecount == 0:
                        current_item.job_type = 'Disk_image'
                        
                        #dismount disk image
                        exitcode = current_item.dismount_iso()
                        if exitcode != 0:
                            print('\nWARNING: failed to dismount disk image!  Moving on to next item...')
                            self.controller.write_list(self.failed_ingest_report, '{}\tFailed to dismount disk image'.format(current_item.identifier))
                            continue
                        
                        #rename to '.dd' file extension
                        timestamp = str(datetime.datetime.now())
                        
                        os.rename(current_item.ripstation_orig_imagefile, current_item.imagefile)
                        
                        #document change to filename
                        current_item.record_premis(timestamp, 'filename change', 0, 'os.rename({}, {})'.format(current_item.ripstation_orig_imagefile, current_item.imagefile), 'Modified the filename, changing extension from .ISO to .DD to ensure consistency with IUL BDPL practices', 'Python %s' % sys.version.split()[0])
                    
                        #next, get technical metadata from disk image and replicate files so we can run additional analyses (this step will also involve creating DFXML and correcting MAC times)
                        current_item.disk_image_info()
                        current_item.disk_image_replication()
                    
                    else:
                        current_item.job_type = 'DVD'
                        
                        current_item.normalize_dvd_content(titlecount, drive_letter)
                        
                        #dismount disk image
                        print('\nDISMOUNTING DISK IMAGE FILE...') 
                        exitcode = current_item.dismount_iso()
                        if exitcode != 0:
                            print('\nWARNING: failed to dismount disk image!  Moving on to next item...')
                            self.controller.write_list(self.failed_ingest_report, '{}\tFailed to dismount disk image'.format(current_item.identifier))
                            continue
                            
                        #rename to '.dd' file extension
                        os.rename(current_item.ripstation_orig_imagefile, current_item.imagefile)
                    
                    #record successful status if files exist; otherwise note failure
                    if current_item.check_files(current_item.files_dir):
                        self.controller.write_list(self.replicated_report, current_item.identifier)
                    else:
                        print('\nWARNING: failed to replicate files!  Moving on to next item...')
                        self.controller.write_list(self.failed_ingest_report, '{}\tFailed to replicate files'.format(current_item.identifier))
                        continue
            
            if not self.controller.check_list(self.analyzed_report, current_item.identifier):
                current_item.run_item_analysis()
                
                #check procedures
                jobs = ['virus check', 'metadata extraction', 'message digest calculation', 'format identification']
                
                if current_item.job_type == 'Disk_image':
                    jobs.append('sensitive data scan')
                
                failed_analysis_jobs = []

                for job in jobs:
                    if not current_item.check_premis(job):
                        failed_analysis_jobs.append(job)
                
                if len(failed_analysis_jobs) > 0:
                    print('\nWARNING: analysis did not complete with:\n\t{}'.format('\n\t'.join(failed_analysis_jobs)))
                    self.controller.write_list(self.failed_ingest_report, '{}\tFailed analysis job(s): {}'.format(current_item.identifier, ', '.join(failed_analysis_jobs)))
                    continue
                else:
                    self.controller.write_list(self.analyzed_report, current_item.identifier)
                        
                        
    def clean_up(self):
        
        #move log and userdata file to ripstation_reports
        
        pass
        
class SdaBatchDeposit(Shipment):
    def __init__(self, controller):
        Shipment.__init__(self, controller)
        
        #set up variables
        self.controller = controller
        self.separations_status = self.controller.separations_status.get()
        self.separations_file = self.controller.separations_file.get()
             
        self.bdpl_archiver_target = os.path.join(self.controller.bdpl_archiver_spool_dir, self.controller.tabs['SdaDeposit'].archiver_dir.get())
        
        #set up deposit directories
        self.bag_report_dir = os.path.join(self.ship_dir, 'bag_reports')
        self.deaccession_dir = os.path.join(self.ship_dir, 'deaccessioned')
        self.unaccounted_dir = os.path.join(self.ship_dir, 'unaccounted') 
        self.deposit_dirs = [self.bag_report_dir, self.deaccession_dir, self.unaccounted_dir]
        
        #set up shelve to track status
        self.sda_status = os.path.join(self.bag_report_dir, 'sda_status')
        self.status_db = shelve.open(self.sda_status, writeback=True)
        
        self.db_lists = [
            'spreadsheet_barcodes', #list of all barcodes in spreadsheet
            'directory_barcodes', #list of all barcodes in ship_dir
            'missing_from_ship_dir', #list of barcodes that are in spreadsheet, but not ship_dir
            'mco_deposit', #list of barcodes that need to be deposited to MCO
            'deaccessioned', #barcodes that will be deaccessioned
            'unaccounted', #barcodes not listed in spreadsheet
            'started', #barcodes that have started deposit process 
            'prepped', #barcodes that passed inital preparations
            'separations_completed', #'separated' content has been removed from the barcode
            'bagged', #barcodes that have been successfully bagged
            'tarred', #barcodes that have been successfully tarred
            'moved', #barcodes that have been successfully moved to Archiver dropbox
            'metadata_written', #barcodes that have metadata  successfully written to master spreadsheet                   
            'cleaned', #barcodes that have been deleted from shipment directory
            'puid_report', #cumulative stats on PUIDs in shipment               
        ]
            
        self.db_dicts = [
            'failed', #record any failures
            'format_report', #cumulative stats on formats in the shipment
            'shipment_stats', #general stats on shipment
            'separation-stats', #stats on separations
            'other_action' #barcodes with 'other' final appraisal decisions
        ]
            
        if len(self.status_db) == 0:
 
            for ls in self.db_lists:
                self.status_db[ls] = []
            
            for dc in self.db_dicts:
                self.status_db[dc] = {}         
        
    def prep_sda_batch(self, master_spreadsheet, shipment_spreadsheet):
        '''
        Check variables and spreadsheets
        '''
        #make sure key variables are present
        status, msg = self.controller.check_main_vars()
        if not status:
            return (status, msg)
        
        #verify master spreadsheet
        status, msg = master_spreadsheet.verify_master_spreadsheet()
        if not status:
            return (status, msg)
                
        #make sure master spreadsheet is not open
        if master_spreadsheet.already_open():
            return (False, '\n\nWARNING: {} is currently open.  Close file before continuing and/or contact digital preservation librarian if other users are involved.'.format(master_spreadsheet.spreadsheet))
            
        #open master_spreadsheet; get worksheets
        master_spreadsheet.open_wb()
        
        #verify shipment spreadsheet
        status, msg = shipment_spreadsheet.verify_spreadsheet()
        if not status:
            return (status, msg)
                
        #make sure shipment spreadsheet is not open
        if shipment_spreadsheet.already_open():
            return (False, '\n\nWARNING: {} is currently open.  Close file before continuing and/or contact digital preservation librarian if other users are involved.'.format(shipment_spreadsheet.spreadsheet))
        
        #open shipment_spreadsheet; get worksheets
        shipment_spreadsheet.open_wb()
        
        #make deposit folders
        for dir in self.deposit_dirs:
            if not os.path.exists(dir):
                os.mkdir(dir)
        
        '''    
        Determine which barcodes are present and need to be deposited to SDA
        '''
        #move into ship_dir
        os.chdir(self.ship_dir)
        
        #barcodes in ship_dir
        self.status_db['directory_barcodes'] = [d for d in os.listdir(self.ship_dir) if os.path.isdir(d) and not d in ['review', 'bag_reports', 'unaccounted', 'deaccessioned', 'ripstation_reports', 'mco_prep', 'reports']] #including legacy directory names; at some point, we will need to simplify this list
        
        #barcodes in shipment spreadsheet
        for barcode in shipment_spreadsheet.app_ws['A'][1:]:
            if not barcode.value is None and not str(barcode.value) in self.status_db['spreadsheet_barcodes']:
                self.status_db['spreadsheet_barcodes'].append(str(barcode.value))
        
        #see if any barcode folders are missing
        self.status_db['missing_from_ship_dir'] = list(set(self.status_db['spreadsheet_barcodes']) - set(self.status_db['directory_barcodes']))
        
        #don't include completed items or alternative appraisal decisions 
        for ls in ['cleaned', 'deaccessioned', 'other_action']:
            if len(self.status_db[ls]) > 0:
                for item in self.status_db[ls]:
                    try:
                        self.status_db['directory_barcodes'].remove(item)
                    except ValueError:
                        pass
            
        #check if there are any folders in the shipment NOT in spreadsheet.  
        self.status_db['unaccounted'] = list(set(self.status_db['directory_barcodes']) - set(self.status_db['spreadsheet_barcodes']))
        
        #If there are any barcodes that do not appear in the spreadsheet, move them to the 'unaccounted_dir'
        if len(self.status_db['unaccounted']) > 0:
            for item in self.status_db['unaccounted']:
                try:
                    shutil.move(item, self.unaccounted_dir)
                    self.status_db['directory_barcodes'].remove(item)
                except (PermissionError, OSError) as e:
                    self.write_db('failed', item, 'Move unaccounted failure\t{}'.format(e))
        
        #get stats on duration of ingest
        if len(self.status_db['directory_barcodes']) > 0:
            latest_date = datetime.datetime.fromtimestamp(os.stat(max(self.status_db['directory_barcodes'], key=os.path.getmtime)).st_ctime).strftime('%Y%m%d')
            
            earliest_date = datetime.datetime.fromtimestamp(os.stat(min(self.status_db['directory_barcodes'], key=os.path.getmtime)).st_ctime).strftime('%Y%m%d')
        
        #if we've already recorded duration information, update values; otherwise, add values to shipment_stats dict
        if self.status_db['shipment_stats'].get('ingest_start_date'):      
            if earliest_date < self.status_db['shipment_stats']['ingest_start_date']:
                self.status_db['shipment_stats']['ingest_start_date'] = earliest_date               
            if latest_date > self.status_db['shipment_stats']['ingest_end_date']:
                self.status_db['shipment_stats']['ingest_end_date'] = latest_date       
        else:
            self.status_db['shipment_stats']['ingest_start_date'] = earliest_date
            self.status_db['shipment_stats']['ingest_end_date'] = latest_date             
        
        #calculate total duration for ingest; use 1 day as minimum timedelta
        tdelta = datetime.datetime.strptime(self.status_db['shipment_stats']['ingest_end_date'], '%Y%m%d') - datetime.datetime.strptime(self.status_db['shipment_stats']['ingest_start_date'], '%Y%m%d')
        
        if tdelta < datetime.timedelta(days=1):
            self.status_db['shipment_stats']['ingest_duration'] = 1
        else:
            self.status_db['shipment_stats']['ingest_duration'] = int(str(tdelta).split()[0])
            
        #save info
        self.status_db.sync()
        
        #get columns for current appraisal spreadsheet
        self.appraisal_ws_columns = shipment_spreadsheet.get_spreadsheet_columns(shipment_spreadsheet.app_ws)
        
        return (True, 'Ready to deposit!')
    
    def deposit_barcodes_to_sda(self, master_spreadsheet, shipment_spreadsheet):
        
        for item in self.status_db['directory_barcodes']:

            #set identifier variable; create barcode object
            self.controller.identifier.set(item.strip())
            current_item = DigitalObject(self.controller)
            
            print('\nWorking on item: {}'.format(current_item.identifier))
            
            #continue to next item if we've already completed item
            if current_item.identifier in self.status_db['cleaned']:
                print('\n{} completed.'.format(current_item.identifier))
                continue            
            
            #load metadata
            current_item.load_item_metadata(shipment_spreadsheet)
            
            #record status
            if not current_item.identifier in self.status_db['started']:
                self.write_db('started', current_item.identifier)
            
            '''Check final_appraisal information for disposition of content'''
            ###ALSO CHECK IF BARCODE IS IN AN 'MCO-COMPLETED' LIST...
            if 'mco' in current_item.current_dict['final_appraisal'].lower():
                print('\n\nContent will be deposited to Media Collections Online. Moving on to next item...')
                self.write_db('mco_deposit', current_item.identifier)
                continue
                
            elif current_item.current_dict['final_appraisal'] == "Delete content":
                try:
                    print('\n\tContent will not be transferred to SDA.  Continuing with next item.')
                    shutil.move(current_item.barcode_dir, self.deaccession_dir)
                    self.write_db('deaccessioned', current_item.identifier)
                
                except (PermissionError, OSError) as e:
                    self.write_db('failed', current_item.identifier, 'deaccession\t{}'.format(e))
                        
                continue
            
            #allow for additional transfer instructions (i.e., transfer to both MCO and SDA)
            elif 'transfer' and 'sda' in current_item.current_dict['final_appraisal'].lower():
                        
                '''PREPARE ITEM: VERIFY CONTENT IS PRESENT AND GET STATS'''
                if not current_item.identifier in self.status_db['prepped']:
                    
                    #check if there are any reported files; double check image_dir
                    if current_item.current_dict['item_file_count'] is None or current_item.current_dict['item_file_count'] == 0:
                    
                        if not current_item.check_files(current_item.files_dir) and not current_item.check_files(current_item.image_dir):
                            
                            print('\n\tItem has no files or disk image!  Moving on...')
                            
                            self.write_db('failed', current_item.identifier,  'check_folder\tNO CONTENT IN BARCODE FOLDER; CHANGE APPRAISAL DECISION?')
                            
                            continue
                    
                    #set up barcode dict to collection format
                    if not self.status_db['format_report'].get(current_item.identifier):
                        self.status_db['format_report'][current_item.identifier] = {}
                    
                    #get file format info
                    format_csv = os.path.join(self.reports_dir, 'formatVersions.csv')
                    if os.path.exists(format_csv):
                        with open(format_csv, 'r') as fi:
                            fi = csv.reader(fi)
                            #skip header row
                            next(fi)
                            #loop through format csv; create a dictionary for each row, recording the PUIDs (with format names and versions) and a count of each
                            for line in fi:
                                puid = line[1]
                                self.status_db['format_report'][current_item.identifier][puid] = {'format' : line[0], 'version' : line[2], 'count' : int(line[3])}
                            self.status_db.sync()
                    
                    #item prepped: record status
                    self.write_db('prepped', current_item.identifier)
                        
                '''COMPLETE SEPARATIONS AND REMOVE TEMP & B_E FILES'''
                if not current_item.identifier in self.status_db['separations_completed']:
                    print('\n\tSeparating unnecessary files...\n')
                    
                    #remove folders
                    for dir in [current_item.temp_dir, current_item.bulkext_dir, current_item.assets_target]:
                        if os.path.exists(dir):
                            shutil.rmtree(dir)
                    
                    #remove temp files/reports
                    for f in ["duplicates.csv", "errors.csv", "formats.csv", "formatVersions.csv", "mimetypes.csv", "unidentified.csv", "uniqueyears.csv", "years.csv", 'email_domain_histogram.txt', 'find_histogram.txt', 'telephone_histogram.txt', 'report.html']:
                        report = os.path.join(current_item.reports_dir, f)
                        if os.path.exists(report):
                            os.remove(report)
                    
                    #address separations, if indicated
                    if self.separations_status:
                        
                        #set up log file
                        current_item.separations_log = os.path.join(current_item.log_dir, 'separations.txt')
                        
                        #get content-to-be-separated from the item barcode
                        with open(self.separations_file, 'r') as f:
                            sep_list = [file for file in f.read().replace('"', '').replace("'", "").splitlines() if current_item.identifier in file]
                        
                        #make sure we have separations for this barcode
                        if len(sep_list) > 0:
                            
                            for item in sep_list:
                                #set up list to hold all files to be separated
                                files_to_be_separated = []
                            
                                #split path at shipment_date in case drive letters differ with BDPL workstation and collecting unit that prepared separations_file
                                item = item.split('{}\\'.format(self.shipment_date))[1] 
                                
                                #if a wildcard is used, we will use glob to build a list of all files/folders matching pattern
                                if '\\**' in item:
                                    files_to_be_separated = glob.glob(item, recursive=True)
                                
                                elif '\\*' in sep_item:
                                    files_to_be_separated = glob.glob(item)
                                
                                #build recursive list of all files in the folder   
                                elif os.path.isdir(item):                                
                                    for root, dirs, files in os.walk(item):
                                        for f in files:
                                            files_to_be_separated.append(os.path.join(root, f))
                                    #also add parent folder so that we can remove it.
                                    files_to_be_separated.append(item)
                                            
                                elif os.path.isfile(item):
                                    files_to_be_separated.append(item)

                                else:
                                    print('\n\tUnable to identify: {}'.format(item))
                                    files_to_be_separated.append('FAIL: {}'.format(item))
                            
                            #check to see if we failed to identify any separation targets; if so, fail barcode so we can troublshoot
                            if [f for f in files_to_be_separated if 'FAIL' in f]:
                                self.write_db('failed', current_item.identifier,  'separations\t{}'.format(','.join([f for f in files_to_be_separated if 'FAIL' in f])))
                                continue
                            #if no failures, move forward with separations
                            else:                                
                                #separate items and gather stats
                                status = self.separate_content(current_item, files_to_be_separated, shipment_spreadsheet)
                                
                                if not status:
                                    continue
                                else:
                                    self.write_db('separations_completed', current_item.identifier)
                
                '''BAG FOLDER'''
                if not current_item.identifier in self.status_db['bagged']:
                
                    print('\n\tCreating bag for barcode folder...')
                    
                    #set metadata for bag.
                    current_item.current_dict['bag_description'] = 'Source: {}. | Label: {}. | Title: {}. | Appraisal notes: {}. | Date range: {}-{}'.format(current_item.current_dict['content_source_type'], current_item.current_dict['label_transcription'], current_item.current_dict.get('item_title', '-'),  current_item.current_dict['appraisal_notes'], current_item.current_dict['begin_date'], current_item.current_dict['end_date'])
                    
                    #make sure we haven't added a temp_dir if we had to restart packaging
                    if os.path.exists(current_item.temp_dir):
                        shutil.rmtree(current_item.temp_dir)
                    
                    try:
                        #create bag
                        bagit.make_bag(current_item.barcode_dir, {"Source-Organization" : current_item.unit_name, "External-Description" : current_item.current_dict['bag_description'], "External-Identifier" : current_item.identifier}, checksums=["md5"])
                        
                        print('\tBagging complete.')
                        
                        #record completion
                        self.write_db('bagged', current_item.identifier)
                    
                    #continue on to next item if failure
                    except (RuntimeError, PermissionError, bagit.BagError, OSError) as e:
                        print("\tUnexpected error: ", e)
                        
                        self.write_db('failed', current_item.identifier, 'bagit\t{}'.format(e))
                        
                        continue
                
                '''CREATE TAR'''
                #make sure file hasn't already been tarred
                if not current_item.identifier in self.status_db['tarred']:
                
                    #Make sure we have enough space to create tar file (just to be sure; we should check first, as a rule)
                    print('\n\tChecking available space...')
                    
                    #first check available space
                    (total_space, used_space, free_space) = shutil.disk_usage(os.getcwd())
                    
                    #now get size of barcode_dir
                    cmd = 'du -s {}'.format(current_item.barcode_dir)
                    
                    output = subprocess.run(cmd, shell=True, text=True, capture_output=True)
                    
                    dir_size = int(output.stdout.split()[0])
                    
                    #check if the new archive will have sufficient space on disk; include addition 10240 bytes for tar file. Ff so, continue.  If not, exit with a warning
                    available_space = int(free_space) - (dir_size * 2 + 10240)
                    
                    #fail item if not enough space to create tar
                    if available_space <= 0:
                        print('\n\tWARNING! Insufficient space to create tar archive.\n\t\tAvailable space: %s\n\t\tSize needed for archive: %s' % (free_space, string(dir_size)))
                        
                        self.write_db('failed', current_item.identifier, 'Insufficient space\t need minimum of {} bytes'.format(dir_size))
                        
                        continue
                    
                    else:
                        print('\tCheck complete; sufficient space for tar file.')
                        
                    #make sure we haven't added a temp_dir to our bag...
                    if os.path.exists(current_item.temp_dir):
                        shutil.rmtree(current_item.temp_dir)
                        
                    print('\n\tCreating tar archive...')
                    
                    try:
                        with tarfile.open(current_item.tar_file, "w") as tar:
                            tar.add(current_item.barcode_dir, arcname=current_item.identifier)
                            
                        print('\tTar archive created')
                        
                        self.write_db('tarred', current_item.identifier)
                        
                        
                    except (RuntimeError, PermissionError, IOError, EnvironmentError) as e:
                        
                        print("\tUnexpected error: ", e)
                        
                        self.write_db('failed', current_item.identifier, 'tar\t{}'.format(e))
                        
                        continue
                
                '''MOVE TAR TO ARCHIVER LOCATION'''
                if not current_item.identifier in self.status_db['moved']:
                
                    print('\n\tMoving tar file to Archiver folder...')
                    
                    #get some stats on SIP and store values in current_item.current_dict
                    print('\tGenerating SIP statistics...')
                    current_item.current_dict['sip_extent'] = current_item.get_size(current_item.tar_file)
                    
                    current_item.current_dict['sip_md5'] = current_item.md5(current_item.tar_file)
                    
                    current_item.current_dict['sip_filename'] = os.path.basename(current_item.tar_file)
                    
                    current_item.current_dict['sip_creation_date'] = datetime.datetime.fromtimestamp(os.path.getmtime(current_item.tar_file)).isoformat()
                    
                    #save current_dict
                    current_item.pickle_dump('current_dict', current_item.current_dict)
                    
                    try:
                        shutil.move(current_item.tar_file, self.bdpl_archiver_target)
                        
                        print('\tTar file moved.')
                        
                        self.write_db('moved', current_item.identifier)
                        
                    except (RuntimeError, PermissionError, IOError, EnvironmentError) as e:
                    
                        print("\tUnexpected error: ", e)
                        
                        self.write_db('failed', current_item.identifier, 'move\t{}'.format(e))
                        
                        continue
                
                '''WRITE STATS TO MASTER SPREADSHEET'''
                if not current_item.identifier in self.status_db['metadata_written']:
                    
                    master_spreadsheet.write_to_spreadsheet(current_item.current_dict, master_spreadsheet.item_ws)
                    
                    #set up additional shipment stats keys if not already done so
                    if not self.status_db['shipment_stats'].get('sip_count'):
                        self.status_db['shipment_stats']['unit_name'] = current_item.unit_name
                        self.status_db['shipment_stats']['shipment_date'] = current_item.shipment_date
                        self.status_db['shipment_stats']['sip_count'] = 0
                        self.status_db['shipment_stats']['extent_raw'] = 0
                        self.status_db['shipment_stats']['item_file_count'] = 0
                        self.status_db['shipment_stats']['sips_extent'] = 0
                    
                    #update statistics & save shelve
                    self.status_db['shipment_stats']['sip_count'] += 1
                    self.status_db['shipment_stats']['extent_raw'] += current_item.current_dict['extent_raw']
                    self.status_db['shipment_stats']['item_file_count'] += current_item.current_dict['item_file_count']
                    self.status_db['shipment_stats']['sips_extent'] += current_item.current_dict['sip_extent']                   
                    self.status_db.sync()
                    
                    #record completion
                    self.write_db('metadata_written', current_item.identifier)
                
                '''CLEAN ORIGINAL BARCODE FOLDER'''
                #remove original folder
                if not current_item.identifier in self.status_db['cleaned']:
                    print('\n\tRemoving original folder...')
                    
                    cmd = 'RD /S /Q "{}"'.format(current_item.barcode_dir)
                    
                    try:
                        subprocess.check_output(cmd, shell=True)
                        print('\tFolder removed')
                        self.write_db('cleaned', current_item.identifier)
                        
                    #if unsuccessful, note failure and continue
                    except (PermissionError, subprocess.CalledProcessError, OSError) as e:
                        print("\tUnexpected error: ", e)
                        self.write_db('failed', current_item.identifier, 'clean_original\t{}'.format(e))
                        continue
                        
                '''BARCODE IS NOW DONE!'''
                print('\n\t{} COMPLETED\n---------------------------------------------------------------'.format(current_item.identifier))
                
                #if barcode had previously failed, remove it from list.
                if current_item.identifier in self.status_db['failed']:
                    del self.status_db['failed'][current_item.identifier]
            
            #if other appraisal decision is indicated, note barcode in 'other_action' list
            else:
            
                print('\n\tAlternate appraisal decision: {}. \n\tConfer with collecting unit as needed.'.format(current_item.current_dict['final_appraisal']))
                
                if current_item.identifier not in self.status_db['other_action']:
                    self.write_db('other_action', current_item.identifier, current_item.current_dict['final_appraisal'])
                continue
        
        '''LOOP THROUGH DIRECTORY BARCODES IS NOW COMPLETE; REPORT RESULTS'''        
        #get lists from status files: how many barcodes are in each list
        reports = ['started', 'cleaned', 'unaccounted', 'deaccessioned', 'other_action', 'failed']
        
        print('\nBATCH COMPLETED:')
        ls_report = ''
        for report in reports:
            if report == 'cleaned':
                label_ = 'completed'
            elif report == 'other_action':
                label_ = 'with other appraisal decisions'
                if len(self.status_db[report]) > 0:
                    ls_report = ['{}:\t{}'.format(k, v) for k, v in self.status_db[report].items()]
            elif report == 'failed':
                if len(self.status_db[report]) > 0:
                    ls_report = ['{}:\t{}'.format(k, v) for k, v in self.status_db[report].items()]
            elif report == 'unaccounted':
                label_ = 'not in shipment'
                if len(self.status_db[report]) > 0:
                    ls_report = self.status_db['unaccounted']
            else:
                label_ = report
            
            print('\n\tItems {}: {}'.format(label_, len(self.status_db[report])))
            
            if ls_report != '':
                print('\t{}'.format('\n\t'.join(ls_report)))
            
        '''UPDATE CUMULATIVE INFORMATION'''
        print('\n\n------------------------------------------------------------\n\nUPDATING MASTER SPREADSHEET')
        
        print('\nWriting shipment stats...')
        
        #write shipment stats to master_spreadsheet
        master_spreadsheet.write_to_spreadsheet(self.status_db['shipment_stats'], master_spreadsheet.cumulative_ws)
        
        #now write format information to master_spreadsheet.  Create new sheet; if it already exists, remove it and rewrite
        print('\nWriting format information...')
        
        puids = 'puids_{}_{}'.format(current_item.unit_name, current_item.shipment_date)
    
        #if this puid sheet already exists, we'll just remove it and start anew...
        if puids in master_spreadsheet.wb.sheetnames:
            master_spreadsheet.wb.remove(master_spreadsheet.wb[puids])
        
        master_spreadsheet.puid_ws = master_spreadsheet.wb.create_sheet(puids)
        
        #set up a header
        puid_header = []
        for barcode in self.status_db['format_report']:
            for puid in self.status_db['format_report'][barcode]:
                if not puid in puid_header and self.status_db['format_report'][barcode][puid]['count'] > 0:
                    puid_header.append(puid)
                    
        #natural-order sort the header
        sort_puids(puid_header)
        
        #insert 'barcode' as the first item in list
        puid_header.insert(0, 'barcode')
        
        #append header to puid sheet
        master_spreadsheet.puid_ws.append(puid_header)
        
        #create a dictionary to use to refer to puid columns in the sheet.  Add 1 to index, as 1st column is 1 (not 0) in openpyxl
        puid_cols = {}
        [puid_cols.update( {x : puid_header.index(x)+1} ) for x in puid_header]
        
        #now loop through all barcodes
        for barcode in self.status_db['format_report']:
            #get a new row for each barcode; write in barcode value
            row = master_spreadsheet.puid_ws.max_row+1
            
            master_spreadsheet.puid_ws.cell(row=row, column=puid_cols['barcode'], value=barcode)
            
            #loop through the puids of each barcode; write count to spreadsheet
            for puid in self.status_db['format_report'][barcode]:
                if self.status_db['format_report'][barcode][puid]['count'] < 1:
                    continue   
                master_spreadsheet.puid_ws.cell(row=row, column=puid_dict[puid], value=self.status_db['format_report'][barcode][puid]['count'])
                
        #Finally, tally the total # of each PUID in the shipment
        row = master_spreadsheet.puid_ws.max_row+1
        master_spreadsheet.puid_ws.cell(row=row, column=1, value='Totals:')
        
        #loop through sheet and sum each column
        #Parameters for iter_cols: min_col=None, max_col=None, min_row=None, max_row=None
        for col in fws.iter_cols(2, master_spreadsheet.puid_ws.max_column, 2, master_spreadsheet.puid_ws.max_row):
            count = 0
            for c in col:
                if not c.value is None:
                    count += c.value
                    colno = c.column
            master_spreadsheet.puid_ws.cell(row=row, column=colno, value=count)
        
        #save master_spreadsheet; add a copy to SDA
        master_spreadsheet.wb.save(master_spreadsheet.spreadsheet)
        shutil.copy(master_spreadsheet.spreadsheet, self.controller.bdpl_archiver_general_dir)
        
        #copy shipment spreadsheet to 'completed shipments' in archiver_dir and unit_home
        shutil.copy(shipment_spreadsheet.spreadsheet, self.controller.bdpl_archiver_completed_spreadsheets)
        shutil.copy(shipment_spreadsheet.spreadsheet, self.completed_shpt_dir)
        
        print('\nPackaging for shipment {}{} completed!!'.format(current_item.unit_name, current_item.shipment_date))
    
    def write_db(self, db, identifier, message=None):
    
        if db in self.db_dicts:
            self.status_db[db][identifier] = message
        
        elif db in self.db_lists:
            self.status_db[db].append(identifier)
            
        self.status_db.sync()
        
    def separate_content(self, current_item, files_to_be_separated, shipment_spreadsheet):
        
        #create timestamp for premis
        timestamp = str(datetime.datetime.now())
        event_detail = ''
        
        #open log file
        outfile = open(current_item.separations_log, 'a')
        
        #if first time through, set up a separations dict for barcode; also write header to log file
        if not self.status_db['separation-stats'].get(current_item.identifier):
            
            self.status_db['separation-stats'][current_item.identifier] = {'files' : [], 'sep_file_count' : 0, 'sep_size_tally' : 0, 'sep_disk_image_count' : 0, 'separated_puids' : [], 'failed_items' : []}
            
            outfile.write('{}\t{}\t{}\t{}\n'.format('filename', 'type', 'size', 'last modified date'))
        
        #set up variable to help track success of operation
        success = True
        
        for file in files_to_be_separated:
            
            #check if we've already separated file; if so, continue
            if file in self.status_db['separation-stats'][current_item.identifier]['files']:
                continue

            print('\n\tSeparating: {}'.format(file))
            
            #set destination; create folder if it doesn't exist
            sep_destination = os.path.join(self.deaccession_dir, os.path.dirname(file))
            
            if not os.path.exists(sep_destination):
                os.makedirs(sep_destination)
            
            #check if we have a folder (in case an entire folder is being removed).  If a file, get statistics; different procedure if 'disk image' vs. 'extracted file'
            if os.path.isfile(file):
                is_file = True
                
                if 'disk-image' in file:
                    type = 'disk-image'
                    size = os.path.getsize(file)
                    last_mod_date = datetime.datetime.fromtimestamp(os.path.getmtime(file)).isoformat()
                    puid = ''
                
                else:
                    type = 'extracted-file'
                
                    with open(current_item.sf_file, 'r', encoding=utf8) as f:
                        csvreader = csv.reader(f)
                        for row in csvreader:
                            if file in row[0]:
                                size = row[1]
                                last_mod_date = row[2]
                                puid = row[5]
                                break
            
            #now move item; we've had some permission issues in the past--try to catch those
            try:
                shutil.move(file, sep_destination)
                result = 'Moved'
                
                #record item in list of completed files; add info to cumulative stats
                
                self.status_db['separation-stats'][current_item.identifier]['files'].append(file)
                
                if is_file and type == 'extracted-file':
                    self.status_db['separation-stats'][current_item.identifier]['sep_file_count'] += 1
                    
                    self.status_db['separation-stats'][current_item.identifier]['sep_size_tally'] += size
                    
                    self.status_db['separation-stats'][current_item.identifier]['separated_puids'].append(puid)
                    
                elif is_file and type == 'disk-image':
                    self.status_db['separation-stats'][current_item.identifier]['sep_disk_image_count'] += 1
                    
                    try:
                        os.rmdir(current_item.image_dir)
                    except OSError:
                        pass
                        
                #write info to our separations log
                outfile.write('{}\t{}\t{}\t{}\n'.format(file, type, size, last_mod_date))
                
                self.status_db.sync()
                
                        
            except (shutil.Error, OSError, IOError, PermissionError) as e:
                result = e
                success = False    
                self.status_db['separation-stats'][current_item.identifier]['failed_items'].append('{}\t{}'.format(file, e))
        
        #if any files failed to be moved, fail this item; return and then continue to next barcode
        if not success:
            print('\n\tWARNING: error(s) with separations; moving on to next item...')
            self.write_db('failed', current_item.identifier, 'separations\t{}'.format(' | '.join(self.status_db['separation-stats'][current_item.identifier]['failed_items'])))
            
            #close log and sync status_db
            outfile.close()
            self.status_db.sync()
            
            return success
    
        #otherwise, print results and update barcode stats/shipment spreadsheet
        print('\n\tSeparations completed:')
        
        if self.status_db['separation-stats'][current_item.identifier]['sep_file_count'] > 0:
            
            print('\t\t{} files separated ({} bytes)'.format(self.status_db['separation-stats'][current_item.identifier]['sep_file_count'], self.status_db['separation-stats'][current_item.identifier]['sep_size_tally']))
        
            event_detail = event_detail + 'removed {} files ({} bytes); '.format(self.status_db['separation-stats'][current_item.identifier]['sep_file_count'], self.status_db['separation-stats'][current_item.identifier]['sep_size_tally'])
        
        if self.status_db['separation-stats'][current_item.identifier]['sep_disk_image_count'] > 0:
            
            print('\t\t{} disk image(s) separated'.format(self.status_db['separation-stats'][current_item.identifier]['sep_disk_image_count']))
            
            event_detail = event_detail + 'removed {} disk image(s); '.format(self.status_db['separation-stats'][current_item.identifier]['sep_disk_image_count'])
        
        #update our puid count for barcode
        temp_puid_dict = dict(Counter(self.status_db['separation-stats'][current_item.identifier]['separated_puids']))
        
        for puid, count in temp_puid_dict.items():
            self.status_db['format_report'][current_item.identifier][puid]['count'] -= count
                
        #update shipment spreadsheet: size and file count
        
        #old spreadsheets may not have calculated an extent in bytes; need to catch those outliers (***Can probably remove at some point***)
        if current_item.current_dict['extent_raw'] is None:
            current_item.current_dict['extent_raw'] = current_item.get_size(current_item.files_dir) 
        else:
            current_item.current_dict['extent_raw'] -= self.status_db['separation-stats'][current_item.identifier]['sep_size_tally']
    
        current_item.current_dict['item_file_count'] -= self.status_db['separation-stats'][current_item.identifier]['sep_file_count']
        
        #write info to spreadsheet
        shipment_spreadsheet.write_to_spreadsheet(current_item.current_dict)
        
        #record premis information
        event_type = 'deaccession'
        event_outcome = 0
        event_detail = event_detail + 'see "./logs/separations.txt" for list of deaccessioned content.'
        event_detail_note = 'Formal removal of an object from the inventory of a repository.'
        agent_id = 'SdaBatchDeposit.separate_content()'
        
        current_item.record_premis(timestamp, event_type, event_outcome, event_detail, event_detail_note, agent_id)
        
        #write premis information to file
        current_item.print_premis()
        
        #close log and sync status_db
        outfile.close()
        self.status_db.sync()
        
        return success
    
    def tryint(self, s):
        try:
            return int(s)
        except ValueError:
            return s
        
    def alphanum_key(self, s):
        """ Turn a string into a list of string and number chunks.
            "z23a" -> ["z", 23, "a"]
        """
        return [ tryint(c) for c in re.split('([0-9]+)', s) ]

    def sort_puids(self, l):
        """ Sort the given list in the way that humans expect.  This and associated functions provided by https://nedbatchelder.com/blog/200712/human_sorting.html
        """
        l.sort(key=alphanum_key)