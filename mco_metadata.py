import openpyxl
import glob
import os
import sys
import shutil

from bdpl_ingest import *

def output_columns(ws):
    
    ws_headers = {}
    header_count = 0
    
    for header in ws[2]:
        if header.value == 'File':
            header_count += 1
            header_name = '%s_%s' % (header.value, header_count)
            ws_headers[header_name] = header.column
        elif header.value == 'Label':
            header_name = '%s_%s' % (header.value, header_count)
            ws_headers[header_name] = header.column
        else:
            ws_headers[header.value] = header.column
    
    return ws_headers

def main():

    while True:
        unit_name = input('\nEnter unit abbreviation: ')
        
        shipmentDate = input('\nEnter shipment date: ')
        
        ship_dir = os.path.join('Z:\\', unit_name, 'ingest', shipmentDate)
        
        avalon_dropbox = input('\nEnter path to Avalon dropbox: ')
        
        if os.path.exists(ship_dir) and os.path.exists(avalon_dropbox):
            break
    
    source_spread = glob.glob(os.path.join(ship_dir, '%s_%s.xlsx' % (unit_name, shipmentDate)))[0]
    
    if not os.path.exists(source_spread):
        print('\nWARNING: shipment spreadsheet not found! Closing...')
        sys.exit(1)
    
    while True:
        out_spread = input('Output spreadsheet: ')
        
        if not os.path.exists(source_spread):
            print('\nWARNING: MCO template not found! Enter a valid path...')
            continue
        else:
            break
    
    wb_source = openpyxl.load_workbook(source_spread)
    wb_out = openpyxl.load_workbook(out_spread)
    
    ws_source = wb_source['Appraisal']
    ws_out = wb_out['Sheet1']
    
    #get our columns four source spreadsheet; adjust so we can use them with iter_rows()
    ws_source_cols = get_spreadsheet_columns(ws_source)
    for key in ws_source_cols.keys():
        ws_source_cols[key] = ws_source_cols[key] - 1
    
    #iterate over input spreadsheet
    ws_iterrows = ws_source.iter_rows()
    next(ws_iterrows)
    
    for data in ws_iterrows:
        
        #get our columns
        ws_out_cols = output_columns(ws_out)
        
        #first, collect data for each item in a dictionary
        metadata_dict = {}
        
        for key in ws_source_cols.keys():
            metadata_dict[key] = data[ws_source_cols[key]].value
            
        #make sure this content will be going to SDA and MCO
        if not metadata_dict['initial_appraisal'] is None:
            if not all(x in metadata_dict['initial_appraisal'] for x in ['SDA', 'MCO']):
                continue
            
        print('\nWorking on:', metadata_dict['item_barcode'])
        
        #continue if barcode dir doesn't exist
        target = os.path.join(ship_dir, metadata_dict['item_barcode'])
        if not os.path.exists(target):
            print('\nBarcode folder does not exist; moving on...')
            continue
            
        #get row in outfile
        status, ws_out_newrow = return_spreadsheet_row(ws_out, metadata_dict['item_barcode'])
        
        #write values; 
        #Other Identifer & Type (will use item_barcode)
        ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Other Identifier Type"], value="local")
        ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Other Identifier"], value=metadata_dict['item_barcode'])
        
        #Title = item title; if it doesn't exist, use label transcription
        item_title = metadata_dict.get('item_title', metadata_dict['label_transcription'])
        ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Title"], value=item_title)
        
        #Creator: use collection_creator.  Alternately: put this in 'contributor' field
        ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Creator"], value=metadata_dict['collection_creator'])
        
        #Date Issued: use date range information.  Alternatively, should we use the extracted begin/end dates for 'Date Created' and instead employ an 'assigned' date?  
        if metadata_dict['begin_date'] == metadata_dict['end_date']:
            ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Date Issued"], value=metadata_dict['begin_date'])
        else:
            ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Date Issued"], value="%s/%s" % (metadata_dict['begin_date'], metadata_dict['end_date']))
        
        #Abstract: use item_description, if present 
        if 'item_description' in metadata_dict:
            ws_out.cell(row=ws_out_newrow, column=ws_out_cols["Abstract"], value=metadata_dict['item_description'])
        
        #Include Physical Description? Would be 'optical disc'--however, this could get a little messy, depending on the type pf transfer.
        
        #any other fields?
        
        #now get files; assume that they are directly within files_dir (i.e., no directories)
        '''
        Can or should we deposit .CUE file with .WAV?
        For listing files; should we include / or \ in path?
        '''
        
        files_dir = os.path.join(target, 'files')
        
        file_list = sorted(os.listdir(files_dir))
        
        if len(file_list) < 1:
            print('\WARNING: no files to be placed in MCO!')
            continue
        else:
            #compare our # of files to the 'File' and 'Label' columns in output spreadsheet
            file_count = len(file_list)
            header_count = len([x for x in ws_out_cols.keys() if 'File' in x])
            
            print('file count:', file_count)
            print('header_count:', header_count)
            
            #if file count greater than our current header count, add additional columns
            if file_count > header_count:
                
                #get the difference; we'll add twice this many columns (i.e., we'll add one 'File' and one 'Label' column per extra file)
                diff = file_count - header_count
                
                #get the current # for our file/label headers; we'll resume numbering from this point
                current_no = int(max([x for x in ws_out_cols.keys() if 'File' in x]).split('_')[1])
                
                #get the index for 'Offset'--this is where we'll insert extra columns
                current_pos = ws_out_cols['Offset']
                
                #add new columns
                ws_out.insert_cols(current_pos, diff*2)
                
                #now add headers to new columns
                for i in range(0, diff):
                    current_no += 1
                    
                    ws_out.cell(row=2, column=current_pos, value='File')
                    current_pos += 1
                    
                    ws_out.cell(row=2, column=current_pos, value='Label')
                    current_pos += 1
                
              
                
                #get a fresh dict of column indices
                ws_out_cols = output_columns(ws_out)
            
            #now loop through and add our files to spreadsheet & copy to avalon_dropbox
            for i in range(0, len(file_list)):
            
                '''Transcoding options:
                ffmpeg -i [input] -c:a aac -b:a 128k -ar 44100 [output].medium.mp4
                ffmpeg -i [input] -c:a aac -b:a 320k -ar 44100 [output].medium.mp4
                
                '''
                
                ind = i+1
                
                ws_out.cell(row=ws_out_newrow, column=ws_out_cols["File_%s" % ind], value='contents/%s' % file_list[i])
                
                file = os.path.join(files_dir, file_list[i])
                
                shutil.copy(file, avalon_dropbox)
                
            wb_out.save(out_spread)
                              
if __name__ == '__main__':
    main()               
        
        
        
        
        
            