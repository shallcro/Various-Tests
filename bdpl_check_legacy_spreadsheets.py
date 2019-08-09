'''script to check if barcodes in legacy BDPL spreadsheets from 2015-2018 have already been deposited to the SDA'''

import openpyxl
import os
import datetime

def main():
    sda = input('Enter Python-appropriate path to XLSX export of SDA stats for general/mediaimages: ')
    master = input('Enter Python-appropriate path to XLSX copy of BDPL master spreadsheet: ')
    current = input('Enter Python-appropriate path to XLSX copy of current transfer log: ')

    unit = input('Unit name: ')
    item_unit = unit.split(' (')[0]

    sda_wb = openpyxl.load_workbook(sda)
    sda_ws = sda_wb['Sheet1']
    
    master_wb = openpyxl.load_workbook(master)
    master_ws = master_wb['Item']
    
    current_wb = openpyxl.load_workbook(current)
    current_ws = current_wb['2018']
    
    #get a list of all files already in master_ws
    master_iter = master_ws.iter_rows()
    next(master_iter)
    master_list = []
    for row in master_iter:
        if not row[17].value is None:
            master_list.append(row[17].value)
    
    #get all sda info in a dictionary, including barcode-filename, size, and date deposited
    sda_iter = sda_ws.iter_rows()
    sda_dict = {}
    
    for row in sda_iter:
        if not row[4] is None:
            sda_dict[row[4].value] = {'size' : row[3].value, 'date' : row[1].value.strftime('%Y-%m-%d')}
        
    
    current_iter = current_ws.iter_rows()
    next(current_iter)
    
    total_sips = { 'number' : 0, 'size' : 0, 'dates' : [], 'extracted_file_count' : 0}
    
    already_recorded = []
    
    #now loop through the current working spreadsheet and see if any barcodes have already been added to the SDA.
    #NOTE: index positions for data will need to be changed due to irregular data-entry practices with legacy spreadsheets.
    for row in current_iter:
        barcode_value = row[0].value
        
        #may need to adjust column used for BDPL accession number
        if not barcode_value is None:
            
            match = [k for k in sda_dict.keys() if str(barcode_value) in k]
        
            if len(match) == 1:
            
                if match[0] in master_list:
                    print('\nNOTE: %s already recorded in master spreadsheet' % match[0])
                    already_recorded.append(match[0])
                    continue
                else:
                    
                    #hint F = row[5]
                    if not row[4].value is None:
                        collection_title = str(row[4].value)
                    else:
                        collection_title = ''
                        
                    if not row[7].value is None:
                        source_type = str(row[7].value)
                    else:
                        source_type = ''
                        
                    if not row[6].value is None:
                        description = str(row[6].value)
                    else:
                        description = ''
                
                    #may need to adjust which row has SDA checksum. Hint: Z is row[25]
                    if not row[25].value is None:
                        checksum = row[25].value
                    else:
                        if not row[23].value is None:
                            checksum = row[23].value
                        else:
                            checksum = 'not recorded'
                    
                    #likely won't have # of extracted files
                    extracted_file_count = ''
                    #extracted_file_count = row[27].value
                    
                    row_list = [barcode_value, item_unit, sda_dict[match[0]]['date'].replace('-', ''), collection_title, '', '', source_type, description, '', '', '', '', sda_dict[match[0]]['date'], extracted_file_count, '', sda_dict[match[0]]['size'], checksum, match[0]]
                    
                    #write to item
                    master_ws.append(row_list)
                    master_wb.save(master)
                    
                    print('\n\nAdding %s to spreadsheet' % barcode_value)
                    
                    total_sips['number'] += 1
                    total_sips['size'] += sda_dict[match[0]]['size']
                    total_sips['dates'].append(sda_dict[match[0]]['date'].replace('-', ''))
                    
                    #include this if we have extracted file count; otherwise, comment out
                    if isinstance(extracted_file_count, int):
                        total_sips['extracted_file_count'] += extracted_file_count
                
            elif len(match) == 0:
                print('\nItem not found on SDA spreadsheet. Check list.')
            
            else:
                print('Found: ', ', '.join(match))

    #now write cumulative
    #find date range
    latest_date = max(total_sips['dates'])
    earliest_date = min(total_sips['dates'])
    
    tdelta = datetime.datetime.strptime(latest_date, '%Y%m%d') - datetime.datetime.strptime(earliest_date, '%Y%m%d')
        
    #use 1 day as minimum timedelta
    if tdelta < datetime.timedelta(days=1):
        duration = 1
    else:
        duration = int(str(tdelta).split()[0])
        
    master_cumulative = master_wb['Cumulative']
    
    if earliest_date == latest_date:
        dates = earliest_date
    else:
        dates = '%s-%s' % (earliest_date, latest_date)
    
    if total_sips['extracted_file_count'] == 0:
        total_sips['extracted_file_count'] = ''
    
    rowlist = [ unit, 'legacy: %s' % dates, total_sips['number'], total_sips['extracted_file_count'], '', total_sips['size'], earliest_date, latest_date, duration ]
    
    master_cumulative.append(rowlist)
    master_wb.save(master)
    
    if len(already_recorded) > 0:
        print('\nThe following items were on %s but were previously added to master spreadsheet:\n\t', '\n\t'.join(already_recorded))
    
if __name__ == '__main__':
    main()