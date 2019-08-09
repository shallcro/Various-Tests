'''
This script created to test process to gather statistics about files separated during BDPL bag preparation and SDA deposit process.
'''

import subprocess
import os
import shutil
import bagit
import sys
import openpyxl
import hashlib
import datetime
import pickle
import glob
import csv
from lxml import etree
import uuid

def separate_content(sep_dest, file, log, report_dir, barcode):
    print('working on: ', file)
    
    #create folder structure, if needed
    if not os.path.exists(sep_dest):
        os.makedirs(sep_dest)
    
    #get separation stats
    try:
        size = os.path.getsize(file)
    except OSError:
        pass
        
    #We will store results for this file in a temp file for the whole barcode.
    temp_stat_list = os.path.join(report_dir, '%s-separation-stats.txt' % barcode)
    temp_list = []
    
    #if this temp file exists, retrieve it and check if this file has already been moved.
    if os.path.exists(temp_stat_list):
        with open(temp_stat_list, 'rb') as f:
            temp_list = pickle.load(f)    
        
        #if we already have a list of 
        for f in temp_list:
            if f['file'] == file:
                #if file previously failed, remove it from the list.
                if f['result'] != 'Moved':
                    temp_list.remove(f)
                    break
                #if the file was moved, return to main() and go on to next one.
                else:
                    return
    
    try:
        print(file, sep_dest)
        with open(log, 'a') as f:
            f.write('%s\n' % file)
        result = 'Moved'
            
    except (shutil.Error, OSError, IOError) as e:
        print('\n\tError separating %s: %s' % (file, e))
        result = e

    #add file to our temp list, noting result of operation and file size.  Write temp list to file.
    temp_list.append({'file' : file, 'size' : size, 'result' : result})
    
    with open(temp_stat_list, 'wb') as f:
        pickle.dump(temp_list, f)

def raw_text(text):
    """Returns a raw string representation of text"""
    
    escape_dict={'\a':r'\a',
           '\b':r'\b',
           '\c':r'\c',
           '\f':r'\f',
           '\n':r'\n',
           '\r':r'\r',
           '\t':r'\t',
           '\v':r'\v',
           '\'':r'\'',
           '\"':r'\"',
           '\0':r'\0',
           '\1':r'\1',
           '\2':r'\2',
           '\3':r'\3',
           '\4':r'\4',
           '\5':r'\5',
           '\6':r'\6',
           '\7':r'\7',
           '\8':r'\8',
           '\9':r'\9'}
    
    new_string=''
    for char in text:
        try: 
            new_string+=escape_dict[char]
        except KeyError: 
            new_string+=char
    return new_string

def list_write(list_name, barcode, message=None):
    with open(list_name, 'a') as current_list:
        if message is None:
            current_list.write('%s\n' % barcode)
        else:
            current_list.write('%s\t%s\n' % (barcode, message))

def get_size(start_path):
    total_size = 0
    if os.path.isfile(start_path):
        total_size = os.path.getsize(start_path)
    else:
        for dirpath, dirnames, filenames in os.walk(start_path):
            for f in filenames:
                fp = os.path.join(dirpath, f)
                # skip if it is symbolic link
                if not os.path.islink(fp):
                    total_size += os.path.getsize(fp)
    return total_size

def main():    
    separations_manifest = "C:/BDPL/TEST34/separations.txt" 
    source = "C:/BDPL/TEST34/"
    shipmentID = 'TEST34'
    report_dir ="C:/BDPL/TEST34/reports" 
    separated_list = os.path.join(report_dir, 'separated-content.txt')
    stats_doc = os.path.join(report_dir, 'shipment_stats.txt')

    spreadsheet = "C:/BDPL/TEST34/UAC_2017_TEST - Copy.xlsx" 
    wb = openpyxl.load_workbook(spreadsheet)
    ws = wb['Appraisal']
    iterrows = ws.iter_rows()
    next(iterrows)
    
    for row in iterrows:

        barcode = str(row[0].value)
        target = os.path.join(source, barcode)
        
        if os.path.isfile(separations_manifest):
            #set up a log file
            separations_log = os.path.join(report_dir, 'separations.txt')
            
            #get a list of relevant lines from the separations manifest, splitting at the barcode (to avoid any differences with absolute paths)
            to_be_separated = []
            with open(separations_manifest, 'r') as f:
                sep_list = f.read().splitlines()
            for file in sep_list:
                if barcode in file:
                    name = raw_text(file.replace('"', '').rstrip())
                    to_be_separated.append(name.split('%s\\' % shipmentID, 1)[1])
            
            #if we've found any files, create a barcode folder in 'deaccessions' direcetory and then loop through list
            if len(to_be_separated) > 0:
            
                separations_dir = os.path.join('deaccessioned', barcode)
                if not os.path.exists(separations_dir):
                    os.mkdir(separations_dir)
                
                for item in to_be_separated:
                    #if a wildcard is used, we will use glob to build a list of all files/folders matching pattern
                    if '*' in item:
                        wildcard_list = []
                        
                        #recursive option
                        if '\\**' in item:
                            wildcard_list = glob.glob(item, recursive=True)
                    
                        #wildcard at one level
                        elif '\\*' in item:
                            wildcard_list = glob.glob(item)
                        
                        #now loop through this list of files/folders identified by glob
                        for wc in wildcard_list:
                            sep_dest = os.path.join(separations_dir, os.path.dirname(wc))
                            separate_content(sep_dest, wc, separations_log, report_dir, barcode)
                    
                    else:
                        sep_dest = os.path.join(separations_dir, os.path.dirname(item))
                        separate_content(sep_dest, item, separations_log, report_dir, barcode)                        
        
                #compile stats and check to see if any errors reported
                temp_stat_list = os.path.join(report_dir, '%s-separation-stats.txt' % barcode)
                temp_list = []
                if os.path.exists(temp_stat_list):
                    with open(temp_stat_list, 'rb') as f:
                        temp_list = pickle.load(f)
                    
                    #loop though list of stats; add to file and byte counts and note any failures
                    sep_files = 0
                    sep_size = 0
                    success = True
                    for f in temp_list:
                        if f['result'] != 'Moved':
                            success = False
                            pass
                        else:
                            sep_files += 1
                            sep_size += f['size']
                    sep_size = 0
                    sep_files = 0
                            
                    #if we have any failures, this barcode will fail: we need to make sure any files designated for separation have been removed.        
                    if not success:
                        print('\tOne or more errors with separations.')
                        list_write(failed_list, '%s\tSee %s for details.' % (barcode, temp_stat_list))
                        continue
                    else:
                        print('\tSeparations completed.')
                        list_write(separated_list, '%s', '%s\t%s' % (barcode, sep_files, sep_size))
                        #os.remove(temp_stat_list)
                    
            #check to see if files were separated; if so, get stats and adjust previous totals.
            if os.path.exists(separated_list):
                with open(separated_list, 'r') as fi:
                    fi = csv.reader(fi, delimiter='\t')
                    for line in fi:
                        try:
                            if barcode == line[0]:
                                sep_files = int(line[1])
                                sep_size = int(line[2])
                                break
                        except IndexError:
                            continue
                                
                        
            
            #Recalculate size of extracted files
            if row[16].value is None:
                extracted_size = get_size(os.path.join(target, 'data', 'files')) 
            else:
                extracted_size = (int(row[16].value) - sep_size)
                print('original: %s\nseparation: %s\nnew: %s\n\n' % (row[16].value, sep_size, extracted_size))
            
            #write corrected size back to spreadsheet
            #ws.cell(row=row[0].row, column=17, value=extracted_size)
            
            #if there are any files separated, adjust extracted file count
            try:
                if sep_files > 0:
                    extracted_no = (int(row[17].value) - sep_files)
                    print('original: %s\nseparation: %s\nnew: %s\n\n' % (row[17].value, sep_files, extracted_no))
                #ws.cell(row=row[0].row, column=18, value=extracted_no)
            except:
                print('whoops')
            # #in case previous step failed, may need to retrieve SIP info
            # SIP_stats = os.path.join(report_dir, 'SIP_%s.txt' % barcode)
            # try:
                # SIP_dict
            # except NameError:
                # SIP_dict = {}
                # with open(SIP_stats, 'rb') as file:
                    # SIP_dict = pickle.load(file)
                
            # rowlist = [barcode, unit, shipmentID, str(row[2].value), str(row[3].value), str(row[4].value), str(row[6].value), str(row[7].value), str(row[8].value), str(row[9].value), str(row[10].value), str(row[12].value), str(datetime.datetime.now()), extracted_no, extracted_size, SIP_dict['size'], SIP_dict['md5']]
            
            # item_ws.append(rowlist)        
            
            #add info to stats for shipment
            shipment_stats = {}
           
            #if we already have shipment stats, retrieve from file and update.  Otherwise, create dictionary values
            if os.path.exists(stats_doc):
                with open(stats_doc, 'rb') as file:
                    shipment_stats = pickle.load(file)
                #shipment_stats['sip_count'] += 1
                shipment_stats['extracted_no'] += extracted_no
                shipment_stats['extracted_size'] += extracted_size
                #shipment_stats['SIP_size'] += SIP_dict['size']
            else:
                shipment_stats = {'extracted_no' : extracted_no, 'extracted_size' : extracted_size}
            
            #Write shipment stat back to file
            with open(stats_doc, 'wb') as file:
                pickle.dump(shipment_stats, file)
            
            #list_write(metadata_list, barcode)
            

    if os.path.exists(stats_doc):
        shipment_stats = {}
        with open(stats_doc, 'rb') as file:
            shipment_stats = pickle.load(file)
        print('total #: ' % shipment_stats['extracted_no'])
        print('total size: ' % shipment_stats['extracted_size'])

if __name__ == '__main__':
    
    main()