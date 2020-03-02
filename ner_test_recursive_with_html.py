import spacy
import os
from lxml import etree
from itertools import islice
import statistics
from collections import Counter
from tika import parser
from sys import argv
import tika
tika.TikaClientOnly = True
import gensim
import gensim.corpora as corpora
from gensim.utils import simple_preprocess
from gensim.models import CoherenceModel
from gensim.models import ldamodel
from spacy.lemmatizer import Lemmatizer
from pprint import pprint
import glob
import openpyxl
import matplotlib.pyplot as plt
import numpy as np
import textwrap
import chardet
import shutil
import sys


def take(n, iterable):
    "Return first n items of the iterable as a list"
    return list(islice(iterable, n))
    
def return_pairs(mydict):
    return {k: mydict[k] for k in list(mydict)[:2]}

def remove_stopwords(doc):
    # This will remove stopwords and punctuation.
    # Use token.text to return strings, which we'll need for Gensim.
    doc = [token.text for token in doc if token.is_stop != True and token.is_punct != True]
    return doc
    
def lemmatizer(doc):
    # This takes in a doc of tokens from the NER and lemmatizes them. 
    # Pronouns (like "I" and "you" get lemmatized to '-PRON-', so I'm removing those.
    nlp = spacy.load("en_core_web_sm")
    doc = [token.lemma_ for token in doc if token.lemma_ != '-PRON-']
    doc = u' '.join(doc)
    return nlp.make_doc(doc)

def absolute_value(pct, allvals):
    absolute = int(pct/100.*np.sum(allvals))
    return "{:d} hits".format(absolute)

def entity_separator(doc_text, nlp, person, norp, fac, org, gpe, loc, product, event, doc_list):
    
    doc = nlp(doc_text)
    
    for ent in doc.ents:
        if ent.label_ == 'PERSON':
            person.append(ent.text) 
        elif ent.label_ == 'NORP':
            norp.append(ent.text)
        elif ent.label_ == 'FAC':
            fac.append(ent.text)
        elif ent.label_ == 'ORG':
            org.append(ent.text)
        elif ent.label_ == 'GPE':
            gpe.append(ent.text)
        elif ent.label_ == 'LOC':
            loc.append(ent.text)
        elif ent.label_ == 'PRODUCT':
            product.append(ent.text)
        elif ent.label_ == 'EVENT':
            event.append(ent.text)
        else:
            continue
            
    #remove stopwords and add to our document list for topic modeling
    lemmatized = lemmatizer(doc)
    no_stops = remove_stopwords(lemmatized)
    doc_list.append(no_stops)
    
    return person, norp, fac, org, gpe, loc, product, event, doc_list

def main():
    #load English model. Small provides good enough NER while being faster than en_core_web_md
    nlp = spacy.load("en_core_web_sm")
        
    #set variables
    try:
        ship_dir = argv[1]
    except IndexError:
        print('\n\nWarning: missing "shipment" argument.  Include a python-friendly path to the directory you want to run NER/topic modeling on.')
        sys.exit()
    
    shipmentID = os.path.basename(ship_dir)
    
    output_folder = os.path.join('C:/temp', 'ner-testing', shipmentID)
    output_files = os.path.join(output_folder, 'files')
    output = os.path.join(output_folder, '{}_ner-report.html'.format(shipmentID))

    css_file = 'C:/BDPL/resources/NER/mystyle.css'
    
    if not os.path.exists(output_files):
        os.makedirs(output_files) 
    
    
    #get spreadsheet and set openpyxl variables    
    spreadsheet = glob.glob(os.path.join(ship_dir, '*.xlsx'))[0]
    wb = openpyxl.load_workbook(spreadsheet)
    ws = wb['Appraisal']
    
    #start html doc
    html = etree.Element('html')
    head = etree.SubElement(html, 'head')
    script = etree.SubElement(head, 'script', src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js")
    link = etree.SubElement(head, 'link')
    link.attrib['rel'] = "stylesheet"
    link.attrib['type'] = "text/css"
    link.attrib['href'] = "./files/mystyle.css" 
    body = etree.SubElement(html, 'body')
    h1 = etree.SubElement(body, 'h1')
    h1.text ="Content Analysis"
    
    div = etree.SubElement(body, 'div')
    div.attrib['class'] = 'scrollable'
    
    table = etree.SubElement(div, 'table')
    tr = etree.SubElement(table, 'tr')
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Object'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Creator'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Label Transcription'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Initial Appraisal'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: People'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Nationalities or religious/political groups'
   
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Buildings and Facilities'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Organizations'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Countries, cities, states'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Geographic Features/Locations'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Products'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Named Entities: Events'
    
    th = etree.SubElement(tr, 'th')
    th.text = 'Topic Modeling'
    
    for item_barcode in [d for d in os.listdir(ship_dir) if os.path.isdir(os.path.join(ship_dir, d))]:
        # if not '30000152027102' in item_barcode:
            # continue
        print('\nWorking on: ', item_barcode)
        
        found = False
        for cell in ws['A']:
            if (cell.value is not None):
                if item_barcode == str(cell.value).strip():
                    found = True
                    creator = ws.cell(row=cell.row, column=5).value
                    transcription = ws.cell(row=cell.row, column=8).value
                    appraisal = ws.cell(row=cell.row, column=9).value
        if not found:
            creator = 'N/A'
            transcription = 'N/A'
            appraisal = 'N/A'
        
        #add barcode to html
        tr = etree.SubElement(table, 'tr')
        td = etree.SubElement(tr, 'td')
        td.text = item_barcode
        td = etree.SubElement(tr, 'td') #creator
        td.text = creator
        td = etree.SubElement(tr, 'td') #label transcription
        td.text = transcription
        td = etree.SubElement(tr, 'td') #initial appraisal notes
        td.text = appraisal
        
        files_dir = os.path.join(ship_dir, item_barcode, 'files')
        
        if not os.path.exists(files_dir):
            continue
        
        #create lists to store extracted entities and text for gensim
        doc_list = []
        person = []
        norp = []
        fac = []
        org = []
        gpe = []
        loc = []
        product = []
        event = []
        
        #Loop through files in our item file_dir. 
        for root, dirs, files, in os.walk(files_dir):
            for f in files:
                
                ner_target = os.path.join(root, f)
                
                print('\tProcessing: {}\n'.format(ner_target))
                
                #use python-tika 'parser' to pull text from file; skip if we get UnicodeEncodeError (will try to figure out how to handle those later)
                try:
                    content = parser.from_file(ner_target)
                except UnicodeEncodeError:
                    continue
                
                #isolate the 'content' from python-tika parser output
                if 'content' in content:
                    text = content['content']
                else:
                    continue
                
                #to make our results a little more manageable, 'split' the content to remove newlines, tabs, etc.  Then rejoin everything (separating each word by a space)
                text = str(text).split()
                combined_text = ' '.join(t for t in text)
                
                #We will now send our text to the 'entity_separator' function--created so that we could handle all of our text in the same way, as we discovered that Spacy will run out of memory if the text is too big.  To handle that, we check size; if text is too big, we will just process one chunk of it at a time and concatenate the results.
                if len(combined_text) > 1000000:
                    continue
                    for chunk in textwrap.wrap(combined_text, 900000):
                        person, norp, fac, org, gpe, loc, product, event, doc_list = entity_separator(chunk, nlp, person, norp, fac, org, gpe, loc, product, event, doc_list)
                else:
                    person, norp, fac, org, gpe, loc, product, event, doc_list = entity_separator(combined_text, nlp, person, norp, fac, org, gpe, loc, product, event, doc_list)
        
        #now loop through each list of entities we created
        for ls in [person, norp, fac, org, gpe, loc, product, event]:
           
            #tally the number of unique entities in each list and sort the resulting dictionary so we can present results in descending order, with the most frequent first.  Reconciling near matches or eliminating false positives would require too much human intervention
            tally = dict(Counter(ls))
            
            #sort the resulting dictionary so we can present results in descending order, with the most frequent first
            sorted_tally = {k: v for k, v in sorted(tally.items(), key=lambda item: item[1], reverse=True)}
            
            #if no results (i.e., empty list), we just note entity is 'N/A' and we do not create a pie-chart or full text file of entities
            if len(sorted_tally) == 0:
                td = etree.SubElement(tr, 'td')
                ul = etree.SubElement(td, 'ul')
                li = etree.SubElement(ul, 'li')
                li.text ='N/A'
            #if we do have results, we will write the top
            else:
                #set graph title
                if ls == person:
                    graph_title = 'people'
                elif ls == org:
                    graph_title = 'organizations'
                elif ls == gpe:
                    graph_title = 'countries-cities-states'  
                elif ls == norp:
                    graph_title = 'groups'
                elif ls == loc:
                    graph_title = 'geographic-features'
                elif ls == fac:
                    graph_title = 'buildings-facilities'
                elif ls == product:
                    graph_title = 'products'
                elif ls == event:
                    graph_title = 'events'
                
                current_chart = os.path.join(output_files, '{}-{}.png'.format(item_barcode, graph_title))
                current_report = os.path.join(output_files, '{}-{}.txt'.format(item_barcode, graph_title)) 
                
                #write full list of entities to file.  
                with open(current_report, 'wb') as f:
                    f.write('{} Entities for {}\n\n'.format(graph_title.upper(), item_barcode).encode())
                    
                    #The code below adds some spaces so that this text file is better formatted
                    for k in sorted_tally.keys():
                        if len(k) < 30:
                            diff = 30 - len(k)
                            spaces = len(k) + diff
                            value = k.rjust(spaces, ' ') 
                        elif len(k) > 30:
                            content = textwrap.wrap(k, 30)
                            for i, c in enumerate(content):
                                if len(c) < 30:
                                    diff = 30 - len(c)
                                    spaces = len(c) + diff
                                    content[i] = c.rjust(spaces, ' ')
                            value = '\n'.join(content)
                        else:
                            value = k
                        #if we have UnicodeEncodeError, use chardet to try and ID the encoding.
                        try:
                            f.write("{} : {}\n\n".format(value, sorted_tally[k]).encode())
                        except UnicodeEncodeError:
                            enc = chardet.detect(value.encode())['encoding']
                            f.write("{} : {}\n\n".format(value.encode(encoding = enc, errors="ignore"), str(sorted_tally[k]).encode()))
                            
                
                #create link to full list of entities; add cell to html
                td = etree.SubElement(tr, 'td')
                a = etree.SubElement(td, 'a')
                a.attrib['href'] = './{}/{}'.format(os.path.basename(os.path.dirname(current_report)), os.path.basename(current_report))
                a.attrib['target'] = "_blank"
                p = etree.SubElement(a, 'p')
                p.text = 'Full list of entities'
                
                ul = etree.SubElement(td, 'ul')
                
                #determine which results are above average; the rest are 'others'.  If there are too many and chart is unreadable' what is upper limit?
                ner_median = statistics.median(sorted(set(list(sorted_tally.values()))))
                
                #we are only going to include results in the 90th percentile and above on our html file.  Change variable as needed
                percentile = 90
                
                ner_percentile = np.percentile(sorted(set(list(sorted_tally.values()))), percentile)
                
                #If we only have 10 or fewer entities, all will get included in piechart and html; if there are more, we will report the top percentiles and then lump everything else into 'other' for our pie chart and html report
                if len(sorted_tally) <= 10:
                    results = sorted_tally
                    others = {}
                else:
                    results = {k:v for k, v in sorted_tally.items() if v > ner_median}
                    others = {k:v for k, v in sorted_tally.items() if v <= ner_median}
                
                #if we have too many results, we won't be able to view results; use 'percentile' as cut-off
                if len(results) > 25:
                    results = {k:v for k, v in sorted_tally.items() if v > ner_percentile}
                    others = {k:v for k, v in sorted_tally.items() if v <= ner_percentile}
                
                #write to html in a list; for each list, include a link to pie chart
                labels = []
                hits = []
                
                a = etree.SubElement(ul, 'a')
                a.attrib['href'] = './{}/{}'.format(os.path.basename(os.path.dirname(current_chart)), os.path.basename(current_chart))
                a.attrib['target'] = "_blank"
                
                for k, v in results.items():
                    li = etree.SubElement(a, 'li')
                    li.text ='{} : {}'.format(k, sorted_tally[k])
                    
                    #add entities and # of hits to list for pie charts
                    labels.append(k)
                    hits.append(v)
                
                #if there are 'other' results, add these to list as a group
                if len(others) > 0:
                    labels.append('Other')
                    hits.append(sum(others.values()))
                
                #create pie chart; clear previous one first
                plt.clf()
                plt.title('{} Entities for {}'.format(graph_title.upper(), item_barcode))
                plt.pie(hits, labels=labels, autopct=lambda pct: absolute_value(pct, hits))
                plt.axis('equal')
                plt.savefig(current_chart, bbox_inches='tight')
    
        #now do topic modeling
        td = etree.SubElement(tr, 'td')
        ul = etree.SubElement(td, 'ul')

        # Creates, which is a mapping of word IDs to words.
        words = corpora.Dictionary(doc_list)

        # Turns each document into a bag of words.
        corpus = [words.doc2bow(doc) for doc in doc_list]
        
        #use gensim LdaModel class to produce topics; 'num_topics' defines # of topics that will be created
        try:
            lda_model = ldamodel.LdaModel(corpus=corpus, id2word=words, num_topics=6, random_state=2, update_every=1, passes=10,  alpha='auto', per_word_topics=True)
            
            x=lda_model.show_topics(num_topics=6, num_words=8,formatted=False)
            topics_words = [(tp[0], [wd[0] for wd in tp[1]]) for tp in x]
            
            #Print topic modeling output
            if len(topics_words) > 0:
                for topic,words in topics_words:
                    li = etree.SubElement(ul, 'li')
                    li.text = " ".join(words)
            else:
                li = etree.SubElement(ul, 'li')
                li.text = "N/A"
        
        except ValueError:
            li = etree.SubElement(ul, 'li')
            li.text = "N/A"
        
    
    #write html to file
    html_doc = etree.ElementTree(html)
    html_doc.write(output, method="html", pretty_print=True)      

    shutil.copy(css_file, output_files)


if __name__ == "__main__":
    main()

