'''
Script to compile cumulative master_stats
'''

import openpyxl
import os
import math
import shutil
import subprocess

def convert_size(size):
    # convert size to human-readable form
    if (size == 0):
        return '0 bytes'
    size_name = ("bytes", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
    i = int(math.floor(math.log(size,1024)))
    p = math.pow(1024,i)
    s = round(size/p)
    s = str(s)
    s = s.replace('.0', '')
    return '%s %s' % (s,size_name[i])

def main():
    if not os.path.exists('W:/spreadsheets/bdpl_master_spreadsheet.xlsx'):
        book = input('\nPath to master spreadsheet: ')
    else:
        book = 'W:/spreadsheets/bdpl_master_spreadsheet.xlsx'
    
    spreadsheet_copy = os.path.join('C:/BDPL/', '%s_COPY.xlsx' % os.path.basename(book))
    
    shutil.copy(book, spreadsheet_copy)

    wb = openpyxl.load_workbook(spreadsheet_copy)

    ws_master_all = wb['Cumulative']

    iterrows = ws_master_all.iter_rows()

    next(iterrows)

    master_stats = {}
    
    master_output = 'C:/BDPL/deposited_content_stats.txt'
    if os.path.exists(master_output):
        os.remove(master_output)

    with open(master_output, 'w') as f:
        for row in iterrows:
            unit = row[0].value.split()[0]
            if not unit in master_stats.keys():
                master_stats[unit] = {'count' : 1, 'items' : int(row[2].value), 'size' : int(row[5].value)}
            else:
                master_stats[unit]['count'] += 1
                master_stats[unit]['items'] += int(row[2].value)
                master_stats[unit]['size'] += int(row[5].value)
        
        unit_totals = {}
        for key, value in master_stats.items():
            sized = convert_size(value['size'])
            print('Unit: %s\nItems: %s\nSize: %s\n' % (key, value['items'], sized))
            unit_totals[key] = {'items' : value['items'], 'size' : sized}
            
            
        ws_master_item = wb['Item']
        
        iterrows2 = ws_master_item.iter_rows()
        next(iterrows2)
        
        stats_items = {}
        by_year = {}
        for row in iterrows2:
            unit = row[1].value
            year = str(row[14].value)[:4]
            
            if year not in by_year.keys():
                by_year[year] = [[int(row[17].value)], [1]]
            else:
                by_year[year][0].append(int(row[17].value))
                by_year[year][1].append(1)
            
            if not unit in stats_items.keys():
                stats_items[unit] = {year : {'items' : 1, 'size' : int(row[17].value)}}
            else:
                if not year in stats_items[unit].keys():
                    stats_items[unit][year] = {'items' : 1, 'size' : int(row[17].value)}
                else:
                    
                    stats_items[unit][year]['items'] += 1
                    stats_items[unit][year]['size'] += int(row[17].value)
        
        for unit, data in stats_items.items():
            f.write('%s\n' % unit)
            print(unit)
            for year, info in sorted(data.items()):
                print('\t', year)
                print('\t\tNumber of items: ', info['items'])
                sized = convert_size(info['size'])
                print('\t\tOverall size: ', sized)
                
                f.write('\t%s\n\t\tNumber of items: %s\n\t\tOverall size: %s\n' % (year, info['items'], sized))
            
            print('\tTOTAL:\n\t\tNumber of items: %s\n\t\tOverall size: %s' % (unit_totals[unit]['items'], unit_totals[unit]['size']))
            f.write('\tTOTAL:\n\t\tNumber of items: %s\n\t\tOverall size: %s\n' % (unit_totals[unit]['items'], unit_totals[unit]['size']))
        
        print('\n\n')
        f.write('\n\n')
        
        for key, values in sorted(by_year.items()):
            print(key, ':', convert_size(sum(values[0])), '(%s items)' % sum(values[1]))
            f.write('%s : %s (%s items)\n' % (key, convert_size(sum(values[0])), sum(values[1])))
        
    print('\n\nText file with these statistics located at: %s' % master_output)
    
    os.remove(spreadsheet_copy)
    
    cmd = 'notepad %s' % master_output
    subprocess.call(cmd)
    
    
    
if __name__ == '__main__':
    main()