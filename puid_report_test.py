import openpyxl
import os
import csv

os.chdir('C:/BDPL/TEST34')



book = "C:/BDPL/TEST34/UAC_2019_TEST.xlsx"

wb = openpyxl.load_workbook(book)
fws = wb.create_sheet('test5')

dirs = list(dir for dir in os.listdir('C:/BDPL/TEST34') if os.path.isdir(os.path.join('C:/BDPL/TEST34', dir)))

format_report = []

puid_list = []

for barcode in dirs:
    formats = os.path.join(barcode, 'metadata', 'reports', 'formatVersions.csv')
    
    if not os.path.exists(formats):
        continue
    
    
    temp_list = []
    with open(formats, 'r') as fi:
        fi = csv.reader(fi)
        next(fi)
        for line in fi:
            #create a dictionary for each row
            temp_dict = {}
            temp_dict['puid'] = line[1]
            temp_dict['format'] = line[0]
            temp_dict['version'] = line[2]
            temp_dict['count'] = int(line[3])
            
            #add all puids to a master list
            puid_list.append(line[1])
            
            #add temp dict to a temp list
            temp_list.append(temp_dict)
    #once we're done with a given barcode, 
    format_report.append({barcode : temp_list})

#remove any duplicate values from our puid list and add 'barcode' as a header for the first column
puid_list = list(set(puid_list)).sort()
puid_list.insert(0, 'barcode')

#add this list to our spreadsheet
fws.append(puid_list)

#now create a puid dictionary so that we can refer to columns in our 'formats' worksheet.  Add one to each enumerator value so we align with openpyxl column index
puid_dict = {}
for pu, id in enumerate(puid_list):
    puid_dict['id'] = pu+1
    
#now go through our format report; for a given puid, write the count

for index in range(len(format_report)):
    for key in format_report[index]:
        newrow = fws.max_row+1
        fws.cell(row=newrow, column=puid_dict['barcode'], value=key)
        for item in format_report[index][key]:
            fws.cell(row=fws.max_row, column=puid_dict[item['puid']], value=item['count'])
            
            
            
            # temp_list = []
            # temp_list.append([(item['count'], item['puid'])])
        # for i in temp_list:
            # if len(temp_list) > 1:
                # print('%s instance(s) of %s' % (temp_list[0], temp_list[1]))
                