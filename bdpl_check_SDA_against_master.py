'''script to determine if barcodes from legacy SDA deposits have been entered into BDPL master spreadsheet.'''

import openpyxl
import os
import datetime

def main():
    sda = input('Enter Python-appropriate path to XLSX export of SDA stats for general/mediaimages: ')
    master = input('Enter Python-appropriate path to XLSX copy of BDPL master spreadsheet: ') 
    
    sda_wb = openpyxl.load_workbook(sda)
    sda_ws = sda_wb['Sheet1']
    
    master_wb = openpyxl.load_workbook(master)
    master_ws = master_wb['Item']
    
    master_iter = master_ws.iter_rows()
    next(master_iter)
    
    master_list = []
    
    for row in master_iter:
        if not row[17].value is None:
            master_list.append(row[17].value)
    
    
    
    sda_iter = sda_ws.iter_rows()
    
    
    missing_from_master = []
    medialogs = []
    
    for row in sda_iter:
        if not row[4].value is None:
            if not str(row[4].value) in master_list:
                if '.xlsx' in str(row[4].value):
                    medialogs.append(str(row[4].value))
                else:
                    missing_from_master.append(str(row[4].value))
                
    if len(missing_from_master) == 0:
        print('All SDA content accounted for')
    else:
        print('The following SDA content is not on the master spreadsheet:\n', '\n\t'.join(missing_from_master))
        
if __name__ == '__main__':
    main()
            
    